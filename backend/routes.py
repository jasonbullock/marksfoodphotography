import csv
import copy
import time
import hashlib
import hmac
import io
import mailer
import notifier
import structure_form
import json
import os
import random
import re
import threading
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote
from xml.etree import ElementTree

import requests

from flask import Blueprint, current_app, jsonify, request, session

from airtable import airtable
from config import Config
from receiving_photo_storage import (
    ReceivingPhotoConfigError,
    ReceivingPhotoCollisionError,
    ReceivingPhotoStorage,
    ReceivingPhotoStorageError,
    ReceivingPhotoValidationError,
    sanitize_path_segment,
)
api = Blueprint("api", __name__)

_CREATIVE_FORCE_LAST_WEBHOOK = None
# A reset in Creative Force fires one event per downstream step, so a single action
# arrives as a burst. Keeping only the last one made the card's step effectively
# arbitrary and left nothing to diagnose with.
_CREATIVE_FORCE_RECENT_WEBHOOKS = []
CREATIVE_FORCE_RECENT_LIMIT = 40


def _record_creative_force_webhook(entry):
    global _CREATIVE_FORCE_LAST_WEBHOOK
    _CREATIVE_FORCE_LAST_WEBHOOK = entry
    _CREATIVE_FORCE_RECENT_WEBHOOKS.insert(0, entry)
    del _CREATIVE_FORCE_RECENT_WEBHOOKS[CREATIVE_FORCE_RECENT_LIMIT:]
    return entry

C = Config  # shorthand
UPLOAD_DIR = Path(__file__).resolve().parent / "uploads" / "receiving"
AUTH_SESSION_KEY = "marks_auth"
PUBLIC_ENDPOINTS = {
    "api.health",
    "api.airtable_status",
    "api.auth_login",
    "api.auth_me",
    "api.auth_logout",
    "api.auth_users",
    "api.receiving_photo",
    "api.creative_force_webhook",
}


def err(msg, status=400, **details):
    payload = {"error": msg, **details}
    return jsonify(payload), status


def airtable_err(error):
    response = getattr(error, "response", None)
    status = getattr(response, "status_code", 500)
    if status in {403, 404}:
        return err("Airtable table is not configured yet.", 501)
    message = "Airtable request failed."
    try:
        payload = response.json() if response is not None else {}
        airtable_message = (payload.get("error") or {}).get("message")
        if airtable_message:
            message = airtable_message
    except ValueError:
        pass
    return err(message, status)


@api.before_request
def require_authenticated_session():
    if request.method == "OPTIONS":
        return None
    if request.endpoint in PUBLIC_ENDPOINTS:
        return None
    if not _session_user():
        return err("Authentication required", 401)
    return None


def _photo_storage():
    return ReceivingPhotoStorage(C)


def _is_unknown_field_error(error, field_name):
    response = getattr(error, "response", None)
    try:
        payload = response.json() if response is not None else {}
    except ValueError:
        return False
    message = str((payload.get("error") or {}).get("message") or "")
    return "Unknown field name" in message and field_name in message


def _now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _now_utc():
    return datetime.now(timezone.utc)


def _as_list(value):
    if not value:
        return []
    return value if isinstance(value, list) else [value]


def _field_value(record, field):
    return (record.get("fields", {}) if record else {}).get(field)


def _first_link(record, field):
    values = _field_value(record, field)
    return values[0] if isinstance(values, list) and values else None


def _current_user():
    user = _session_user()
    if not user:
        return None
    return _record_from_shaped_user(user)


def _current_user_id():
    user = _session_user()
    return user.get("id") if user else None


def _current_user_display_name():
    """Who is acting, as plain text.

    "Merchandise Verified By" is a text field, not a link to Users: verification
    is a stamp on the merchandise, and linking would put a Merchandise back-link
    on every user record for no benefit.
    """
    user = _session_user()
    if not user:
        return ""
    return user.get("displayName") or user.get("name") or ""


def _permission_context():
    user = _current_user()
    if user is None:
        return {"all": False, "client_ids": set()}
    fields = user.get("fields", {})
    if fields.get(C.F_USER_ALL_CLIENTS):
        return {"all": True, "client_ids": None}
    return {"all": False, "client_ids": set(fields.get(C.F_USER_CLIENTS, []) or [])}


def _client_permitted(client_id, permissions=None):
    if not client_id:
        return False
    permissions = permissions or _permission_context()
    return permissions["all"] or client_id in permissions["client_ids"]


def _client_ids_permitted(client_ids, permissions=None):
    ids = set(_as_list(client_ids))
    if not ids:
        return False
    permissions = permissions or _permission_context()
    return permissions["all"] or bool(ids & permissions["client_ids"])


def _receipt_client_permitted(client_ids, permissions=None):
    ids = set(_as_list(client_ids))
    if not ids:
        return True
    permissions = permissions or _permission_context()
    return permissions["all"] or bool(ids & permissions["client_ids"])


def _filter_by_client_field(records, field):
    permissions = _permission_context()
    if permissions["all"]:
        return records
    return [record for record in records if _client_ids_permitted(record.get("fields", {}).get(field, []), permissions)]


def _filter_receipts_by_access(records):
    permissions = _permission_context()
    if permissions["all"]:
        return records
    return [record for record in records if _receipt_client_permitted(record.get("fields", {}).get(C.F_RECEIPT_CLIENT, []), permissions)]


def _list_all_record_ids(table_name):
    record_ids = []
    params = {"pageSize": 100}
    while True:
        data = airtable.list_records(table_name, params=params, by_field_id=False)
        record_ids.extend(record["id"] for record in data.get("records", []))
        offset = data.get("offset")
        if not offset:
            return record_ids
        params["offset"] = offset


# Clients, Locations and Users are read on nearly every request and change
# perhaps weekly. They were 8 of the 22 Airtable calls a single page load made,
# with Clients alone read six times. Cached briefly, and dropped the moment this
# application writes to the table, so the only staleness possible is an edit made
# directly in Airtable - which the TTL bounds.
REFERENCE_TABLE_TTL_SECONDS = 60
# One page load makes nine requests within about a second, and they re-read the
# same data tables: Merchandise three times, Products, Issues, Shipments and the
# workstream cards twice each. A short window collapses that burst into one read.
# Anything this application writes drops the cache immediately, so the only
# staleness possible is an edit made directly in Airtable during the window.
DATA_TABLE_TTL_SECONDS = 10
_REFERENCE_TABLE_CACHE = {}


def _reference_table_records(table_name):
    """One cached read of a whole reference table, shared by every caller."""
    cached = _cached_reference_records(table_name)
    if cached is not None:
        return list(cached)
    records = airtable.list_records(table_name, by_field_id=False).get("records", [])
    _remember_reference_records(table_name, list(records))
    return records


def _client_records():
    return _reference_table_records(C.CLIENTS_TABLE)


def _reference_tables():
    return {C.CLIENTS_TABLE, C.LOCATIONS_TABLE, C.USERS_TABLE}


def invalidate_reference_cache(table_name=None):
    if table_name is None:
        _REFERENCE_TABLE_CACHE.clear()
    else:
        _REFERENCE_TABLE_CACHE.pop(table_name, None)


def _table_cache_ttl(table_name):
    return REFERENCE_TABLE_TTL_SECONDS if table_name in _reference_tables() else DATA_TABLE_TTL_SECONDS


def _cached_reference_records(table_name):
    entry = _REFERENCE_TABLE_CACHE.get(table_name)
    if entry and (time.monotonic() - entry[0]) < _table_cache_ttl(table_name):
        return entry[1]
    return None


def _remember_reference_records(table_name, records):
    _REFERENCE_TABLE_CACHE[table_name] = (time.monotonic(), records)


def _install_reference_cache_invalidation():
    """Any write through the client drops that table's cache."""
    client = airtable
    if getattr(client, "_reference_cache_installed", False):
        return
    for name in ("create_record", "update_record", "delete_records"):
        original = getattr(client, name)

        def wrapper(table_name, *args, _original=original, **kwargs):
            invalidate_reference_cache(table_name)
            return _original(table_name, *args, **kwargs)

        setattr(client, name, wrapper)
    client._reference_cache_installed = True


_install_reference_cache_invalidation()


def _list_all_records(table_name, params=None):
    # Only unfiltered scans are cached; a filtered read asks a different question.
    cacheable = params is None
    if cacheable:
        cached = _cached_reference_records(table_name)
        if cached is not None:
            return list(cached)
    records = []
    merged = {"pageSize": 100, **(params or {})}
    while True:
        data = airtable.list_records(table_name, params=merged, by_field_id=False)
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            if cacheable:
                _remember_reference_records(table_name, list(records))
            return records
        merged["offset"] = offset


def _delete_records_in_batches(table_name, record_ids):
    deleted = 0
    for index in range(0, len(record_ids), 10):
        batch = record_ids[index:index + 10]
        if batch:
            airtable.delete_records(table_name, batch)
            deleted += len(batch)
    return deleted


def _permitted_client_records(records):
    permissions = _permission_context()
    if permissions["all"]:
        return records
    return [record for record in records if record.get("id") in permissions["client_ids"]]


def _record_client_ids(table, record_id):
    if not record_id:
        return []
    try:
        record = airtable.get_record(table, record_id, by_field_id=False)
    except requests.HTTPError:
        return []
    fields = record.get("fields", {})
    if table == C.PRODUCTS_TABLE:
        return fields.get(C.F_ITEM_CLIENT, []) or []
    if table == C.SHIPMENTS_TABLE:
        return fields.get(C.F_RECEIPT_CLIENT, []) or []
    return []




def _item_client_ids(item_id):
    return _record_client_ids(C.PRODUCTS_TABLE, item_id)


def _client_ids_for_issue(record):
    fields = record.get("fields", {})
    client_ids = set()
    for item_id in fields.get(C.F_ISSUE_ITEM, []) or []:
        client_ids.update(_item_client_ids(item_id))
    return list(client_ids)


def _client_ids_for_history(record):
    fields = record.get("fields", {})
    client_ids = set()
    for item_id in fields.get(C.F_HISTORY_ITEM, []) or []:
        client_ids.update(_item_client_ids(item_id))
    return list(client_ids)


def _filter_indirect_client_records(records, resolver):
    permissions = _permission_context()
    if permissions["all"]:
        return records
    return [record for record in records if _client_ids_permitted(resolver(record), permissions)]


def _all_linked_records_permitted(table, ids):
    return all(_client_ids_permitted(_record_client_ids(table, record_id)) for record_id in _as_list(ids))


def _forbidden():
    return err("You do not have access to that Client.", 403)


def _linked_location_names(ids):
    names = []
    for record_id in _as_list(ids):
        try:
            record = airtable.get_record(C.LOCATIONS_TABLE, record_id, by_field_id=False)
        except requests.HTTPError:
            continue
        name = record.get("fields", {}).get(C.F_LOCATION_NAME)
        if name:
            names.append(name)
    return ", ".join(names)


def _history_signature(fields):
    parts = [
        fields.get(C.F_HISTORY_EVENT, ""),
        ",".join(fields.get(C.F_HISTORY_ITEM, []) or []),
        ",".join(fields.get(C.F_HISTORY_MERCHANDISE, []) or []),
        fields.get(C.F_HISTORY_FROM, ""),
        fields.get(C.F_HISTORY_TO, ""),
    ]
    return " | ".join(parts)


def _history_exists(fields):
    params = {"maxRecords": 100, "sort[0][field]": C.F_HISTORY_DATE, "sort[0][direction]": "desc"}
    try:
        data = airtable.list_records(C.HISTORY_TABLE, params=params, by_field_id=False)
    except requests.HTTPError:
        return False

    signature = _history_signature(fields)
    now = datetime.now(timezone.utc)
    for record in data.get("records", []):
        existing = record.get("fields", {})
        event_date = existing.get(C.F_HISTORY_DATE)
        if event_date:
            try:
                parsed = datetime.fromisoformat(event_date.replace("Z", "+00:00"))
            except ValueError:
                parsed = None
            if parsed and (now - parsed).total_seconds() > 120:
                continue
        existing_fields = {
            C.F_HISTORY_EVENT: existing.get(C.F_HISTORY_EVENT, ""),
            C.F_HISTORY_ITEM: existing.get(C.F_HISTORY_ITEM, []),
            C.F_HISTORY_MERCHANDISE: existing.get(C.F_HISTORY_MERCHANDISE, []),
            C.F_HISTORY_FROM: existing.get(C.F_HISTORY_FROM, ""),
            C.F_HISTORY_TO: existing.get(C.F_HISTORY_TO, ""),
        }
        if _history_signature(existing_fields) == signature:
            return True
    return False


def _create_history_event(event, *, item_ids=None, user_ids=None, merchandise_ids=None, from_value=None, to_value=None):
    fields = {
        C.F_HISTORY_EVENT: event,
        C.F_HISTORY_DATE: _now_iso(),
    }
    if item_ids:
        fields[C.F_HISTORY_ITEM] = _as_list(item_ids)
    if merchandise_ids:
        fields[C.F_HISTORY_MERCHANDISE] = _as_list(merchandise_ids)
    if user_ids:
        fields[C.F_HISTORY_USER] = _as_list(user_ids)
    if from_value not in (None, ""):
        fields[C.F_HISTORY_FROM] = str(from_value)
    if to_value not in (None, ""):
        fields[C.F_HISTORY_TO] = str(to_value)

    if _history_exists(fields):
        return None
    try:
        return airtable.create_record(C.HISTORY_TABLE, fields, by_field_id=False)
    except requests.HTTPError:
        return None


# ── Health ────────────────────────────────────────────────────────────────────

@api.get("/health")
def health():
    """Presence only. Never report a configured value, just whether it is set:
    this endpoint is unauthenticated, and "which settings are missing" is the
    question worth answering from outside the box."""
    return jsonify({
        "status": "ok",
        "service": "Marks Food Photography API",
        "airtableConfigured": C.airtable_ready(),
        "configured": {
            name: bool(getattr(C, name, ""))
            for name in (
                "AIRTABLE_API_KEY",
                "AIRTABLE_BASE_ID",
                "SECRET_KEY",
                "R2_ACCOUNT_ID",
                "R2_ACCESS_KEY_ID",
                "R2_SECRET_ACCESS_KEY",
                "R2_BUCKET_NAME",
                "R2_PUBLIC_BASE_URL",
                "CREATIVE_FORCE_WEBHOOK_SECRET",
            )
        },
        "corsOrigins": C.cors_origins(),
    })


@api.get("/airtable/status")
def airtable_status():
    return jsonify({
        "configured": airtable.is_configured,
        "baseIdPresent": bool(C.AIRTABLE_BASE_ID),
        "apiKeyPresent": bool(C.AIRTABLE_API_KEY),
    })


# ── Clients ───────────────────────────────────────────────────────────────────

TOPCO_READINESS_PROFILE = {
    "mode": "activation_driven",
    "label": "Activation-driven",
    "matchingTarget": "Activation row linked to received Merchandise",
    "matchKeys": ["UPC"],
    "readyForPhotoRequires": [
        "Merchandise received",
        "Activation confirmed",
        "Activation row linked",
        "Deliverables confirmed",
    ],
    "sources": [
        {"label": "Activation Package", "description": "Marks creates and stores the Topco project readiness package."},
        {"label": "Shipments", "description": "Marks captures received quantity, photos, and physical handling."},
    ],
    "notRequiredFromActivation": [
        "Quantity received",
        "Storage location",
        "Individual file names",
        "Post-photo tracking statuses",
    ],
    "sourceCheckRules": {
        "version": 1,
        "sourceIdentityFields": ["productName", "upc"],
        "activationField": "requestType",
        "requiredToProceed": {
            "Packaging": ["productName", "upc", "jobNumber"],
            "Ecomm": ["productName", "upc", "cvid"],
        },
        "sourceFieldMappings": [
            {
                "sourceField": "Prod Descrip",
                "usedAs": ["File Name Description", "Product Description"],
                "note": "For this Topco slice, Prod Descrip satisfies the packaging filename/handoff token and the display Product Description.",
            },
        ],
        "requestTypeMappings": {
            "ecomm only": {"requiredDeliverables": ["Ecomm"], "label": "Ecomm required to proceed"},
            "pack only": {"requiredDeliverables": ["Packaging"], "label": "Packaging required to proceed"},
            "ecomm pack": {"requiredDeliverables": ["Ecomm", "Packaging"], "label": "Ecomm + Packaging required to proceed"},
            "ecomm and pack": {"requiredDeliverables": ["Ecomm", "Packaging"], "label": "Ecomm + Packaging required to proceed"},
            "pack thr3d": {
                "requiredDeliverables": ["Packaging"],
                "shipmentContext": ["Thr3d"],
                "label": "Packaging required to proceed + Thr3d shipment context",
            },
            "pack and thr3d": {
                "requiredDeliverables": ["Packaging"],
                "shipmentContext": ["Thr3d"],
                "label": "Packaging required to proceed + Thr3d shipment context",
            },
            "thr3d only": {
                "requiredDeliverables": [],
                "noWalnutWorkExpected": True,
                "alertIfReceived": True,
                "label": "No Walnut work expected; alert if received",
            },
            "not needed": {
                "requiredDeliverables": [],
                "noWalnutWorkExpected": True,
                "alertIfReceived": True,
                "label": "No Walnut work expected; alert if received",
            },
        },
    },
    "sourceRefresh": {
        "enabled": C.TOPCO_SOURCE_REFRESH_ENABLED,
        "intervalSeconds": C.TOPCO_SOURCE_REFRESH_INTERVAL_SECONDS,
        "limit": C.TOPCO_SOURCE_REFRESH_LIMIT,
        "provider": "topco",
    },
}


def _normalize_source_refresh_config(value, fallback=None):
    fallback = fallback or {}
    if value in (None, ""):
        value = {}
    if not isinstance(value, dict) or isinstance(value, list):
        raise ValueError("Source refresh must be a JSON object.")
    try:
        interval_seconds = int(value.get("intervalSeconds", fallback.get("intervalSeconds", 300)) or 300)
    except (TypeError, ValueError):
        raise ValueError("Source refresh interval must be a number of seconds.")
    try:
        refresh_limit = int(value.get("limit", fallback.get("limit", 100)) or 100)
    except (TypeError, ValueError):
        raise ValueError("Source refresh limit must be a number.")
    return {
        "enabled": bool(value.get("enabled", fallback.get("enabled", False))),
        "intervalSeconds": max(60, interval_seconds),
        "limit": max(1, refresh_limit),
        "provider": str(value.get("provider") or fallback.get("provider") or "").strip(),
    }


def _client_readiness_profile(client_name, client_fields=None):
    if (client_name or "").strip().lower() == "topco":
        profile = copy.deepcopy(TOPCO_READINESS_PROFILE)
        if client_fields:
            requirements, _ = _parse_photo_production_requirements(
                client_fields.get(C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS, ""),
                client_name=client_name,
            )
            refresh_override = requirements.get("sourceRefresh") if isinstance(requirements, dict) else None
            if isinstance(refresh_override, dict):
                profile["sourceRefresh"] = _normalize_source_refresh_config(refresh_override, profile.get("sourceRefresh"))
        return profile
    return None


def source_refresh_client_configs():
    data = {"records": _client_records()}
    configs = []
    for record in data.get("records", []):
        client_name = _client_name(record)
        profile = _client_readiness_profile(client_name, record.get("fields", {})) or {}
        refresh = profile.get("sourceRefresh") if isinstance(profile, dict) else None
        if not isinstance(refresh, dict) or not refresh.get("enabled"):
            continue
        try:
            interval_seconds = int(refresh.get("intervalSeconds") or 300)
        except (TypeError, ValueError):
            interval_seconds = 300
        try:
            refresh_limit = int(refresh.get("limit") or 100)
        except (TypeError, ValueError):
            refresh_limit = 100
        configs.append({
            "clientId": record.get("id", ""),
            "clientName": client_name,
            "provider": str(refresh.get("provider") or "").strip(),
            "intervalSeconds": max(60, interval_seconds),
            "limit": max(1, refresh_limit),
        })
    return configs


@api.get("/clients")
def list_clients():
    data = airtable.list_records(
        C.CLIENTS_TABLE,
        params={"sort[0][field]": C.F_CLIENT_NAME, "sort[0][direction]": "asc"},
        by_field_id=False,
    )
    raw_records = data.get("records", [])
    if request.args.get("all") == "1":
        management_error = _require_user_management()
        if management_error:
            return management_error
        permitted = raw_records
    else:
        permitted = _permitted_client_records(raw_records)
    records = [_shape_client(r) for r in permitted]
    clients = [{"id": r["id"], "client": r["name"]} for r in records]
    return jsonify({"records": records, "clients": clients})


@api.post("/clients")
def create_client():
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    body = request.get_json(silent=True) or {}
    try:
        fields = _client_fields_from_body(body, creating=True)
    except ValueError as e:
        return err(str(e))
    try:
        record = airtable.create_record(C.CLIENTS_TABLE, fields, by_field_id=False, typecast=True)
    except requests.HTTPError as e:
        return airtable_err(e)
    schema_sync = _ensure_creative_force_feed_schema() if C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS in fields else None
    response = {"client": _shape_client(record)}
    if schema_sync is not None:
        response["creativeForceFeedSchema"] = schema_sync
    return jsonify(response), 201


@api.put("/clients/<client_id>")
@api.patch("/clients/<client_id>")
def update_client(client_id):
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    body = request.get_json(silent=True) or {}
    try:
        fields = _client_fields_from_body(body)
    except ValueError as e:
        return err(str(e))
    try:
        if fields:
            record = airtable.update_record(C.CLIENTS_TABLE, client_id, fields, by_field_id=False, typecast=True)
        else:
            record = airtable.get_record(C.CLIENTS_TABLE, client_id, by_field_id=False)
    except requests.HTTPError as e:
        return airtable_err(e)
    schema_sync = _ensure_creative_force_feed_schema() if C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS in fields else None
    response = {"client": _shape_client(record)}
    if schema_sync is not None:
        response["creativeForceFeedSchema"] = schema_sync
    return jsonify(response)


def _ensure_creative_force_feed_schema():
    """Reconcile feed columns after Client photo requirements change."""
    try:
        from ensure_creative_force_product_feed import ensure_feed_schema

        result = ensure_feed_schema()
        return {"verified": True, **result}
    except Exception as error:  # Airtable availability must not erase the saved Client config.
        current_app.logger.exception("Creative Force feed schema reconciliation failed")
        return {"verified": False, "error": str(error)}


def _shape_client(r):
    f = r.get("fields", {})
    name = f.get(C.F_CLIENT_NAME, "")
    product_import_profiles, product_import_profiles_error = _parse_product_import_profiles(
        f.get(C.F_CLIENT_PRODUCT_IMPORT_PROFILES, "")
    )
    photo_production_requirements, photo_production_requirements_error = _parse_photo_production_requirements(
        f.get(C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS, ""), client_name=name
    )
    readiness_profile = _client_readiness_profile(name, f)
    return {
        "id": r["id"],
        "name": name,
        "codeType": f.get(C.F_CLIENT_IDENTIFIER_TYPE, ""),
        "identifierLabel": f.get(C.F_CLIENT_IDENTIFIER_LABEL, "") or "UPC / Product ID",
        "primaryMatchKeyLabel": f.get(C.F_CLIENT_IDENTIFIER_LABEL, "") or "UPC / Product ID",
        "requiredToShoot": f.get(C.F_CLIENT_REQUIRED_TO_SHOOT, []) or ["Identifier"],
        "artworkRequirement": f.get(C.F_CLIENT_ARTWORK_REQUIREMENT, "") or "Optional",
        "merchandiseRequired": f.get(C.F_CLIENT_MERCHANDISE_REQUIRED, True),
        "photoReleaseRecipients": f.get(C.F_CLIENT_PHOTO_RELEASE_RECIPIENTS, ""),
        # The URL itself never leaves the server: holding it is enough to post to
        # the channel. The Admin screen only needs to know whether one is set.
        "teamsWebhookConfigured": bool(str(f.get(C.F_CLIENT_TEAMS_WEBHOOK, "") or "").strip()),
        "holdDays": f.get(C.F_CLIENT_HOLD_DAYS),
        "dispoDays": f.get(C.F_CLIENT_DISPO_DAYS),
        "active": f.get(C.F_CLIENT_ACTIVE, False),
        "productImportProfiles": product_import_profiles,
        "productImportProfilesRaw": f.get(C.F_CLIENT_PRODUCT_IMPORT_PROFILES, "") or "",
        "productImportProfilesError": product_import_profiles_error,
        "photoProductionRequirements": photo_production_requirements,
        "photoProductionRequirementsRaw": f.get(C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS, "") or "",
        "photoProductionRequirementsError": photo_production_requirements_error,
        "readinessProfile": readiness_profile,
        "sourceCheckRules": (readiness_profile or {}).get("sourceCheckRules"),
    }


def _empty_product_import_profiles():
    return {"defaultProfile": "", "profiles": {}}


def _normalize_product_import_profiles(value):
    if value in (None, ""):
        return _empty_product_import_profiles()
    if not isinstance(value, dict) or isinstance(value, list):
        raise ValueError("Product Import Profiles must be a JSON object.")

    profiles = value.get("profiles") or {}
    if not isinstance(profiles, dict) or isinstance(profiles, list):
        raise ValueError("Product Import Profiles profiles must be an object.")

    normalized_profiles = {}
    for profile_name, profile in profiles.items():
        clean_name = str(profile_name or "").strip()
        if not clean_name:
            continue
        if not isinstance(profile, dict) or isinstance(profile, list):
            raise ValueError(f"Product Import Profile {clean_name} must be an object.")
        normalized_profile = {}
        for key in ["sourceHeaders", "targetMapping", "referenceDataTargets"]:
            raw_mapping = profile.get(key) or {}
            if not isinstance(raw_mapping, dict) or isinstance(raw_mapping, list):
                raise ValueError(f"Product Import Profile {clean_name}.{key} must be an object.")
            normalized_profile[key] = raw_mapping
        required_targets = profile.get("requiredTargets") or []
        if not isinstance(required_targets, list):
            raise ValueError(f"Product Import Profile {clean_name}.requiredTargets must be a list.")
        normalized_profile["requiredTargets"] = [str(target) for target in required_targets if str(target).strip()]
        normalized_profiles[clean_name] = normalized_profile

    default_profile = str(value.get("defaultProfile") or "").strip()
    if default_profile and default_profile not in normalized_profiles:
        raise ValueError("Product Import Profiles defaultProfile must match a named profile.")
    return {
        "defaultProfile": default_profile,
        "profiles": normalized_profiles,
    }


def _parse_product_import_profiles(raw):
    if not raw:
        return _empty_product_import_profiles(), ""
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
        return _normalize_product_import_profiles(parsed), ""
    except (TypeError, ValueError, json.JSONDecodeError):
        return _empty_product_import_profiles(), "Malformed Product Import Profiles JSON."


def _product_import_profiles_json(value):
    if isinstance(value, str):
        if not value.strip():
            return ""
        try:
            value = json.loads(value)
        except json.JSONDecodeError as e:
            raise ValueError("Product Import Profiles must be valid JSON.") from e
    normalized = _normalize_product_import_profiles(value)
    if normalized == _empty_product_import_profiles():
        return ""
    return json.dumps(normalized, sort_keys=True)


PHOTO_PRODUCTION_REQUIREMENT_FIELDS = {
    "productName": "Product Name",
    "upc": "UPC / Product ID",
    "cvid": "CVID",
    "jobNumber": "WKFT Job Number",
    "brandPrefix": "Brand Prefix",
    "fileNameDescription": "File Name Description",
    "productType": "Product Type",
    "ecommPhotoNotes": "Ecomm Photo Notes",
    "pathToArt": "Valid Artwork Path",
}
# Entered on the release form, not carried on the Product, so they are kept apart
# from requiredProductFields - listing them there would leave Planning asking for a
# Product field that does not exist.
PHOTO_RELEASE_REQUIREMENT_FIELDS = {
    "artworkPath": "Artwork Path",
    "uploadLocation": "Upload Location",
}
PHOTO_PRODUCTION_CATEGORY_FIELDS = {
    "clientName": "Client Name",
    "productName": "Product Name",
    "brandPrefix": "Brand Prefix",
    "productType": "Product Type",
    "custom": "Custom value",
}


def _empty_photo_production_requirements():
    return {"version": 1, "workstreams": {}}


def _normalize_photo_production_requirements(value):
    if value in (None, ""):
        return _empty_photo_production_requirements()
    if not isinstance(value, dict) or isinstance(value, list):
        raise ValueError("Photo Production Requirements must be a JSON object.")
    workstreams = value.get("workstreams") or {}
    if not isinstance(workstreams, dict) or isinstance(workstreams, list):
        raise ValueError("Photo Production Requirements workstreams must be an object.")
    normalized = {}
    for workstream_type, config in workstreams.items():
        name = str(workstream_type or "").strip()
        if name not in C.WORKSTREAM_TYPE_OPTIONS:
            raise ValueError(f"Unsupported photo workstream: {name}.")
        if not isinstance(config, dict) or isinstance(config, list):
            raise ValueError(f"Photo requirements for {name} must be an object.")
        fields = config.get("requiredProductFields") or []
        if not isinstance(fields, list):
            raise ValueError(f"Photo requirements for {name}.requiredProductFields must be a list.")
        invalid = [str(field) for field in fields if str(field) not in PHOTO_PRODUCTION_REQUIREMENT_FIELDS]
        if invalid:
            raise ValueError(f"Unsupported Product requirement for {name}: {', '.join(invalid)}.")
        naming = config.get("naming") or {}
        if not isinstance(naming, dict) or isinstance(naming, list):
            raise ValueError(f"Photo requirements for {name}.naming must be an object.")
        tokens = naming.get("tokens") or []
        if not isinstance(tokens, list):
            raise ValueError(f"Photo requirements for {name}.naming.tokens must be a list.")
        views = naming.get("views") or []
        if not isinstance(views, list):
            raise ValueError(f"Photo requirements for {name}.naming.views must be a list.")
        separator = naming.get("separator", "_")
        if separator not in ["", "_", "-", ".", " "]:
            raise ValueError(f"Photo requirements for {name}.naming.separator must be one of: _, -, ., space, or none.")
        normalized_config = {
            "requiredProductFields": [str(field) for field in fields if str(field) in PHOTO_PRODUCTION_REQUIREMENT_FIELDS],
            "naming": {
                "template": str(naming.get("template") or "").strip(),
                "tokens": [str(token).strip() for token in tokens if str(token).strip()],
                "views": [str(view).strip() for view in views if str(view).strip()],
                "separator": separator,
            },
        }
        release = config.get("release") or {}
        if not isinstance(release, dict) or isinstance(release, list):
            raise ValueError(f"Photo requirements for {name}.release must be an object.")
        release_fields = release.get("requiredFields") or []
        if not isinstance(release_fields, list):
            raise ValueError(f"Photo requirements for {name}.release.requiredFields must be a list.")
        invalid_release = [str(field) for field in release_fields if str(field) not in PHOTO_RELEASE_REQUIREMENT_FIELDS]
        if invalid_release:
            raise ValueError(f"Unsupported release requirement for {name}: {', '.join(invalid_release)}.")
        if release_fields:
            normalized_config["release"] = {"requiredFields": [str(field) for field in release_fields]}
        paths = config.get("paths") or {}
        if not isinstance(paths, dict) or isinstance(paths, list):
            raise ValueError(f"Photo requirements for {name}.paths must be an object.")
        normalized_paths = {
            key: str(paths.get(key) or "").strip()
            for key in ("artwork", "upload")
            if str(paths.get(key) or "").strip()
        }
        if normalized_paths:
            normalized_config["paths"] = normalized_paths
        if "creativeForce" in config:
            creative_force = config.get("creativeForce") or {}
            if not isinstance(creative_force, dict) or isinstance(creative_force, list):
                raise ValueError(f"Photo requirements for {name}.creativeForce must be an object.")
            product_code_field = str(creative_force.get("productCodeField") or "").strip()
            if product_code_field and product_code_field not in PHOTO_PRODUCTION_REQUIREMENT_FIELDS:
                raise ValueError(f"Unsupported Creative Force Product Code field for {name}: {product_code_field}.")
            category_field = str(creative_force.get("categoryField") or creative_force.get("categorySource") or "clientName").strip()
            if category_field not in PHOTO_PRODUCTION_CATEGORY_FIELDS:
                raise ValueError(f"Unsupported Creative Force Category source for {name}: {category_field}.")
            category_value = str(creative_force.get("categoryValue") or "").strip()
            if category_field == "custom" and not category_value:
                raise ValueError(f"Creative Force Category custom value is required for {name}.")
            normalized_config["creativeForce"] = {
                "productCodeField": product_code_field,
                "categoryField": category_field,
                "categoryValue": category_value,
            }
        normalized[name] = normalized_config
    normalized_requirements = {"version": 1, "workstreams": normalized}
    if "sourceRefresh" in value:
        normalized_requirements["sourceRefresh"] = _normalize_source_refresh_config(
            value.get("sourceRefresh"),
            TOPCO_READINESS_PROFILE.get("sourceRefresh"),
        )
    return normalized_requirements


def _parse_photo_production_requirements(raw, client_name=""):
    if not raw:
        return _empty_photo_production_requirements(), ""
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
        return _normalize_photo_production_requirements(parsed), ""
    except (TypeError, ValueError, json.JSONDecodeError):
        return _empty_photo_production_requirements(), "Malformed Photo Production Requirements JSON."


def _photo_production_requirements_json(value):
    if isinstance(value, str):
        if not value.strip():
            return ""
        try:
            value = json.loads(value)
        except json.JSONDecodeError as error:
            raise ValueError("Photo Production Requirements must be valid JSON.") from error
    normalized = _normalize_photo_production_requirements(value)
    if normalized == _empty_photo_production_requirements():
        return ""
    return json.dumps(normalized, sort_keys=True)


def _client_fields_from_body(body, *, creating=False):
    fields = {}
    if creating:
        name = (body.get("name") or body.get("client") or "").strip()
        if not name:
            raise ValueError("name is required")
        fields[C.F_CLIENT_NAME] = name
        fields[C.F_CLIENT_ACTIVE] = True
    for key, const in [
        ("name", C.F_CLIENT_NAME),
        ("client", C.F_CLIENT_NAME),
        ("codeType", C.F_CLIENT_IDENTIFIER_TYPE),
        ("identifierLabel", C.F_CLIENT_IDENTIFIER_LABEL),
        ("primaryMatchKeyLabel", C.F_CLIENT_IDENTIFIER_LABEL),
        ("artworkRequirement", C.F_CLIENT_ARTWORK_REQUIREMENT),
    ]:
        if key in body and not (creating and const == C.F_CLIENT_NAME):
            fields[const] = (body.get(key) or "").strip()
    if "requiredToShoot" in body:
        required = body.get("requiredToShoot") or []
        if not isinstance(required, list):
            raise ValueError("requiredToShoot must be a list.")
        fields[C.F_CLIENT_REQUIRED_TO_SHOOT] = [str(field) for field in required if str(field).strip()]
    if "merchandiseRequired" in body:
        fields[C.F_CLIENT_MERCHANDISE_REQUIRED] = bool(body["merchandiseRequired"])
    if "active" in body:
        fields[C.F_CLIENT_ACTIVE] = bool(body["active"])
    if "holdDays" in body:
        fields[C.F_CLIENT_HOLD_DAYS] = _int_or_none(body.get("holdDays"))
    if "dispoDays" in body:
        fields[C.F_CLIENT_DISPO_DAYS] = _int_or_none(body.get("dispoDays"))
    if "productImportProfiles" in body:
        fields[C.F_CLIENT_PRODUCT_IMPORT_PROFILES] = _product_import_profiles_json(body.get("productImportProfiles"))
    elif "productImportProfilesRaw" in body:
        fields[C.F_CLIENT_PRODUCT_IMPORT_PROFILES] = _product_import_profiles_json(body.get("productImportProfilesRaw"))
    if "photoProductionRequirements" in body:
        fields[C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS] = _photo_production_requirements_json(body.get("photoProductionRequirements"))
    elif "photoProductionRequirementsRaw" in body:
        fields[C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS] = _photo_production_requirements_json(body.get("photoProductionRequirementsRaw"))
    return fields


def _json_field(value, fallback):
    if isinstance(value, (dict, list)):
        return value
    if not value:
        return fallback
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return fallback


def _int_or_none(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _shape_activation(r):
    f = r.get("fields", {})
    client_ids = f.get(C.F_ACTIVATION_CLIENT, []) or []
    return {
        "id": r["id"],
        "name": f.get(C.F_ACTIVATION_NAME, ""),
        "clientIds": client_ids,
        "status": f.get(C.F_ACTIVATION_STATUS, "") or "Draft",
        "projectReference": f.get(C.F_ACTIVATION_PROJECT_REFERENCE, ""),
        "activationPackage": f.get(C.F_ACTIVATION_PACKAGE, ""),
        "activationDate": f.get(C.F_ACTIVATION_DATE, ""),
        "dueUrgency": f.get(C.F_ACTIVATION_DUE_URGENCY, ""),
        "walnutScope": f.get(C.F_ACTIVATION_WALNUT_SCOPE, ""),
        "numberOfSkus": _int_or_none(f.get(C.F_ACTIVATION_NUMBER_OF_SKUS)),
        "imagesPerBundle": _int_or_none(f.get(C.F_ACTIVATION_IMAGES_PER_BUNDLE)),
        "totalImages": _int_or_none(f.get(C.F_ACTIVATION_TOTAL_IMAGES)),
        "artworkPath": f.get(C.F_ACTIVATION_ARTWORK_PATH, ""),
        "uploadLocation": f.get(C.F_ACTIVATION_UPLOAD_LOCATION, ""),
        "skuDetails": _json_field(f.get(C.F_ACTIVATION_SKU_DETAILS_JSON, ""), []),
        "skuDetailsRaw": f.get(C.F_ACTIVATION_SKU_DETAILS_JSON, ""),
        "deliverables": f.get(C.F_ACTIVATION_DELIVERABLES, []) or [],
        "linkedMerchandiseIds": f.get(C.F_ACTIVATION_MATCHED_MERCHANDISE, []) or [],
        "matchedMerchandiseIds": f.get(C.F_ACTIVATION_MATCHED_MERCHANDISE, []) or [],
        "notes": f.get(C.F_ACTIVATION_NOTES, ""),
        "emailSubject": f.get(C.F_ACTIVATION_EMAIL_SUBJECT, ""),
        "emailBodyHtml": f.get(C.F_ACTIVATION_EMAIL_BODY_HTML, ""),
    }


def _activation_text(body, key):
    return str(body.get(key) or "").strip()


def _activation_number(body, key):
    value = body.get(key)
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        raise ValueError(f"{key} must be a number.")


def _activation_json_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return json.dumps(parsed, ensure_ascii=False)
        except ValueError:
            return value.strip()
    return json.dumps(value, ensure_ascii=False)


def _dedupe_activation_sku_details(value):
    if not isinstance(value, list):
        return value
    seen_merchandise_ids = set()
    result = []
    for row in value:
        if not isinstance(row, dict):
            result.append(row)
            continue
        merchandise_id = str(row.get("merchandiseId") or "").strip()
        if merchandise_id and merchandise_id in seen_merchandise_ids:
            continue
        if merchandise_id:
            seen_merchandise_ids.add(merchandise_id)
        result.append(row)
    return result


def _activation_package_missing(shaped, client=None):
    missing = []
    for label, value in [
        ("Name", shaped.get("name")),
        ("Walnut Scope", shaped.get("walnutScope")),
    ]:
        if not str(value or "").strip():
            missing.append(label)
    deliverables = _validate_deliverables(shaped.get("deliverables", []))
    if not isinstance(deliverables, list) or not deliverables:
        missing.append("Deliverables")
        deliverables = []
    requirements = (client or {}).get("photoProductionRequirements") or _empty_photo_production_requirements()
    workstreams = requirements.get("workstreams") if isinstance(requirements, dict) else {}
    required_fields = set()
    for deliverable in deliverables:
        config = workstreams.get(deliverable) if isinstance(workstreams, dict) else None
        if isinstance(config, dict):
            required_fields.update(str(field) for field in (config.get("requiredProductFields") or []))
    if "pathToArt" in required_fields and not str(shaped.get("artworkPath") or "").strip():
        missing.append("Artwork Path")
    if "uploadLocation" in required_fields and not str(shaped.get("uploadLocation") or "").strip():
        missing.append("Upload Location")
    sku_details = shaped.get("skuDetails") if isinstance(shaped.get("skuDetails"), list) else []
    if not sku_details:
        missing.append("Items")
    for index, row in enumerate(sku_details, start=1):
        row = row if isinstance(row, dict) else {}
        required_item_fields = [("Linked Merchandise", "merchandiseId")]
        if "productName" in required_fields or "description" in required_fields:
            required_item_fields.append(("Description", "description"))
        if "upc" in required_fields or "productId" in required_fields:
            required_item_fields.append(("UPC", "upc"))
        if "cvid" in required_fields:
            required_item_fields.append(("CVID", "cvid"))
        for label, key in required_item_fields:
            if not str(row.get(key) or "").strip():
                missing.append(f"Item {index} {label}")
    return missing


def _activation_linked_merchandise_ids(shaped):
    from_rows = [
        row.get("merchandiseId")
        for row in (shaped.get("skuDetails") if isinstance(shaped.get("skuDetails"), list) else [])
        if isinstance(row, dict)
    ]
    return _as_clean_string_list([*from_rows, *shaped.get("linkedMerchandiseIds", [])])


def _move_removed_activation_merchandise_to_waiting(merchandise_ids):
    moved = []
    for merchandise_id in merchandise_ids:
        try:
            entry = airtable.get_record(C.MERCHANDISE_TABLE, merchandise_id, by_field_id=False)
        except requests.HTTPError as exc:
            return airtable_err(exc)
        fields = entry.get("fields", {})
        linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
        if linked_receipts and _first_permitted_receipt(linked_receipts) is None:
            return _forbidden()
        if fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) != PLANNING_STATUS_LABELS["awaiting-photo-release"]:
            continue
        update_fields = {
            C.F_RECEIPT_ENTRY_PLANNING_STATUS: "New",
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED: False,
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY: "",
        }
        if _normalized_merch_status(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS)) != fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS):
            update_fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = _normalized_merch_status(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS))
        try:
            updated = _update_receipt_entry_record(merchandise_id, update_fields)
        except requests.HTTPError as exc:
            return airtable_err(exc)
        moved.append(_shape_verification_entry(updated))
    return moved


@api.get("/activations")
def list_activations():
    client_id = (request.args.get("clientId") or "").strip()
    if client_id and not _client_permitted(client_id):
        return err("You do not have access to this Client.", 403)
    try:
        data = airtable.list_records(
            C.ACTIVATIONS_TABLE,
            params={"sort[0][field]": C.F_ACTIVATION_NAME, "sort[0][direction]": "asc"},
            by_field_id=False,
        )
    except requests.HTTPError as exc:
        return airtable_err(exc)
    records = [_shape_activation(record) for record in data.get("records", [])]
    if client_id:
        records = [record for record in records if client_id in record["clientIds"]]
    else:
        records = [record for record in records if _client_ids_permitted(record["clientIds"])]
    return jsonify({"records": records})


@api.post("/activations")
def create_activation():
    return _save_activation()


@api.patch("/activations/<activation_id>")
def update_activation(activation_id):
    return _save_activation(activation_id)


@api.post("/activations/<activation_id>/move-to-photo")
def move_activation_to_photo(activation_id):
    try:
        activation = airtable.get_record(C.ACTIVATIONS_TABLE, activation_id, by_field_id=False)
    except requests.HTTPError as exc:
        return airtable_err(exc)
    shaped = _shape_activation(activation)
    if not _client_ids_permitted(shaped["clientIds"]):
        return _forbidden()
    try:
        clients = _clients_by_id()
    except requests.HTTPError as exc:
        return airtable_err(exc)
    client = next((clients.get(client_id) for client_id in shaped["clientIds"] if clients.get(client_id)), None)
    missing = _activation_package_missing(shaped, client)
    if missing:
        return err("Cannot release to photo until the photo release is complete.", 400, missing=missing)
    linked_merchandise_ids = _activation_linked_merchandise_ids(shaped)
    if not linked_merchandise_ids:
        return err("Cannot release to photo without linked Merchandise.", 400, missing=["Linked Merchandise"])
    deliverables = _validate_deliverables(shaped.get("deliverables", []))
    if not isinstance(deliverables, list):
        return deliverables
    # A photo release may target one workstream while its parent Merchandise
    # record still carries the union of all selected deliverables. Load the
    # child cards before updating the parent so a sibling can remain waiting.
    try:
        workstream_records = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
    except requests.HTTPError as exc:
        return airtable_err(exc)
    release_types = set(deliverables) - {"Thr3d"}
    moved = []
    for merchandise_id in linked_merchandise_ids:
        try:
            entry = airtable.get_record(C.MERCHANDISE_TABLE, merchandise_id, by_field_id=False)
        except requests.HTTPError as exc:
            return airtable_err(exc)
        fields = entry.get("fields", {})
        linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
        if linked_receipts and _first_permitted_receipt(linked_receipts) is None:
            return _forbidden()
        existing_deliverables = _deliverable_values(fields.get(C.F_RECEIPT_ENTRY_DELIVERABLES, []))
        combined_deliverables = existing_deliverables + [
            value for value in deliverables if value not in existing_deliverables
        ]
        photo_cards = []
        for workstream_record in workstream_records:
            workstream_fields = workstream_record.get("fields", {})
            if merchandise_id not in _as_list(workstream_fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])):
                continue
            workstream_type = str(workstream_fields.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip()
            if workstream_type in {"Packaging", "Ecomm"}:
                photo_cards.append(workstream_fields)
        unreleased_siblings = {
            str(card.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip()
            for card in photo_cards
            if str(card.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) or "").strip() != PLANNING_STATUS_LABELS["awaiting-photo-release"]
            and str(card.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip() not in release_types
        }
        update_fields = {
            C.F_RECEIPT_ENTRY_DELIVERABLES: combined_deliverables,
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED: True,
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT: datetime.now(timezone.utc).isoformat(),
        }
        if not unreleased_siblings:
            update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["awaiting-photo-release"]
            # This is a release, so it is stamped like one. Without it the board
            # could not tell a card released last week from one never released.
            if not fields.get(C.F_RECEIPT_ENTRY_RELEASED):
                update_fields[C.F_RECEIPT_ENTRY_RELEASED] = True
                update_fields[C.F_RECEIPT_ENTRY_RELEASED_AT] = _now_iso()
                releaser_id = _current_user_id()
                if releaser_id:
                    update_fields[C.F_RECEIPT_ENTRY_RELEASED_BY] = [releaser_id]
        if _normalized_merch_status(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS)) != fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS):
            update_fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = _normalized_merch_status(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS))
        current_user_id = _current_user_id()
        if current_user_id:
            update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY] = _current_user_display_name()
        try:
            updated = _update_receipt_entry_record(merchandise_id, update_fields)
        except requests.HTTPError as exc:
            return airtable_err(exc)
        # The client only needs the move result here. Avoid re-shaping the full
        # merchandise record, which performs unrelated linked-record lookups.
        moved.append({
            "id": updated.get("id", merchandise_id),
            "planningStatusLabel": update_fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) or fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS, ""),
            "deliverables": combined_deliverables,
        })
    # Release only the selected photo workstream types. Merchandise keeps the
    # aggregate deliverable set; each child card owns its own queue status.
    linked_types = release_types
    ready_workstream_cards = []
    for workstream_record in workstream_records:
        workstream_fields = workstream_record.get("fields", {})
        received_merchandise = _as_list(workstream_fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, []))
        workstream_type = str(workstream_fields.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip()
        if not set(received_merchandise).intersection(linked_merchandise_ids) or workstream_type not in linked_types:
            continue
        if workstream_fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) == PLANNING_STATUS_LABELS["awaiting-photo-release"]:
            # An already-ready card still needs to be projected into the CF
            # feed when this photo group is edited or released again.
            ready_workstream_cards.append(workstream_record)
            continue
        try:
            updated_workstream = airtable.update_record(
                C.WORKSTREAM_CARDS_TABLE,
                workstream_record.get("id"),
                {
                    C.F_WORKSTREAM_CARD_PLANNING_STATUS: PLANNING_STATUS_LABELS["awaiting-photo-release"],
                },
                by_field_id=False,
                typecast=True,
            )
            ready_workstream_cards.append(updated_workstream)
        except requests.HTTPError as exc:
            return airtable_err(exc)
    try:
        if ready_workstream_cards:
            _populate_creative_force_feed_for_ready_cards(ready_workstream_cards)
    except requests.HTTPError as exc:
        return airtable_err(exc)
    try:
        activation = airtable.update_record(
            C.ACTIVATIONS_TABLE,
            activation_id,
            {C.F_ACTIVATION_STATUS: "Released"},
            by_field_id=False,
        )
    except requests.HTTPError as exc:
        return airtable_err(exc)
    shaped_activation = _shape_activation(activation)
    # The release email is sent by whoever released it. Sending it from here would
    # need Microsoft Graph, and the tenant admin consent that requires was declined.
    recipients = mailer.parse_recipients((client or {}).get("photoReleaseRecipients", ""))
    return jsonify({
        "activation": shaped_activation,
        "moved": moved,
        "movedCount": len(moved),
        "emailSent": False,
        "emailDetail": "The release is recorded. Send the email from here.",
        # Returned whether or not it sent, so the board can hand an unsent
        # release to the user's own mail client instead of losing it.
        "email": {
            "subject": shaped_activation.get("emailSubject", ""),
            "html": shaped_activation.get("emailBodyHtml", ""),
            "recipients": recipients,
        },
    })


def _save_activation(activation_id=None):
    body = request.get_json(force=True, silent=True) or {}
    existing_activation = None
    previous_linked_merchandise_ids = []
    if activation_id:
        try:
            existing_activation = airtable.get_record(C.ACTIVATIONS_TABLE, activation_id, by_field_id=False)
        except requests.HTTPError as exc:
            return airtable_err(exc)
        existing_shaped = _shape_activation(existing_activation)
        if not _client_ids_permitted(existing_shaped["clientIds"]):
            return _forbidden()
        previous_linked_merchandise_ids = _activation_linked_merchandise_ids(existing_shaped)
    client_id = _activation_text(body, "clientId")
    if not client_id:
        return err("Client is required.")
    if not _client_permitted(client_id):
        return err("You do not have access to this Client.", 403)
    status = _activation_text(body, "status") or "Draft"
    if status not in C.ACTIVATION_STATUS_OPTIONS:
        return err(f"Status must be one of: {', '.join(C.ACTIVATION_STATUS_OPTIONS)}.")
    deliverables = _validate_deliverables(body.get("deliverables", []))
    if not isinstance(deliverables, list):
        return deliverables
    try:
        number_of_skus = _activation_number(body, "numberOfSkus")
        images_per_bundle = _activation_number(body, "imagesPerBundle")
        total_images = _activation_number(body, "totalImages")
    except ValueError as exc:
        return err(str(exc))
    name = _activation_text(body, "name") or _activation_text(body, "projectReference") or "New Activation"
    sku_details = _dedupe_activation_sku_details(body.get("skuDetails"))
    fields = {
        C.F_ACTIVATION_NAME: name,
        C.F_ACTIVATION_CLIENT: [client_id],
        C.F_ACTIVATION_STATUS: status,
        C.F_ACTIVATION_PROJECT_REFERENCE: _activation_text(body, "projectReference"),
        C.F_ACTIVATION_PACKAGE: _activation_text(body, "activationPackage"),
        C.F_ACTIVATION_DUE_URGENCY: _activation_text(body, "dueUrgency"),
        C.F_ACTIVATION_WALNUT_SCOPE: _activation_text(body, "walnutScope"),
        C.F_ACTIVATION_ARTWORK_PATH: _activation_text(body, "artworkPath"),
        C.F_ACTIVATION_UPLOAD_LOCATION: _activation_text(body, "uploadLocation"),
        C.F_ACTIVATION_SKU_DETAILS_JSON: _activation_json_text(sku_details),
        C.F_ACTIVATION_DELIVERABLES: deliverables,
        C.F_ACTIVATION_NOTES: _activation_text(body, "notes"),
        # Stored as released so the record shows what was actually sent.
        C.F_ACTIVATION_EMAIL_SUBJECT: _activation_text(body, "emailSubject"),
        C.F_ACTIVATION_EMAIL_BODY_HTML: _activation_text(body, "emailBodyHtml"),
    }
    activation_date = _activation_text(body, "activationDate")
    if activation_date:
        fields[C.F_ACTIVATION_DATE] = activation_date
    if number_of_skus is not None:
        fields[C.F_ACTIVATION_NUMBER_OF_SKUS] = number_of_skus
    if images_per_bundle is not None:
        fields[C.F_ACTIVATION_IMAGES_PER_BUNDLE] = images_per_bundle
    if total_images is not None:
        fields[C.F_ACTIVATION_TOTAL_IMAGES] = total_images
    matched_merchandise_ids = [
        item
        for item in _as_clean_string_list(body.get("linkedMerchandiseIds", body.get("matchedMerchandiseIds", [])))
        if item
    ]
    if "linkedMerchandiseIds" in body or "matchedMerchandiseIds" in body:
        fields[C.F_ACTIVATION_MATCHED_MERCHANDISE] = matched_merchandise_ids
    try:
        if activation_id:
            record = airtable.update_record(C.ACTIVATIONS_TABLE, activation_id, fields, by_field_id=False)
        else:
            record = airtable.create_record(C.ACTIVATIONS_TABLE, fields, by_field_id=False)
    except requests.HTTPError as exc:
        return airtable_err(exc)
    removed_linked_ids = [
        merchandise_id
        for merchandise_id in previous_linked_merchandise_ids
        if merchandise_id not in matched_merchandise_ids
    ]
    moved_to_waiting = _move_removed_activation_merchandise_to_waiting(removed_linked_ids) if activation_id and removed_linked_ids else []
    if isinstance(moved_to_waiting, tuple):
        return moved_to_waiting
    return jsonify({"record": _shape_activation(record), "movedToWaiting": moved_to_waiting}), 200 if activation_id else 201


REQUIRED_TO_SHOOT_LABELS = {
    "merchandise_issue": "Merchandise Issue",
    "waiting_for_merchandise": "Waiting for Merchandise",
    "missing_data": "Missing Data",
    "missing_artwork": "Missing Artwork",
    "ready_for_photo": "Ready",
}
MERCHANDISE_ISSUE_TYPES = {"Missing Merch", "Wrong Merch", "Damaged", "Unknown Item"}
RESOLVED_ISSUE_STATUSES = {"Resolved", "Cancelled"}
PRODUCTION_LOCK_STATUSES = {"Production", "Complete", "Cancelled"}


def _identifier_label(client):
    return (client or {}).get("identifierLabel") or "UPC / Product ID"


def _primary_match_key_value(item):
    return item.get("primaryMatchKey") or item.get("productId") or item.get("identifier") or item.get("gtinUpc") or ""


def _required_to_shoot_fields(client):
    fields = (client or {}).get("requiredToShoot") or ["Identifier"]
    return fields if isinstance(fields, list) else [fields]


def _validate_identifier_value(identifier, code_type, label="Primary Match Key"):
    value = str(identifier or "")
    code_type = code_type or ""
    if code_type == "GTIN-14" and not (value.isdigit() and len(value) == 14):
        return f"{label} must be exactly 14 digits."
    if code_type == "GTIN-13" and not (value.isdigit() and len(value) == 13):
        return f"{label} must be exactly 13 digits."
    if code_type == "GTIN-8" and not (value.isdigit() and len(value) == 8):
        return f"{label} must be exactly 8 digits."
    if code_type == "Numeric" and not value.isdigit():
        return f"{label} must contain digits only."
    if code_type in {"Text", "Item #"} and not value:
        return f"{label} is required."
    return ""


def _item_has_merchandise(item):
    return bool(item.get("received"))


def _blocking_merchandise_issues(issues):
    blockers = []
    for issue in issues or []:
        if issue.get("status") in RESOLVED_ISSUE_STATUSES:
            continue
        if issue.get("type") in MERCHANDISE_ISSUE_TYPES:
            blockers.append(issue)
    return blockers


def evaluate_required_to_shoot(item, client=None, issues=None, full=False):
    client = client or {}
    label = _identifier_label(client)
    required_fields = _required_to_shoot_fields(client)
    artwork_requirement = client.get("artworkRequirement") or "Optional"
    merchandise_required = client.get("merchandiseRequired")
    if merchandise_required is None:
        merchandise_required = True

    details = {
        "ready": False,
        "state": "missing_data",
        "label": REQUIRED_TO_SHOOT_LABELS["missing_data"],
        "missing": [],
        "warnings": [],
    }
    if full:
        details["requirements"] = {
            "identifierLabel": label,
            "primaryMatchKeyLabel": label,
            "codeType": client.get("codeType", ""),
            "requiredToShoot": required_fields,
            "artworkRequirement": artwork_requirement,
            "merchandiseRequired": merchandise_required,
        }

    if item.get("status") in PRODUCTION_LOCK_STATUSES:
        details.update({"ready": True, "state": "ready_for_photo", "label": REQUIRED_TO_SHOOT_LABELS["ready_for_photo"]})
        details["warnings"].append("Product is already in production or complete; Required to Shoot will not move it backward.")
        return details

    blockers = _blocking_merchandise_issues(issues)
    if blockers:
        details.update({"state": "merchandise_issue", "label": REQUIRED_TO_SHOOT_LABELS["merchandise_issue"]})
        details["missing"].append("Resolve merchandise issue.")
        if full:
            details["issues"] = blockers
        return details

    if merchandise_required and not _item_has_merchandise(item):
        details.update({"state": "waiting_for_merchandise", "label": REQUIRED_TO_SHOOT_LABELS["waiting_for_merchandise"]})
        details["missing"].append("Merchandise must be received and matched to this Product.")
        return details

    missing_data = []
    for field in required_fields:
        if field in {"Identifier", "ID"}:
            identifier = _primary_match_key_value(item)
            if not identifier:
                missing_data.append(f"{label} is required.")
            else:
                message = _validate_identifier_value(identifier, client.get("codeType", ""), label)
                if message:
                    missing_data.append(message)
        elif field in {"Product Name", "Product/File Name", "Product or File Name"} and not item.get("product"):
            missing_data.append("Product or File Name is required.")
        elif field == "Brand" and not item.get("brand"):
            missing_data.append("Brand is required.")
        elif field == "Artwork" and not item.get("artworkReceived"):
            missing_data.append("Artwork is required.")

    if missing_data:
        details.update({"state": "missing_data", "label": REQUIRED_TO_SHOOT_LABELS["missing_data"], "missing": missing_data})
        return details

    if artwork_requirement == "Required" and not item.get("artworkReceived"):
        details.update({"state": "missing_artwork", "label": REQUIRED_TO_SHOOT_LABELS["missing_artwork"]})
        details["missing"].append("Artwork is required.")
        return details
    if artwork_requirement == "Optional" and not item.get("artworkReceived"):
        details["warnings"].append("Artwork has not been received.")

    details.update({"ready": True, "state": "ready_for_photo", "label": REQUIRED_TO_SHOOT_LABELS["ready_for_photo"]})
    return details


# ── Intake preview ────────────────────────────────────────────────────────────

ALLOWED_INTAKE_EXTENSIONS = {".csv", ".xlsx", ".xls"}
XLSX_MAIN_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
XLSX_RELS_NS = {
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "office": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
INTAKE_FALLBACK_DESCRIPTIONS = {
    "Product Name": "Optional product display name in the app.",
    "Identifier": "Primary value used to match Received Merch to this expected Product. Stored in Airtable as Products.Identifier.",
    "UPC": "UPC match key. Stored separately from the internal Airtable Product record ID.",
    "CVID": "Client product reference used alongside the UPC when provided.",
    "Brand Prefix": "Client naming prefix for this product.",
    "Request Type": "Requested production route for this Product.",
    "Project Status": "Client project status for this Product.",
    "WKFT Job Number": "Client or production job reference.",
    "Mbox Number": "Client merchandise box reference.",
    "Product Type": "Product structure or storage category.",
    "Product Description": "Client-provided product description.",
    "Link to Prepro/Overlays": "Link to preproduction or overlay materials.",
    "Ecomm Photo Notes": "Notes for Ecomm photography.",
    "Path to Art": "Path or reference to product artwork.",
    "Product or File Name": "Product or file name.",
    "Description": "Longer source product description.",
    "Product Job Number": "Row-level job or project number for the product.",
    "Master or Variant": "Whether this product is a master or a variant.",
    "Pickup Job Number": "Previous production job number for variant pickup work.",
    "Brand": "Product brand.",
    "Due Date": "Job due date when present in the source spreadsheet.",
    "Notes": "Source notes that describe the product.",
    "Job Name": "Human-readable job or group name.",
    "Reference Data": "Preserve source values as product reference JSON.",
}
INTAKE_MAPPINGS = {
    "kroger": {
        "item_job_number": "Job #",
        "job_name": "Description",
        "item_name": "Product Received",
        "description": "Description",
        "id": "UPC",
        "brand": "Brand",
        "notes": ["Notes"],
    },
    "unfi": {
        "item_job_number": "Project Number",
        "item_name": "Description",
        "description": "Description",
        "id": "UPC",
        "notes": ["Notes"],
    },
    "smithfield": {
        "item_job_number": "Job #",
        "item_name": "Product Description",
        "description": "Product Description",
        "id": "GAR #",
        "brand": "Brand",
        "notes": ["Notes"],
    },
}


def _import_record_name(filename):
    return f"{filename or 'Spreadsheet'} - {_now_iso()}"


def _import_summary_fields(result, status):
    summary = result.get("summary") or {}
    if status == "Validated":
        return {
            C.F_IMPORT_STATUS: status,
            C.F_IMPORT_ROWS: result.get("totalRows", 0),
            C.F_IMPORT_ITEMS_CREATED: result.get("itemsToCreate", 0),
            C.F_IMPORT_ITEMS_UPDATED: result.get("itemsToUpdate", 0),
            C.F_IMPORT_ROWS_SKIPPED: sum(1 for row in result.get("rows", []) if row.get("errors")),
            C.F_IMPORT_ERRORS: result.get("errorCount", 0),
            C.F_IMPORT_WARNINGS: result.get("warningCount", 0),
        }
    return {
        C.F_IMPORT_STATUS: status,
        C.F_IMPORT_ROWS: result.get("totalRows", 0),
        C.F_IMPORT_ITEMS_CREATED: summary.get("itemsCreated", result.get("itemsToCreate", 0)),
        C.F_IMPORT_ITEMS_UPDATED: summary.get("itemsUpdated", result.get("itemsToUpdate", 0)),
        C.F_IMPORT_ROWS_SKIPPED: summary.get("rowsSkipped", 0),
        C.F_IMPORT_ERRORS: summary.get("errors", result.get("errorCount", 0)),
        C.F_IMPORT_WARNINGS: summary.get("warnings", result.get("warningCount", 0)),
    }


def _create_import_record(client_id, filename, status, rows=None, details=""):
    fields = {
        C.F_IMPORT_NAME: _import_record_name(filename),
        C.F_IMPORT_CLIENT: [client_id],
        C.F_IMPORT_FILE: filename,
        C.F_IMPORT_TYPE: "Spreadsheet",
        C.F_IMPORT_STATUS: status,
        C.F_IMPORT_STARTED: _now_iso(),
    }
    user_id = _current_user_id()
    if user_id:
        fields[C.F_IMPORT_USER] = [user_id]
    if rows is not None:
        fields[C.F_IMPORT_ROWS] = rows
    if details:
        fields[C.F_IMPORT_DETAILS] = details
    if status == "Failed":
        fields[C.F_IMPORT_FINISHED] = _now_iso()
    return airtable.create_record(C.IMPORTS_TABLE, fields, by_field_id=False)


def _update_import_record(import_id, fields):
    if not import_id:
        return None
    return airtable.update_record(C.IMPORTS_TABLE, import_id, fields, by_field_id=False)


def _fail_import_record(import_id, client_id, filename, message):
    fields = {
        C.F_IMPORT_STATUS: "Failed",
        C.F_IMPORT_FINISHED: _now_iso(),
        C.F_IMPORT_DETAILS: message,
    }
    if import_id:
        return _update_import_record(import_id, fields)
    if client_id and filename:
        return _create_import_record(client_id, filename, "Failed", details=message)
    return None


@api.post("/intake/structure-form/preview")
def structure_form_preview():
    """Read one or more Structure Forms and return what they propose.

    Preview only: nothing is written. A form describes expected work, not merchandise
    in hand, so reading one never creates a Product until it is committed.
    """
    uploads = request.files.getlist("files") or (
        [request.files["file"]] if "file" in request.files else []
    )
    if not uploads:
        return err("Upload at least one Structure Form PDF.")

    forms = []
    for uploaded in uploads:
        name = uploaded.filename or "form.pdf"
        if not name.lower().endswith(".pdf"):
            forms.append({"fileName": name, "error": "Not a PDF."})
            continue
        try:
            fields = structure_form.extract_form_fields(io.BytesIO(uploaded.read()))
        except ValueError as error:
            forms.append({"fileName": name, "error": str(error)})
            continue
        except Exception:  # noqa: BLE001
            current_app.logger.exception("Could not read structure form %s", name)
            forms.append({"fileName": name, "error": "This PDF could not be read."})
            continue
        plan = structure_form.parse_structure_form(fields)
        forms.append({"fileName": name, **plan})

    return jsonify({
        "forms": forms,
        "counts": {
            "forms": len(forms),
            "unreadable": sum(1 for form in forms if form.get("error")),
            "products": sum(len(form.get("rows") or []) for form in forms),
        },
    })


@api.post("/intake/structure-form/commit")
def structure_form_commit():
    """Create or update Products from reviewed Structure Form rows.

    Products only. A form says work is expected, not that merchandise has arrived, so
    nothing here creates Merchandise or a match. When the box turns up, receiving
    matches against the Products this created.

    Existing Products are matched by UPC and updated rather than duplicated, so
    re-reading a form is safe.
    """
    body = request.get_json(silent=True) or {}
    client_id = str(body.get("clientId") or "").strip()
    rows = body.get("rows") or []
    if not client_id:
        return err("Choose a Client before creating Products.")
    if not _client_permitted(client_id):
        return _forbidden()
    if not rows:
        return err("No rows to create.")

    try:
        product_records = _list_all_records(C.PRODUCTS_TABLE)
    except requests.HTTPError as error:
        return airtable_err(error)

    created, updated, skipped = [], [], []
    for row in rows:
        upc = str(row.get("upc") or "").strip()
        name = str(row.get("productName") or "").strip()
        if not upc or not name:
            skipped.append({"upc": upc, "reason": "A product needs both a name and a UPC."})
            continue

        patch = {
            "name": name,
            "upc": upc,
            "mboxNumber": str(row.get("mboxNumber") or "").strip(),
            "wkftJobNumber": str(row.get("wkftJobNumber") or "").strip(),
            "projectName": str(row.get("projectName") or "").strip(),
            "studioDestination": str(row.get("studio") or "").strip(),
            "vendor": str(row.get("supplier") or "").strip(),
            "ecommPhotoNotes": str(row.get("ecommPhotoNotes") or "").strip(),
        }
        request_type = _normalize_product_request_type(row.get("requestType"))
        if request_type:
            patch["requestType"] = request_type
        patch = {key: value for key, value in patch.items() if value}

        # The form only proposes a request type. A Product the source sheet owns
        # gets the real one from the sheet, so the guess is not offered at all -
        # not even into an empty field.
        known = _product_identifier_index(client_id, product_records).get(
            normalized_identifier(upc))
        if known is not None and _source_snapshot_for_topco_product(known):
            patch.pop("requestType", None)

        try:
            saved, outcome, _filled = merge_product(
                client_id, upc, patch,
                source=f"structureForm:{str(row.get('fileName') or '').strip()}".rstrip(":"),
                reference={"_structureForm": {
                    "fileName": str(row.get("fileName") or ""),
                    "project": str(row.get("project") or ""),
                    "importedAt": _now_iso(),
                }},
                records=product_records,
            )
        except ValueError as error:
            skipped.append({"upc": upc, "reason": str(error)})
            continue
        except requests.HTTPError as error:
            return airtable_err(error)

        if outcome == "created":
            product_records.append(saved)
            _create_history_event("Item Created", item_ids=[saved["id"]])
            created.append({"id": saved["id"], "upc": upc, "name": name})
        else:
            # "unchanged" is reported as updated: the row was accepted and the
            # Product already said everything the form had to say.
            updated.append({"id": saved["id"], "upc": upc, "name": name})

    return jsonify({
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "summary": {"created": len(created), "updated": len(updated), "skipped": len(skipped)},
    }), 201


@api.post("/intake/preview")
def intake_preview():
    client_id = (request.form.get("clientId") or "").strip()
    header_row = request.form.get("headerRow") or ""
    uploaded = request.files.get("file")

    if client_id and not _client_permitted(client_id):
        return _forbidden()
    if not uploaded or not uploaded.filename:
        return err("Upload an XLSX or CSV file.")

    filename = uploaded.filename
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_INTAKE_EXTENSIONS:
        return err("Invalid file type. Upload .xlsx, .xls, or .csv.")

    content = uploaded.read()
    if not content:
        return err("The uploaded file is empty.")

    try:
        parsed = _parse_spreadsheet(content, ext, header_row=header_row)
        import_record = _create_import_record(client_id, filename, "Parsed", rows=parsed.get("rowCount", 0)) if client_id else None
    except UnicodeDecodeError:
        _fail_import_record(None, client_id, filename, "Unreadable file. The CSV encoding could not be detected.")
        return err("Unreadable file. The CSV encoding could not be detected.")
    except zipfile.BadZipFile:
        _fail_import_record(None, client_id, filename, "Unreadable file. The Excel workbook could not be opened.")
        return err("Unreadable file. The Excel workbook could not be opened.")
    except ValueError as exc:
        _fail_import_record(None, client_id, filename, str(exc))
        return err(str(exc))
    except requests.HTTPError as error:
        return airtable_err(error)
    except Exception:
        _fail_import_record(None, client_id, filename, "Unreadable file. The spreadsheet could not be parsed.")
        return err("Unreadable file. The spreadsheet could not be parsed.")

    return jsonify({
        "fileName": filename,
        "clientId": client_id,
        "importId": import_record.get("id") if import_record else "",
        **parsed,
    })


@api.post("/intake/review")
def intake_review():
    return _intake_import_response(dry_run=True)


@api.post("/intake/import")
def intake_import():
    return _intake_import_response(dry_run=False)


def _intake_import_response(dry_run):
    if request.is_json:
        return _intake_import_json_response(dry_run)

    client_id = (request.form.get("clientId") or "").strip()
    import_id = (request.form.get("importId") or "").strip()
    uploaded = request.files.get("file")
    if not client_id:
        return err("Choose a client before uploading a spreadsheet.")
    if not _client_permitted(client_id):
        return _forbidden()
    if not uploaded or not uploaded.filename:
        return err("Upload an XLSX or CSV file.")

    filename = uploaded.filename
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_INTAKE_EXTENSIONS:
        return err("Invalid file type. Upload .xlsx, .xls, or .csv.")

    content = uploaded.read()
    if not content:
        return err("The uploaded file is empty.")

    try:
        parsed = _parse_spreadsheet(content, ext)
        result = _build_intake_plan(client_id, filename, parsed)
        if not dry_run:
            result = _execute_intake_plan(result)
            fields = _import_summary_fields(result, "Imported")
            fields[C.F_IMPORT_FINISHED] = _now_iso()
            _update_import_record(import_id, fields)
        else:
            _update_import_record(import_id, _import_summary_fields(result, "Validated"))
    except UnicodeDecodeError:
        _fail_import_record(import_id, client_id, filename, "Unreadable file. The CSV encoding could not be detected.")
        return err("Unreadable file. The CSV encoding could not be detected.")
    except zipfile.BadZipFile:
        _fail_import_record(import_id, client_id, filename, "Unreadable file. The Excel workbook could not be opened.")
        return err("Unreadable file. The Excel workbook could not be opened.")
    except ValueError as exc:
        _fail_import_record(import_id, client_id, filename, str(exc))
        return err(str(exc))
    except requests.HTTPError as error:
        _fail_import_record(import_id, client_id, filename, "Airtable request failed.")
        return airtable_err(error)
    except Exception:
        _fail_import_record(import_id, client_id, filename, "Unreadable file. The spreadsheet could not be imported.")
        return err("Unreadable file. The spreadsheet could not be imported.")

    result["dryRun"] = dry_run
    result["importId"] = import_id
    return jsonify(result)


def _intake_import_json_response(dry_run):
    body = request.get_json(silent=True) or {}
    client_id = (body.get("clientId") or "").strip()
    filename = (body.get("fileName") or "Edited import").strip()
    import_id = (body.get("importId") or "").strip()
    rows = body.get("rows") or []
    source_rows = body.get("sourceRows") or []
    headers = body.get("columnHeaders") or []
    mapping = body.get("mapping") or {}
    if not client_id:
        return err("Choose a client before importing.")
    if not _client_permitted(client_id):
        return _forbidden()
    if source_rows and (not isinstance(source_rows, list) or not isinstance(headers, list)):
        return err("Invalid source rows were provided for import.")
    if not source_rows and (not isinstance(rows, list) or not rows):
        return err("No rows were provided for import.")

    try:
        if not import_id:
            row_count = len(source_rows) if source_rows else len(rows)
            import_record = _create_import_record(client_id, filename, "Parsed", rows=row_count)
            import_id = import_record.get("id", "")
        if source_rows:
            result = _build_intake_plan_from_source_rows(client_id, filename, headers, source_rows, mapping)
        else:
            result = _build_intake_plan_from_mapped_rows(client_id, filename, rows)
        if not dry_run:
            result = _execute_intake_plan(result)
            fields = _import_summary_fields(result, "Imported")
            fields[C.F_IMPORT_FINISHED] = _now_iso()
            _update_import_record(import_id, fields)
        else:
            _update_import_record(import_id, _import_summary_fields(result, "Validated"))
    except requests.HTTPError as error:
        _fail_import_record(import_id, client_id, filename, "Airtable request failed.")
        return airtable_err(error)
    except ValueError as exc:
        _fail_import_record(import_id, client_id, filename, str(exc))
        return err(str(exc))

    result["dryRun"] = dry_run
    result["importId"] = import_id
    return jsonify(result)


def _parse_spreadsheet(content, ext, header_row=None):
    if _looks_like_xlsx(content):
        return _parse_xlsx(content, header_row=header_row)
    if ext == ".csv":
        return _parse_csv(content, header_row=header_row)
    if ext == ".xlsx":
        return _parse_xlsx(content, header_row=header_row)
    if _looks_like_text_spreadsheet(content):
        return _parse_delimited_text(content, header_row=header_row)
    return _parse_xls(content, header_row=header_row)


def _looks_like_xlsx(content):
    return bytes(content[:4]) == b"PK\x03\x04"


def _looks_like_text_spreadsheet(content):
    sample = bytes(content[:2048]).lstrip()
    if not sample:
        return False
    return not sample.startswith(b"\xd0\xcf\x11\xe0")


def _parse_delimited_text(content, header_row=None):
    text = _decode_text_spreadsheet(content)
    sample = text[:2048]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    summary = _summarize_rows(rows, header_row=header_row)
    return {
        "sheetNames": [],
        "selectedSheet": "",
        **summary,
    }


def _decode_text_spreadsheet(content):
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8-sig", errors="replace")


@api.get("/imports")
def list_imports():
    limit = request.args.get("limit", "25")
    params = {
        "sort[0][field]": C.F_IMPORT_STARTED,
        "sort[0][direction]": "desc",
        "maxRecords": limit,
    }
    data = airtable.list_records(C.IMPORTS_TABLE, params=params, by_field_id=False)
    records = _filter_by_client_field(data.get("records", []), C.F_IMPORT_CLIENT)
    return jsonify({"records": _shape_import_records(records)})


@api.get("/imports/client-status")
def import_client_status():
    client_id = (request.args.get("clientId") or "").strip()
    if not client_id:
        return err("clientId is required")
    if not _client_permitted(client_id):
        return _forbidden()
    records = _list_all_records(C.IMPORTS_TABLE)
    records = [
        record for record in records
        if client_id in (record.get("fields", {}).get(C.F_IMPORT_CLIENT, []) or [])
    ]
    return jsonify({"clientId": client_id, "hasImports": bool(records), "importCount": len(records)})


@api.get("/imports/<record_id>")
def get_import(record_id):
    record = airtable.get_record(C.IMPORTS_TABLE, record_id, by_field_id=False)
    if not _client_ids_permitted(record.get("fields", {}).get(C.F_IMPORT_CLIENT, [])):
        return _forbidden()
    return jsonify({"record": _shape_import_records([record])[0]})


def _name_map(table, name_field):
    data = airtable.list_records(table, by_field_id=False)
    return {record.get("id"): record.get("fields", {}).get(name_field, "") for record in data.get("records", [])}


def _shape_import_records(records):
    client_names = _name_map(C.CLIENTS_TABLE, C.F_CLIENT_NAME)
    user_names = _name_map(C.USERS_TABLE, C.F_USER_NAME)
    return [_shape_import_record(record, client_names, user_names) for record in records]


def _shape_import_record(record, client_names, user_names):
    fields = record.get("fields", {})
    client_ids = fields.get(C.F_IMPORT_CLIENT, []) or []
    user_ids = fields.get(C.F_IMPORT_USER, []) or []
    return {
        "id": record.get("id"),
        "name": fields.get(C.F_IMPORT_NAME, ""),
        "clientIds": client_ids,
        "client": client_names.get(client_ids[0], "") if client_ids else "",
        "userIds": user_ids,
        "user": user_names.get(user_ids[0], "") if user_ids else "",
        "file": fields.get(C.F_IMPORT_FILE, ""),
        "type": fields.get(C.F_IMPORT_TYPE, ""),
        "status": fields.get(C.F_IMPORT_STATUS, ""),
        "started": fields.get(C.F_IMPORT_STARTED, ""),
        "finished": fields.get(C.F_IMPORT_FINISHED, ""),
        "rows": fields.get(C.F_IMPORT_ROWS, 0),
        "itemsCreated": fields.get(C.F_IMPORT_ITEMS_CREATED, 0),
        "itemsUpdated": fields.get(C.F_IMPORT_ITEMS_UPDATED, 0),
        "rowsSkipped": fields.get(C.F_IMPORT_ROWS_SKIPPED, 0),
        "errors": fields.get(C.F_IMPORT_ERRORS, 0),
        "warnings": fields.get(C.F_IMPORT_WARNINGS, 0),
        "details": fields.get(C.F_IMPORT_DETAILS, ""),
    }


def _parse_csv(content, header_row=None):
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    summary = _summarize_rows(rows, header_row=header_row)
    return {
        "sheetNames": [],
        "selectedSheet": "",
        **summary,
    }


def _parse_xlsx(content, header_row=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _parse_xlsx_xml(content, header_row=header_row)

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise ValueError("No sheets were found in the workbook.")
    selected = workbook[sheet_names[0]]
    rows = [[_spreadsheet_cell_value(cell) for cell in row] for row in selected.iter_rows()]
    summary = _summarize_rows(rows, header_row=header_row)
    return {
        "sheetNames": sheet_names,
        "selectedSheet": selected.title,
        **summary,
    }


def _parse_xls(content, header_row=None):
    try:
        import xlrd
    except ImportError:
        raise ValueError("Unreadable file. Legacy .xls support requires xlrd.")

    workbook = xlrd.open_workbook(file_contents=content)
    sheet_names = workbook.sheet_names()
    if not sheet_names:
        raise ValueError("No sheets were found in the workbook.")
    selected = workbook.sheet_by_index(0)
    rows = [
        [_spreadsheet_value(selected.cell_value(row_index, col_index)) for col_index in range(selected.ncols)]
        for row_index in range(selected.nrows)
    ]
    summary = _summarize_rows(rows, header_row=header_row)
    return {
        "sheetNames": sheet_names,
        "selectedSheet": selected.name,
        **summary,
    }


def _parse_xlsx_xml(content, header_row=None):
    with zipfile.ZipFile(io.BytesIO(content)) as workbook:
        shared_strings = _xlsx_shared_strings(workbook)
        sheets = _xlsx_sheets(workbook)
        if not sheets:
            raise ValueError("No sheets were found in the workbook.")

        selected = sheets[0]
        rows = _xlsx_sheet_rows(workbook, selected["path"], shared_strings)
        summary = _summarize_rows(rows, header_row=header_row)
        return {
            "sheetNames": [sheet["name"] for sheet in sheets],
            "selectedSheet": selected["name"],
            **summary,
        }


def _spreadsheet_cell_value(cell):
    return _spreadsheet_value(cell.value, getattr(cell, "number_format", ""))


def _spreadsheet_value(value, number_format=""):
    if value is None:
        return ""
    padded = _zero_padded_numeric_text(value, number_format)
    if padded is not None:
        return padded
    return str(value)


def _zero_padded_numeric_text(value, number_format):
    width = _zero_padded_number_width(number_format)
    if not width or isinstance(value, bool):
        return None
    if isinstance(value, int):
        digits = str(value)
    elif isinstance(value, float) and value.is_integer():
        digits = str(int(value))
    else:
        return None
    if digits.startswith("-"):
        return None
    return digits.zfill(width)


def _zero_padded_number_width(number_format):
    section = str(number_format or "").split(";", 1)[0]
    section = re.sub(r'"[^"]*"', "", section)
    section = re.sub(r"\\.", "", section)
    section = re.sub(r"\[[^\]]+\]", "", section)
    return len(section) if re.fullmatch(r"0+", section or "") else 0


def _xlsx_shared_strings(workbook):
    try:
        root = ElementTree.fromstring(workbook.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    strings = []
    for item in root.findall("main:si", XLSX_MAIN_NS):
        text = "".join(node.text or "" for node in item.findall(".//main:t", XLSX_MAIN_NS))
        strings.append(text)
    return strings


def _xlsx_sheets(workbook):
    workbook_root = ElementTree.fromstring(workbook.read("xl/workbook.xml"))
    rels_root = ElementTree.fromstring(workbook.read("xl/_rels/workbook.xml.rels"))
    rels = {}
    for rel in rels_root.findall("rel:Relationship", XLSX_RELS_NS):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        if rel_id and target:
            target = target.lstrip("/")
            rels[rel_id] = target if target.startswith("xl/") else "xl/" + target

    sheets = []
    for sheet in workbook_root.findall("main:sheets/main:sheet", XLSX_MAIN_NS):
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        path = rels.get(rel_id)
        if path:
            sheets.append({"name": sheet.attrib.get("name", "Sheet"), "path": path})
    return sheets


def _xlsx_sheet_rows(workbook, path, shared_strings):
    root = ElementTree.fromstring(workbook.read(path))
    rows = []
    for row in root.findall(".//main:sheetData/main:row", XLSX_MAIN_NS):
        try:
            row_number = int(row.attrib.get("r") or len(rows) + 1)
        except (TypeError, ValueError):
            row_number = len(rows) + 1
        while len(rows) < row_number:
            rows.append([])
        values = []
        for cell in row.findall("main:c", XLSX_MAIN_NS):
            index = _xlsx_cell_index(cell.attrib.get("r", ""))
            while len(values) <= index:
                values.append("")
            values[index] = _xlsx_cell_value(cell, shared_strings)
        rows[row_number - 1] = values
    return rows


def _xlsx_cell_index(reference):
    match = re.match(r"([A-Z]+)", reference or "")
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - ord("A") + 1
    return index - 1


def _xlsx_cell_value(cell, shared_strings):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", XLSX_MAIN_NS))

    value_node = cell.find("main:v", XLSX_MAIN_NS)
    value = value_node.text if value_node is not None else ""
    if cell_type == "s":
        try:
            return shared_strings[int(value)]
        except (ValueError, IndexError):
            return ""
    return value or ""


def _summarize_rows(rows, header_row=None):
    normalized = [[("" if value is None else str(value)) for value in row] for row in rows]
    if not normalized or not any(_row_has_value(row) for row in normalized):
        raise ValueError("No usable rows were found in the spreadsheet.")

    if header_row in (None, "", "auto"):
        candidates = [
            (index, row)
            for index, row in enumerate(normalized[:20])
            if _row_has_value(row)
        ]
        header_index = max(
            candidates,
            key=lambda candidate: (
                sum(1 for value in candidate[1] if _cell_has_value(value)),
                -candidate[0],
            ),
        )[0]
        selected_header_row = header_index + 1
    else:
        try:
            selected_header_row = int(header_row)
        except (TypeError, ValueError):
            raise ValueError("Header row must be a positive row number.")
        if selected_header_row < 1 or selected_header_row > len(normalized):
            raise ValueError(f"Header row {selected_header_row} is outside the spreadsheet.")
        header_index = selected_header_row - 1

    header = normalized[header_index]
    data_rows = [row for row in normalized[header_index + 1:] if _row_has_value(row)]
    if not data_rows:
        raise ValueError("No usable data rows were found after the header row.")

    width = max(len(header), *(len(row) for row in data_rows))
    header = _pad_row(header, width)
    data_rows = [_pad_row(row, width) for row in data_rows]
    keep_indexes = [
        index for index in range(width)
        if _cell_has_value(header[index]) or any(_cell_has_value(row[index]) for row in data_rows)
    ]
    header = [header[index] for index in keep_indexes]
    data_rows = [[row[index] for index in keep_indexes] for row in data_rows]

    if not header:
        raise ValueError("No usable columns were found in the spreadsheet.")

    return {
        "rowCount": len(data_rows),
        "headerRow": selected_header_row,
        "columnHeaders": header,
        "previewRows": data_rows[:10],
        "rows": data_rows,
    }


def _pad_row(row, width):
    return row + [""] * (width - len(row))


def _row_has_value(row):
    return any(_cell_has_value(value) for value in row)


def _cell_has_value(value):
    return str(value or "").strip() != ""


def _client_record(client_id):
    return airtable.get_record(C.CLIENTS_TABLE, client_id, by_field_id=False)


def _client_name(client_record):
    return client_record.get("fields", {}).get(C.F_CLIENT_NAME, "")


def _intake_destination_field_map():
    return {
        "Product Name": (C.PRODUCTS_TABLE, C.F_ITEM_NAME),
        "Identifier": (C.PRODUCTS_TABLE, C.F_ITEM_IDENTIFIER),
        "UPC": (C.PRODUCTS_TABLE, C.F_ITEM_UPC),
        "CVID": (C.PRODUCTS_TABLE, C.F_ITEM_CVID),
        "Brand Prefix": (C.PRODUCTS_TABLE, C.F_ITEM_BRAND_PREFIX),
        "Request Type": (C.PRODUCTS_TABLE, C.F_ITEM_REQUEST_TYPE),
        "Project Status": (C.PRODUCTS_TABLE, C.F_ITEM_PROJECT_STATUS),
        "WKFT Job Number": (C.PRODUCTS_TABLE, C.F_ITEM_WKFT_JOB_NUMBER),
        "Mbox Number": (C.PRODUCTS_TABLE, C.F_ITEM_MBOX_NUMBER),
        "Product Type": (C.PRODUCTS_TABLE, C.F_ITEM_PRODUCT_TYPE),
        "File Name Description": (C.PRODUCTS_TABLE, C.F_ITEM_FILE_NAME_DESCRIPTION),
        # Sheets in the wild still carry the old headings.
        "Product Description": (C.PRODUCTS_TABLE, C.F_ITEM_FILE_NAME_DESCRIPTION),
        "Prod Descrip": (C.PRODUCTS_TABLE, C.F_ITEM_FILE_NAME_DESCRIPTION),
        "Link to Prepro/Overlays": (C.PRODUCTS_TABLE, C.F_ITEM_PREPRO_OVERLAYS),
        "Ecomm Photo Notes": (C.PRODUCTS_TABLE, C.F_ITEM_ECOMM_PHOTO_NOTES),
        "Path to Art": (C.PRODUCTS_TABLE, C.F_ITEM_PATH_TO_ART),
        "Product or File Name": (C.PRODUCTS_TABLE, C.F_ITEM_PRODUCT),
        "Product Job Number": (C.PRODUCTS_TABLE, C.F_ITEM_JOB_NUMBER),
        "Master or Variant": (C.PRODUCTS_TABLE, C.F_ITEM_MASTER_VARIANT),
        "Pickup Job Number": (C.PRODUCTS_TABLE, C.F_ITEM_PICKUP_JOB_NUMBER),
        "Brand": (C.PRODUCTS_TABLE, C.F_ITEM_BRAND),
        "Notes": (C.PRODUCTS_TABLE, C.F_ITEM_NOTES),
        "Reference Data": (C.PRODUCTS_TABLE, C.F_ITEM_REFERENCE_DATA),
    }


def _airtable_field_descriptions():
    if not airtable.is_configured:
        return {}
    url = f"https://api.airtable.com/v0/meta/bases/{C.AIRTABLE_BASE_ID}/tables"
    response = requests.get(url, headers=airtable._headers(), timeout=20)
    response.raise_for_status()
    descriptions = {}
    destinations = _intake_destination_field_map()
    for table in response.json().get("tables", []):
        table_name = table.get("name")
        for field in table.get("fields", []):
            field_name = field.get("name")
            for destination, (target_table, target_field) in destinations.items():
                if table_name == target_table and field_name == target_field and field.get("description"):
                    descriptions[destination] = field["description"]
    return descriptions


@api.get("/intake/mapping-targets")
def intake_mapping_targets():
    targets = list(INTAKE_FALLBACK_DESCRIPTIONS)
    try:
        airtable_descriptions = _airtable_field_descriptions()
    except requests.HTTPError as error:
        return airtable_err(error)
    except Exception:
        airtable_descriptions = {}

    return jsonify({
        "targets": [
            {
                "target": target,
                "description": airtable_descriptions.get(target) or INTAKE_FALLBACK_DESCRIPTIONS[target],
                "descriptionSource": "airtable" if target in airtable_descriptions else "local",
            }
            for target in targets
        ],
    })


def _mapping_for_client(client_name):
    mapping = INTAKE_MAPPINGS.get((client_name or "").strip().lower())
    if not mapping:
        raise ValueError(f"No spreadsheet mapping exists for {client_name or 'this Client'}.")
    return mapping


def _row_dict(headers, values):
    return {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}


def _source_value(row, field):
    return str(row.get(field, "") or "").strip() if field else ""


def _mapped_value(row, mapping, key):
    field = mapping.get(key)
    value = _source_value(row, field)
    if value:
        return value
    fallback = mapping.get(f"{key}_fallback")
    return _source_value(row, fallback)


def _mapped_notes(row, mapping):
    notes = []
    for field in mapping.get("notes", []):
        value = _source_value(row, field)
        if value:
            notes.append(value)
    return "\n".join(notes)


def _mapped_reference_data(row, mapping, headers=None):
    reference_data = {}
    for field in headers or []:
        value = _source_value(row, field)
        if value:
            reference_data[str(field)] = value
    for field in mapping.get("reference_data", []):
        value = _source_value(row, field)
        if value:
            reference_data[str(field)] = value
    return reference_data


def _normalize_reference_data_value(raw_value):
    if isinstance(raw_value, dict):
        normalized = {}
        for key, value in raw_value.items():
            clean_key = str(key or "").strip()
            normalized_value = _normalize_reference_data_value(value)
            if clean_key and normalized_value not in ("", {}, []):
                normalized[clean_key] = normalized_value
        return normalized
    if isinstance(raw_value, list):
        normalized = [_normalize_reference_data_value(value) for value in raw_value]
        return [value for value in normalized if value not in ("", {}, [])]
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return raw_value
    return str(raw_value or "").strip()


def _normalize_reference_data(value):
    if not isinstance(value, dict):
        return {}
    normalized = {}
    for key, raw_value in value.items():
        clean_key = str(key or "").strip()
        clean_value = _normalize_reference_data_value(raw_value)
        if clean_key and clean_value not in ("", {}, []):
            normalized[clean_key] = clean_value
    return normalized


def _parse_reference_data(raw):
    if not raw:
        return {}
    if isinstance(raw, dict):
        return _normalize_reference_data(raw)
    text = str(raw or "").strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except (TypeError, ValueError):
        print(f"Malformed Reference Data JSON: {text[:200]}")
        return {"Raw": text}
    if not isinstance(parsed, dict):
        print(f"Malformed Reference Data JSON: {text[:200]}")
        return {"Raw": text}
    return _normalize_reference_data(parsed)


def _reference_data_json(value):
    normalized = _normalize_reference_data(value)
    if not normalized:
        return ""
    return json.dumps(normalized, ensure_ascii=False, sort_keys=True)


def build_source_snapshot(source_row, match_method, actionable_reason):
    row = source_row if isinstance(source_row, dict) else {}
    source_data = row.get("sourceData") if isinstance(row.get("sourceData"), dict) else row
    product_name = str(source_data.get("Product Name") or source_data.get("productName") or "").strip()
    upc = str(source_data.get("UPC") or source_data.get("upc") or "").strip()
    return {
        "client": str(row.get("client") or row.get("clientName") or "Topco").strip() or "Topco",
        "source": str(row.get("source") or row.get("sourceName") or "TOPCO (MARKS) PROJECTS").strip() or "TOPCO (MARKS) PROJECTS",
        "sheetTab": str(row.get("sheetTab") or row.get("sheetName") or C.TOPCO_SOURCE_SHEET_TAB).strip() or C.TOPCO_SOURCE_SHEET_TAB,
        "sourceRowNumber": row.get("sourceRowNumber") or row.get("rowNumber") or "",
        "sourceCheckedAt": str(row.get("sourceCheckedAt") or _now_iso()).strip(),
        "matchMethod": str(match_method or row.get("matchMethod") or "").strip(),
        "actionableReason": str(actionable_reason or "").strip(),
        "sourceIdentity": {
            "productName": product_name,
            "upc": upc,
        },
    }


def merge_product_source_snapshot(product_reference_data, snapshot):
    reference_data = _parse_reference_data(product_reference_data)
    if not isinstance(snapshot, dict) or not snapshot:
        return reference_data
    reference_data["_sourceSnapshot"] = snapshot
    return _normalize_reference_data(reference_data)


def _mapping_from_ui_mapping(ui_mapping):
    if not isinstance(ui_mapping, dict):
        raise ValueError("Invalid column mapping.")

    mapping = {"notes": [], "reference_data": []}
    target_keys = {
        "Item Name": "item_name",
        "Product Name": "item_name",
        "Identifier": "id",
        "UPC": "upc",
        "CVID": "cvid",
        "Brand Prefix": "brand_prefix",
        "Request Type": "request_type",
        "Project Status": "project_status",
        "WKFT Job Number": "wkft_job_number",
        "Mbox Number": "mbox_number",
        "Product Type": "product_type",
        "File Name Description": "file_name_description",
        # Sheets in the wild still carry the old headings.
        "Product Description": "file_name_description",
        "Prod Descrip": "file_name_description",
        "Link to Prepro/Overlays": "prepro_overlays",
        "Ecomm Photo Notes": "ecomm_photo_notes",
        "Path to Art": "path_to_art",
        "Product or File Name": "product",
        "Product/File Name": "product",
        "Description": "description",
        "Item Job Number": "item_job_number",
        "Product Job Number": "item_job_number",
        "Brand": "brand",
        "Job Number": "item_job_number",
        "Parent Job Number": "parent_job_number",
        "Master or Variant": "master_or_variant",
        "Pickup Job Number": "pickup_job_number",
        "Due Date": "due",
        "Job Name": "job_name",
        "Jobs.Job": "job_name",
        "Jobs.Job Number": "parent_job_number",
        "Jobs.Parent Job Number": "parent_job_number",
        "Jobs.Due": "due",
        "Items.Item": "item_name",
        "Products.Product Name": "item_name",
        "Items.Identifier": "id",
        "Items.Product or File Name": "product",
        "Items.Product/File Name": "product",
        "Items.Product Name": "product",
        "Items.Description": "description",
        "Items.Item Job Number": "item_job_number",
        "Products.Product Job Number": "item_job_number",
        "Items.Job Number": "item_job_number",
        "Items.Master or Variant": "master_or_variant",
        "Items.Pickup Job Number": "pickup_job_number",
        "Items.Brand": "brand",
        "Items.Category": "category",
        "Job": "job_name",
        "Job Name": "job_name",
        "Identifier": "id",
        "ID": "id",
        "Product or File Name": "product",
        "Product/File Name": "product",
        "Description": "description",
        "Item Job Number": "item_job_number",
        "Product Job Number": "item_job_number",
        "Job Number": "item_job_number",
        "Master or Variant": "master_or_variant",
        "Pickup Job Number": "pickup_job_number",
        "Brand": "brand",
        "Category": "category",
    }

    target_mapping = ui_mapping.get("__targetMapping")
    if isinstance(target_mapping, dict):
        for target, source in target_mapping.items():
            source_name = str(source or "").strip()
            target_name = str(target or "").strip()
            if not source_name:
                continue
            if target_name in {"Notes", "Items.Notes", "Products.Notes"}:
                mapping["notes"].append(source_name)
                continue
            key = target_keys.get(target_name)
            if key and key not in {"notes", "job_notes"}:
                mapping[key] = source_name

    for source, target in ui_mapping.items():
        if source in {"__targetMapping", "__singleJobName", "__existingJobId", "__existingJobName", "__jobGroupField"}:
            continue
        source_name = str(source or "").strip()
        target_name = str(target or "").strip()
        if not source_name or target_name == "Ignore":
            continue
        if target_name in {"Items.Notes", "Products.Notes", "Notes"}:
            mapping["notes"].append(source_name)
            continue
        if target_name == "Reference Data":
            mapping["reference_data"].append(source_name)
            continue
        key = target_keys.get(target_name)
        if key and key not in mapping:
            mapping[key] = source_name

    return mapping


def _normalize_master_or_variant(value):
    cleaned = str(value or "").strip()
    if not cleaned:
        return ""
    aliases = {
        "master": "Master",
        "m": "Master",
        "variant": "Variant",
        "v": "Variant",
    }
    return aliases.get(cleaned.lower(), cleaned if cleaned in {"Master", "Variant"} else "")


def _normalize_item_job_number(value):
    return str(value or "").strip()


def _normalize_product_type(value):
    text = str(value or "").strip().strip('"').strip("'").strip()
    if not text:
        return ""
    if text.casefold() in {"refridgeration req", "refrigeration req"}:
        # The active Airtable choice is still misspelled; write its exact value
        # until the field option can be corrected in Airtable.
        return "Refridgeration Req"
    return text


PRODUCT_REQUEST_TYPE_OPTIONS = [
    "Ecomm only",
    "Pack only",
    "Thr3d only",
    "Pack & Thr3d",
    "Ecomm & Pack",
]


PRODUCT_TYPECAST_FIELDS = {
    C.F_ITEM_MASTER_VARIANT,
    C.F_ITEM_PRODUCT_TYPE,
    C.F_ITEM_REQUEST_TYPE,
}


def _normalize_product_request_type(value):
    text = str(value or "").strip().strip('"').strip("'").strip()
    if not text:
        return ""
    compact = re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()
    aliases = {
        "ecomm": "Ecomm only",
        "ecomm only": "Ecomm only",
        "ecommerce": "Ecomm only",
        "ecommerce only": "Ecomm only",
        "pack": "Pack only",
        "pack only": "Pack only",
        "packaging": "Pack only",
        "packaging only": "Pack only",
        "thr3d": "Thr3d only",
        "thr3d only": "Thr3d only",
        "threed": "Thr3d only",
        "threed only": "Thr3d only",
        "3d": "Thr3d only",
        "3d only": "Thr3d only",
        "pack thr3d": "Pack & Thr3d",
        "pack and thr3d": "Pack & Thr3d",
        "packaging thr3d": "Pack & Thr3d",
        "packaging and thr3d": "Pack & Thr3d",
        "packaging threed": "Pack & Thr3d",
        "packaging and threed": "Pack & Thr3d",
        "packaging 3d": "Pack & Thr3d",
        "packaging and 3d": "Pack & Thr3d",
        "thr3d pack": "Pack & Thr3d",
        "thr3d and pack": "Pack & Thr3d",
        "ecomm pack": "Ecomm & Pack",
        "ecomm and pack": "Ecomm & Pack",
        "ecommerce pack": "Ecomm & Pack",
        "ecommerce and pack": "Ecomm & Pack",
        "pack ecomm": "Ecomm & Pack",
        "pack and ecomm": "Ecomm & Pack",
    }
    return aliases.get(compact, text if text in PRODUCT_REQUEST_TYPE_OPTIONS else text)


def _normalize_description(value):
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def _readable_item_name(brand, product, identifier):
    if brand and product:
        return f"{brand} {product}"
    return product or identifier


def _received_signal(row):
    for key in ("Received", "Product Received", "Merch Received"):
        value = _source_value(row, key).lower()
        if value in {"yes", "true", "received", "y", "1"}:
            return True
        if value in {"no", "false", "n", "0"}:
            return False
    return None


def _default_item_status(row):
    return "Pending"




def _validate_identifier(identifier, code_type, label="Primary Match Key"):
    return _validate_item_identifier(identifier, code_type, label)


def _normalized_required_fields(client_config):
    return {
        "Identifier" if field == "ID" else "Product or File Name" if field in {"Product Name", "Product/File Name"} else str(field or "").strip()
        for field in (client_config or {}).get("requiredToShoot", [])
        if str(field or "").strip()
    }






def _existing_item_for(existing_items, identifier):
    """Look up by what the sheet said, then by the comparable form of it."""
    if not identifier:
        return None
    raw = str(identifier).strip()
    return existing_items.get(raw) or existing_items.get(normalized_identifier(raw))


def _existing_items_by_identifier(client_id):
    """Products this client already has, keyed for lookup by a sheet row.

    Keyed the same way the merge keys them, and over the same two fields. Indexing
    `Identifier` alone missed every Product created from a Structure Form, because
    those carry a UPC and no Identifier - so importing the sheet made a second
    record for a SKU that already had one. Exact string matching missed them again
    whenever the spreadsheet had stripped a leading zero.

    Both the raw and normalized forms are keyed, because callers look up with
    whatever the sheet gave them.
    """
    records = _filter_by_client_field(_list_all_records(C.PRODUCTS_TABLE), C.F_ITEM_CLIENT)
    items = {}
    for record in records:
        fields = record.get("fields", {})
        if client_id not in (fields.get(C.F_ITEM_CLIENT, []) or []):
            continue
        for candidate in (fields.get(C.F_ITEM_IDENTIFIER), fields.get(C.F_ITEM_UPC)):
            raw = str(candidate or "").strip()
            if raw:
                items.setdefault(raw, record)
            key = normalized_identifier(candidate)
            if key:
                items.setdefault(key, record)
    return items


def _build_intake_plan(client_id, filename, parsed, mapping=None):
    client_record = _client_record(client_id)
    if not _client_permitted(client_id):
        raise ValueError("You do not have access to that Client.")
    client_name = _client_name(client_record)
    client_config = _shape_client(client_record)
    code_type = client_config.get("codeType", "")
    identifier_label = _identifier_label(client_config)
    required_photo_fields = _normalized_required_fields(client_config)
    mapping = mapping or _mapping_for_client(client_name)
    headers = parsed.get("columnHeaders", [])
    rows = [_row_dict(headers, values) for values in parsed.get("rows", [])]
    existing_items = _existing_items_by_identifier(client_id)
    seen_ids = {}
    row_results = []
    items_to_create = 0
    items_to_update = 0
    warning_count = 0
    error_count = 0

    try:
        source_first_row_number = int(parsed.get("headerRow") or 1) + 1
    except (TypeError, ValueError):
        source_first_row_number = 2

    for index, row in enumerate(rows, start=2):
        source_row_number = source_first_row_number + index - 2
        identifier = _mapped_value(row, mapping, "id") or _mapped_value(row, mapping, "upc")
        upc = _mapped_value(row, mapping, "upc") or identifier
        cvid = _mapped_value(row, mapping, "cvid")
        brand_prefix = _mapped_value(row, mapping, "brand_prefix")
        request_type = _normalize_product_request_type(_mapped_value(row, mapping, "request_type"))
        project_status = _mapped_value(row, mapping, "project_status")
        wkft_job_number = _mapped_value(row, mapping, "wkft_job_number")
        mbox_number = _mapped_value(row, mapping, "mbox_number")
        product_type = _normalize_product_type(_mapped_value(row, mapping, "product_type"))
        file_name_description = _mapped_value(row, mapping, "file_name_description")
        prepro_overlays = _mapped_value(row, mapping, "prepro_overlays")
        ecomm_photo_notes = _mapped_value(row, mapping, "ecomm_photo_notes")
        path_to_art = _mapped_value(row, mapping, "path_to_art")
        product = _mapped_value(row, mapping, "product")
        product_source = _source_value(row, mapping.get("product"))
        item_job_number = _normalize_item_job_number(_mapped_value(row, mapping, "item_job_number"))
        description = _normalize_description(_mapped_value(row, mapping, "description"))
        brand = _mapped_value(row, mapping, "brand")
        category = _mapped_value(row, mapping, "category")
        master_or_variant = _normalize_master_or_variant(_mapped_value(row, mapping, "master_or_variant"))
        pickup_job_number = _normalize_item_job_number(_mapped_value(row, mapping, "pickup_job_number"))
        notes = _mapped_notes(row, mapping)
        reference_data = _mapped_reference_data(row, mapping, headers)
        source_snapshot = None
        if client_name.strip().casefold() == "topco":
            source_snapshot = build_source_snapshot(
                {
                    "client": client_name,
                    "sourceData": reference_data,
                    "rowNumber": source_row_number,
                },
                match_method="Import",
                actionable_reason="import_commit",
            )
        item_name = _mapped_value(row, mapping, "item_name")
        problems = []
        warnings = []

        if identifier:
            validation_error = _validate_identifier(identifier, code_type, identifier_label)
            if validation_error:
                problems.append(validation_error)
            if identifier in seen_ids:
                warnings.append(f"Duplicate {identifier_label} also appears on row {seen_ids[identifier]}")
            else:
                seen_ids[identifier] = index
        if "Product or File Name" in required_photo_fields and not product_source:
            problems.append("Missing Product or File Name")
        if "Brand" in required_photo_fields and not brand:
            problems.append("Missing Brand")

        existing_item = _existing_item_for(existing_items, identifier)
        action = "skip" if problems else ("update" if existing_item else "create")
        if action == "create":
            items_to_create += 1
        elif action == "update":
            items_to_update += 1
        warning_count += len(warnings)
        error_count += len(problems)
        row_results.append({
            "rowNumber": index,
            "action": action,
            "id": identifier,
            "upc": upc,
            "cvid": cvid,
            "brandPrefix": brand_prefix,
            "requestType": request_type,
            "projectStatus": project_status,
            "wkftJobNumber": wkft_job_number,
            "mboxNumber": mbox_number,
            "productType": product_type,
            "fileNameDescription": file_name_description,
            "preproOverlays": prepro_overlays,
            "ecommPhotoNotes": ecomm_photo_notes,
            "pathToArt": path_to_art,
            "itemName": item_name or _readable_item_name(brand, product, identifier),
            "product": product,
            "itemJobNumber": item_job_number,
            "description": description,
            "brand": brand,
            "category": category,
            "masterOrVariant": master_or_variant,
            "pickupJobNumber": pickup_job_number,
            "notes": notes,
            "referenceData": reference_data,
            "sourceSnapshot": source_snapshot,
            "status": _default_item_status(row),
            "existingItemId": existing_item.get("id") if existing_item else None,
            "errors": problems,
            "warnings": warnings,
        })

    return {
        "fileName": filename,
        "clientId": client_id,
        "clientName": client_name,
        "codeType": code_type,
        "identifierLabel": identifier_label,
        "primaryMatchKeyLabel": identifier_label,
        "sheetNames": parsed.get("sheetNames", []),
        "selectedSheet": parsed.get("selectedSheet", ""),
        "totalRows": len(rows),
        "itemsToCreate": items_to_create,
        "itemsToUpdate": items_to_update,
        "warningCount": warning_count,
        "errorCount": error_count,
        "rows": row_results,
        "summary": {
            "itemsCreated": 0,
            "itemsUpdated": 0,
            "rowsSkipped": error_count and sum(1 for row in row_results if row["errors"]) or 0,
            "errors": error_count,
            "warnings": warning_count,
        },
    }


def _build_intake_plan_from_source_rows(client_id, filename, headers, source_rows, ui_mapping):
    parsed = {
        "columnHeaders": [str(header or "") for header in headers],
        "rows": source_rows,
        "sheetNames": [],
        "selectedSheet": "",
    }
    mapping = _mapping_from_ui_mapping(ui_mapping)
    return _build_intake_plan(client_id, filename, parsed, mapping=mapping)


def _build_intake_plan_from_mapped_rows(client_id, filename, rows):
    client_record = _client_record(client_id)
    client_name = _client_name(client_record)
    client_config = _shape_client(client_record)
    code_type = client_config.get("codeType", "")
    identifier_label = _identifier_label(client_config)
    required_photo_fields = _normalized_required_fields(client_config)
    existing_items = _existing_items_by_identifier(client_id)
    seen_ids = {}
    row_results = []
    items_to_create = 0
    items_to_update = 0
    warning_count = 0
    error_count = 0

    for index, source in enumerate(rows, start=1):
        identifier = str(source.get("id", "") or source.get("upc", "") or "").strip()
        upc = str(source.get("upc", "") or identifier).strip()
        cvid = str(source.get("cvid", "") or "").strip()
        brand_prefix = str(source.get("brandPrefix", "") or "").strip()
        request_type = _normalize_product_request_type(source.get("requestType"))
        project_status = str(source.get("projectStatus", "") or "").strip()
        wkft_job_number = str(source.get("wkftJobNumber", "") or "").strip()
        mbox_number = str(source.get("mboxNumber", "") or "").strip()
        product_type = _normalize_product_type(source.get("productType", ""))
        file_name_description = str(source.get("fileNameDescription", "") or source.get("productDescription", "") or "").strip()
        prepro_overlays = str(source.get("preproOverlays", "") or "").strip()
        ecomm_photo_notes = str(source.get("ecommPhotoNotes", "") or "").strip()
        path_to_art = str(source.get("pathToArt", "") or "").strip()
        product = str(source.get("product", "") or "").strip()
        item_job_number = _normalize_item_job_number(source.get("itemJobNumber"))
        description = _normalize_description(source.get("description"))
        brand = str(source.get("brand", "") or "").strip()
        category = str(source.get("category", "") or "").strip()
        master_or_variant = _normalize_master_or_variant(source.get("masterOrVariant"))
        pickup_job_number = _normalize_item_job_number(source.get("pickupJobNumber"))
        notes = str(source.get("notes", "") or "").strip()
        reference_data = _normalize_reference_data(source.get("referenceData") or {})
        source_snapshot = None
        if client_name.strip().casefold() == "topco":
            source_snapshot = build_source_snapshot(
                {
                    "client": client_name,
                    "sourceData": reference_data,
                    "rowNumber": source.get("rowNumber") or index,
                },
                match_method=source.get("sourceSnapshotMatchMethod") or "Import",
                actionable_reason=source.get("actionableReason") or "import_commit",
            )
        status = source.get("status") or "Pending"
        item_name = str(source.get("itemName", "") or "").strip()
        problems = []
        warnings = []

        if identifier:
            validation_error = _validate_identifier(identifier, code_type, identifier_label)
            if validation_error:
                problems.append(validation_error)
            if identifier in seen_ids:
                warnings.append(f"Duplicate {identifier_label} also appears on row {seen_ids[identifier]}")
            else:
                seen_ids[identifier] = index
        if "Product or File Name" in required_photo_fields and not product:
            problems.append("Missing Product or File Name")
        if "Brand" in required_photo_fields and not brand:
            problems.append("Missing Brand")

        existing_item = ({"id": source.get("existingItemId")} if source.get("existingItemId")
                         else _existing_item_for(existing_items, identifier))
        action = "skip" if problems else ("update" if existing_item else "create")
        if action == "create":
            items_to_create += 1
        elif action == "update":
            items_to_update += 1
        warning_count += len(warnings)
        error_count += len(problems)
        row_results.append({
            "rowNumber": source.get("rowNumber") or index,
            "action": action,
            "id": identifier,
            "upc": upc,
            "cvid": cvid,
            "brandPrefix": brand_prefix,
            "requestType": request_type,
            "projectStatus": project_status,
            "wkftJobNumber": wkft_job_number,
            "mboxNumber": mbox_number,
            "productType": product_type,
            "fileNameDescription": file_name_description,
            "preproOverlays": prepro_overlays,
            "ecommPhotoNotes": ecomm_photo_notes,
            "pathToArt": path_to_art,
            "itemName": item_name or _readable_item_name(brand, product, identifier),
            "product": product,
            "itemJobNumber": item_job_number,
            "description": description,
            "brand": brand,
            "category": category,
            "masterOrVariant": master_or_variant,
            "pickupJobNumber": pickup_job_number,
            "notes": notes,
            "referenceData": reference_data,
            "sourceSnapshot": source_snapshot,
            "clearBlankSourceFields": bool(source.get("clearBlankSourceFields")),
            "existingItemId": existing_item.get("id") if existing_item else None,
            "errors": problems,
            "warnings": warnings,
        })

    return {
        "fileName": filename,
        "clientId": client_id,
        "clientName": client_name,
        "codeType": code_type,
        "identifierLabel": identifier_label,
        "primaryMatchKeyLabel": identifier_label,
        "sheetNames": [],
        "selectedSheet": "",
        "totalRows": len(rows),
        "itemsToCreate": items_to_create,
        "itemsToUpdate": items_to_update,
        "warningCount": warning_count,
        "errorCount": error_count,
        "rows": row_results,
        "summary": {
            "itemsCreated": 0,
            "itemsUpdated": 0,
            "rowsSkipped": sum(1 for row in row_results if row["errors"]),
            "errors": error_count,
            "warnings": warning_count,
        },
    }




def _item_fields_from_row(client_id, row):
    fields = {
        C.F_ITEM_NAME: row["itemName"],
        C.F_ITEM_CLIENT: [client_id],
    }
    clear_blank_source_fields = bool(row.get("clearBlankSourceFields"))
    if C.F_ITEM_PRODUCT != C.F_ITEM_NAME and row.get("product"):
        fields[C.F_ITEM_PRODUCT] = row["product"]
    if row.get("id"):
        fields[C.F_ITEM_IDENTIFIER] = row["id"]
    if row.get("upc"):
        fields[C.F_ITEM_UPC] = row["upc"]
    if row.get("cvid"):
        fields[C.F_ITEM_CVID] = row["cvid"]
    if row.get("brandPrefix"):
        fields[C.F_ITEM_BRAND_PREFIX] = row["brandPrefix"]
    for key, field in {
        "requestType": C.F_ITEM_REQUEST_TYPE,
        "wkftJobNumber": C.F_ITEM_WKFT_JOB_NUMBER,
        "mboxNumber": C.F_ITEM_MBOX_NUMBER,
        "projectName": C.F_ITEM_PROJECT_NAME,
        "productType": C.F_ITEM_PRODUCT_TYPE,
        "fileNameDescription": C.F_ITEM_FILE_NAME_DESCRIPTION,
        "preproOverlays": C.F_ITEM_PREPRO_OVERLAYS,
        "ecommPhotoNotes": C.F_ITEM_ECOMM_PHOTO_NOTES,
        "pathToArt": C.F_ITEM_PATH_TO_ART,
    }.items():
        if row.get(key):
            if key == "productType":
                fields[field] = _normalize_product_type(row[key])
            elif key == "requestType":
                fields[field] = _normalize_product_request_type(row[key])
            else:
                fields[field] = row[key]
        elif clear_blank_source_fields and key == "requestType":
            fields[field] = None
    if row.get("brand"):
        fields[C.F_ITEM_BRAND] = row["brand"]
    if row.get("itemJobNumber"):
        fields[C.F_ITEM_JOB_NUMBER] = row["itemJobNumber"]
    if row.get("masterOrVariant"):
        fields[C.F_ITEM_MASTER_VARIANT] = row["masterOrVariant"]
    if row.get("pickupJobNumber"):
        fields[C.F_ITEM_PICKUP_JOB_NUMBER] = row["pickupJobNumber"]
    if row.get("category"):
        fields[C.F_ITEM_CATEGORY] = row["category"]
    if row.get("notes"):
        fields[C.F_ITEM_NOTES] = row["notes"]
    reference_data = row.get("referenceData")
    if row.get("sourceSnapshot"):
        reference_data = merge_product_source_snapshot(reference_data, row.get("sourceSnapshot"))
    reference_data = _reference_data_json(reference_data)
    if reference_data:
        fields[C.F_ITEM_REFERENCE_DATA] = reference_data
    return fields


def _execute_intake_plan(plan):
    client_id = plan["clientId"]
    summary = {
        "itemsCreated": 0,
        "itemsUpdated": 0,
        "rowsSkipped": 0,
        "errors": plan["errorCount"],
        "warnings": plan["warningCount"],
    }

    for row in plan["rows"]:
        if row["errors"]:
            summary["rowsSkipped"] += 1
            continue
        fields = _item_fields_from_row(client_id, row)
        if row.get("existingItemId"):
            previous = airtable.get_record(C.PRODUCTS_TABLE, row["existingItemId"], by_field_id=False)
            data = airtable.update_record(C.PRODUCTS_TABLE, row["existingItemId"], fields, by_field_id=False)
            _log_item_changes(data["id"], previous, data, fields)
            row["action"] = "updated"
            summary["itemsUpdated"] += 1
        else:
            data = airtable.create_record(C.PRODUCTS_TABLE, fields, by_field_id=False)
            row["existingItemId"] = data["id"]
            row["action"] = "created"
            _create_history_event("Item Created", item_ids=[data["id"]])
            summary["itemsCreated"] += 1

    plan["summary"] = summary
    return plan


# ── Locations ─────────────────────────────────────────────────────────────────

@api.get("/airtable/single-select-options")
def airtable_single_select_options():
    """Choices for one single-select field, read server-side.

    The browser used to call Airtable's metadata API directly, which meant
    shipping an Airtable token in the bundle for a carrier dropdown. The token
    belongs on the server; the browser only needs the resulting list.
    """
    table_name = (request.args.get("tableName") or "").strip()
    field_name = (request.args.get("fieldName") or "").strip()
    if not table_name or not field_name:
        return err("tableName and fieldName are required.")
    if not airtable.is_configured:
        return jsonify({"options": []})
    try:
        response = requests.get(
            f"https://api.airtable.com/v0/meta/bases/{C.AIRTABLE_BASE_ID}/tables",
            headers=airtable._headers(),
            timeout=20,
        )
        response.raise_for_status()
    except requests.HTTPError as error:
        return airtable_err(error)
    except requests.RequestException:
        current_app.logger.exception("Could not read Airtable table metadata")
        return jsonify({"options": []})

    table = next((t for t in response.json().get("tables", []) if t.get("name") == table_name), None)
    field = next((f for f in (table or {}).get("fields", []) if f.get("name") == field_name), None)
    choices = ((field or {}).get("options") or {}).get("choices") or []
    return jsonify({"options": [c.get("name", "") for c in choices if c.get("name")]})


@api.get("/locations")
def list_locations():
    data = airtable.list_records(
        C.LOCATIONS_TABLE,
        params={"sort[0][field]": C.F_LOCATION_NAME, "sort[0][direction]": "asc"},
        by_field_id=False,
    )
    records = _filter_locations(data.get("records", []))
    records = [_shape_location(r) for r in records]
    return jsonify({"records": records})


def _filter_locations(records):
    permissions = _permission_context()
    if permissions["all"]:
        return records
    permitted_item_ids = _permitted_record_ids(C.PRODUCTS_TABLE, C.F_ITEM_CLIENT)
    permitted_receipt_ids = _permitted_record_ids(C.SHIPMENTS_TABLE, C.F_RECEIPT_CLIENT)
    filtered = []
    for record in records:
        fields = record.get("fields", {})
        item_ids = set((fields.get("Products", []) or []) + (fields.get("Items", []) or []))
        receipt_ids = set((fields.get("Shipments", []) or []) + (fields.get("Receipts", []) or []))
        if item_ids & permitted_item_ids or receipt_ids & permitted_receipt_ids:
            filtered.append(record)
    return filtered


def _permitted_record_ids(table, client_field):
    data = airtable.list_records(table, by_field_id=False)
    records = _filter_by_client_field(data.get("records", []), client_field)
    return {record.get("id") for record in records}


def _shape_location(r):
    f = r.get("fields", {})
    return {
        "id": r["id"],
        "name": f.get(C.F_LOCATION_NAME, ""),
        "type": f.get(C.F_LOCATION_TYPE, ""),
        "active": f.get(C.F_LOCATION_ACTIVE, False),
        "notes": f.get(C.F_LOCATION_NOTES, ""),
    }


# ── Users ─────────────────────────────────────────────────────────────────────

@api.get("/users")
def list_users():
    management_error = _require_user_management()
    if management_error:
        return management_error
    data = airtable.list_records(
        C.USERS_TABLE,
        params={"sort[0][field]": C.F_USER_NAME, "sort[0][direction]": "asc"},
        by_field_id=False,
    )
    records_data = data.get("records", [])
    if not _is_admin():
        records_data = [record for record in records_data if not _record_is_admin_user(record)]
    records = [_shape_user(r) for r in records_data]
    return jsonify({"records": records})


def _shape_user(r):
    f = r.get("fields", {})
    return {
        "id": r["id"],
        "name": f.get(C.F_USER_NAME, ""),
        "firstName": f.get(C.F_USER_FIRST_NAME, ""),
        "lastName": f.get(C.F_USER_LAST_NAME, ""),
        "displayName": f.get(C.F_USER_DISPLAY_NAME, ""),
        "email": f.get(C.F_USER_EMAIL, ""),
        "role": f.get(C.F_USER_ROLE, ""),
        "active": f.get(C.F_USER_ACTIVE, False),
        "clientIds": f.get(C.F_USER_CLIENTS, []) or [],
        "allClients": f.get(C.F_USER_ALL_CLIENTS, False),
        "avatar": f.get(C.F_USER_AVATAR, ""),
        "hasPIN": bool(f.get(C.F_USER_PIN_HASH, "")),
    }


def _shape_login_user(r):
    user = _shape_user(r)
    return {
        "id": user["id"],
        "name": user["name"],
        "displayName": user["displayName"],
        "avatar": user["avatar"],
        "active": user["active"],
    }


def _user_display_name(user):
    if not user:
        return ""
    fields = user.get("fields", {}) if "fields" in user else {}
    shaped = user if "fields" not in user else _shape_user(user)
    return (
        shaped.get("displayName")
        or shaped.get("name")
        or " ".join(filter(None, [shaped.get("firstName"), shaped.get("lastName")]))
        or fields.get(C.F_USER_DISPLAY_NAME)
        or fields.get(C.F_USER_NAME)
        or ""
    )


def _user_initials(user):
    name = _user_display_name(user)
    parts = [part for part in re.split(r"\s+", name.strip()) if part]
    if not parts:
        return ""
    return (parts[0][0] + (parts[-1][0] if len(parts) > 1 else "")).upper()


def _record_from_shaped_user(user):
    return {
        "id": user.get("id", ""),
        "fields": {
            C.F_USER_NAME: user.get("name", ""),
            C.F_USER_FIRST_NAME: user.get("firstName", ""),
            C.F_USER_LAST_NAME: user.get("lastName", ""),
            C.F_USER_DISPLAY_NAME: user.get("displayName", ""),
            C.F_USER_EMAIL: user.get("email", ""),
            C.F_USER_ROLE: user.get("role", ""),
            C.F_USER_ACTIVE: user.get("active", False),
            C.F_USER_CLIENTS: user.get("clientIds", []) or [],
            C.F_USER_ALL_CLIENTS: user.get("allClients", False),
            C.F_USER_AVATAR: user.get("avatar", ""),
        },
    }


def _session_user():
    user = session.get(AUTH_SESSION_KEY)
    return user if isinstance(user, dict) and user.get("id") else None


def _set_session_user(record):
    user = _shape_user(record)
    session.clear()
    session[AUTH_SESSION_KEY] = user
    session.permanent = True
    return user


def _refresh_session_user():
    user = _session_user()
    if not user:
        return None
    try:
        record = airtable.get_record(C.USERS_TABLE, user["id"], by_field_id=False)
    except requests.HTTPError:
        session.clear()
        return None
    if not record.get("fields", {}).get(C.F_USER_ACTIVE, False):
        session.clear()
        return None
    return _set_session_user(record)


def _is_admin(user=None):
    user = user or _session_user()
    return (user or {}).get("role") in {"Admin", "Administrator"}


def _require_admin():
    if not _is_admin():
        return err("Administrator access required", 403)
    return None


def _is_admin_role_value(role):
    return str(role or "").strip() in {"Admin", "Administrator"}


def _require_user_management():
    if not _session_user():
        return err("Authentication required", 401)
    return None


def _record_is_admin_user(record):
    return _is_admin_role_value((record.get("fields", {}) if record else {}).get(C.F_USER_ROLE))


def _hash_pin(user_id, pin):
    return hashlib.sha256(f"{user_id}:{pin}".encode()).hexdigest()


# ── Auth ──────────────────────────────────────────────────────────────────────

@api.post("/auth/login")
def auth_login():
    body = request.get_json(silent=True) or {}
    user_id = (body.get("userId") or "").strip()
    pin = str(body.get("pin") or "").strip()
    if not user_id or not pin:
        return err("userId and pin are required")
    try:
        record = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
    except requests.HTTPError:
        return err("Invalid credentials", 401)
    fields = record.get("fields", {})
    if not fields.get(C.F_USER_ACTIVE, False):
        return err("Account is not active", 403)
    stored_hash = fields.get(C.F_USER_PIN_HASH, "")
    if stored_hash and _hash_pin(user_id, pin) != stored_hash:
        return err("Invalid credentials", 401)
    # If no PIN hash set, accept any PIN and save it as the new hash
    if not stored_hash:
        record = airtable.update_record(
            C.USERS_TABLE, user_id,
            {C.F_USER_PIN_HASH: _hash_pin(user_id, pin)},
            by_field_id=False,
        )
    return jsonify({"user": _set_session_user(record)})


@api.get("/auth/me")
def auth_me():
    user = _refresh_session_user()
    if not user:
        return err("Authentication required", 401)
    return jsonify({"user": user})


@api.put("/auth/me")
def update_auth_me():
    user = _session_user()
    if not user:
        return err("Authentication required", 401)
    body = request.get_json(silent=True) or {}
    fields = {}
    for key, const in [
        ("firstName", C.F_USER_FIRST_NAME),
        ("lastName", C.F_USER_LAST_NAME),
        ("displayName", C.F_USER_DISPLAY_NAME),
        ("email", C.F_USER_EMAIL),
        ("avatar", C.F_USER_AVATAR),
    ]:
        if key in body:
            fields[const] = (body[key] or "").strip()
    if body.get("pin"):
        fields[C.F_USER_PIN_HASH] = _hash_pin(user["id"], str(body["pin"]))
    try:
        record = airtable.update_record(C.USERS_TABLE, user["id"], fields, by_field_id=False) if fields else _record_from_shaped_user(user)
    except requests.HTTPError as e:
        return airtable_err(e)
    return jsonify({"user": _set_session_user(record)})


@api.post("/auth/logout")
def auth_logout():
    session.clear()
    return jsonify({"ok": True})


@api.get("/auth/users")
def auth_users():
    data = airtable.list_records(
        C.USERS_TABLE,
        params={"sort[0][field]": C.F_USER_NAME, "sort[0][direction]": "asc"},
        by_field_id=False,
    )
    records = [_shape_login_user(r) for r in data.get("records", [])]
    return jsonify({"records": records})


@api.post("/users")
def create_user():
    management_error = _require_user_management()
    if management_error:
        return management_error
    body = request.get_json(silent=True) or {}
    if not _is_admin() and _is_admin_role_value(body.get("role")):
        return err("Only administrators can create administrator users.", 403)
    name = (body.get("name") or "").strip()
    if not name:
        return err("name is required")
    fields = {
        C.F_USER_NAME: name,
        C.F_USER_ACTIVE: True,
    }
    for key, const in [
        ("firstName", C.F_USER_FIRST_NAME),
        ("lastName", C.F_USER_LAST_NAME),
        ("displayName", C.F_USER_DISPLAY_NAME),
        ("email", C.F_USER_EMAIL),
        ("role", C.F_USER_ROLE),
        ("avatar", C.F_USER_AVATAR),
    ]:
        val = (body.get(key) or "").strip()
        if val:
            fields[const] = val
    if "allClients" in body:
        fields[C.F_USER_ALL_CLIENTS] = bool(body["allClients"])
    if "clientIds" in body:
        fields[C.F_USER_CLIENTS] = body["clientIds"] or []
    try:
        record = airtable.create_record(C.USERS_TABLE, fields, by_field_id=False, typecast=True)
    except requests.HTTPError as e:
        return airtable_err(e)
    user_id = record["id"]
    if body.get("pin"):
        airtable.update_record(
            C.USERS_TABLE, user_id,
            {C.F_USER_PIN_HASH: _hash_pin(user_id, str(body["pin"]))},
            by_field_id=False,
        )
    return jsonify({"user": _shape_user(record)}), 201


@api.put("/users/<user_id>")
def update_user(user_id):
    management_error = _require_user_management()
    if management_error:
        return management_error
    body = request.get_json(silent=True) or {}
    if not _is_admin():
        if _is_admin_role_value(body.get("role")):
            return err("Only administrators can assign administrator credentials.", 403)
        try:
            existing = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
        except requests.HTTPError as e:
            return airtable_err(e)
        if _record_is_admin_user(existing):
            return err("Only administrators can edit administrator users.", 403)
    fields = {}
    for key, const in [
        ("name", C.F_USER_NAME),
        ("firstName", C.F_USER_FIRST_NAME),
        ("lastName", C.F_USER_LAST_NAME),
        ("displayName", C.F_USER_DISPLAY_NAME),
        ("email", C.F_USER_EMAIL),
        ("role", C.F_USER_ROLE),
        ("avatar", C.F_USER_AVATAR),
    ]:
        if key in body:
            fields[const] = (body[key] or "").strip()
    if "active" in body:
        fields[C.F_USER_ACTIVE] = bool(body["active"])
    if "allClients" in body:
        fields[C.F_USER_ALL_CLIENTS] = bool(body["allClients"])
    if "clientIds" in body:
        fields[C.F_USER_CLIENTS] = body["clientIds"] or []
    if body.get("pin"):
        fields[C.F_USER_PIN_HASH] = _hash_pin(user_id, str(body["pin"]))
    try:
        if fields:
            record = airtable.update_record(C.USERS_TABLE, user_id, fields, by_field_id=False, typecast=True)
        else:
            record = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
    except requests.HTTPError as e:
        return airtable_err(e)
    return jsonify({"user": _shape_user(record)})


def _source_check_normalized_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _source_check_upc_key(value):
    text = str(value or "").strip()
    if not text or text.casefold() == "no upc":
        return ""
    return re.sub(r"\D+", "", text)


def _source_check_sheet_range(limit):
    header_row = max(1, C.TOPCO_SOURCE_HEADER_ROW)
    first_column, _, last_column = C.TOPCO_SOURCE_COLUMN_RANGE.partition(":")
    first_column = first_column or "A"
    last_column = last_column or "AE"
    end_row = header_row + max(1, limit)
    return f"{first_column}{header_row}:{last_column}{end_row}"


# Matching reads this sheet on every keystroke. Without a cache each one re-downloads
# the whole CSV, which is why the read window had to stay small to stay usable. Caching
# by range means the sheet is fetched at most once per refresh interval however much
# typing happens, so the window can cover the whole list.
_SOURCE_ROWS_CACHE = {}


def _fetch_topco_source_rows(limit, *, max_age_seconds=None):
    row_limit = max(1, min(int(limit or 20), 1000))
    source_range = _source_check_sheet_range(row_limit)
    ttl = C.TOPCO_SOURCE_REFRESH_INTERVAL_SECONDS if max_age_seconds is None else max_age_seconds
    cached = _SOURCE_ROWS_CACHE.get(source_range)
    if cached and ttl > 0 and (time.monotonic() - cached[0]) < ttl:
        return list(cached[1]), source_range
    url = (
        f"https://docs.google.com/spreadsheets/d/{quote(C.TOPCO_SOURCE_SHEET_ID)}/export"
        f"?format=csv&gid={quote(str(C.TOPCO_SOURCE_SHEET_GID))}&range={quote(source_range)}"
    )
    response = requests.get(url, timeout=15)
    response.raise_for_status()
    reader = csv.DictReader(io.StringIO(response.text))
    rows = []
    for index, row in enumerate(reader, start=C.TOPCO_SOURCE_HEADER_ROW + 1):
        if len(rows) >= row_limit:
            break
        source_data = {
            str(key or "").strip(): str(value or "").strip()
            for key, value in row.items()
            if str(key or "").strip() in C.TOPCO_SOURCE_CHECK_FIELDS
        }
        if not any(source_data.values()):
            continue
        rows.append({
            "sourceRowNumber": index,
            "sourceData": source_data,
        })
    _SOURCE_ROWS_CACHE[source_range] = (time.monotonic(), list(rows))
    return rows, source_range


def _topco_source_row_to_import_row(source_row):
    source_data = source_row.get("sourceData") if isinstance(source_row, dict) else {}
    source_data = source_data if isinstance(source_data, dict) else {}
    upc = str(source_data.get("UPC") or "").strip()
    product_name = str(source_data.get("Product Name") or "").strip()
    return {
        "rowNumber": source_row.get("sourceRowNumber") or source_row.get("rowNumber") or "",
        "noJob": True,
        "id": upc,
        "upc": upc,
        "cvid": source_data.get("CVID") or "",
        "brandPrefix": source_data.get("Brand Prefix") or "",
        "requestType": source_data.get("Request Type") or "",
        "wkftJobNumber": source_data.get("WKFT #") or source_data.get("WKFT Job Number") or "",
        "mboxNumber": source_data.get("Mbox #") or "",
        "productType": source_data.get("Product Type") or "",
        "fileNameDescription": source_data.get("Prod Descrip") or source_data.get("File Name Description") or source_data.get("Product Description") or "",
        "preproOverlays": source_data.get("Link to Prepro/Overlays") or "",
        "ecommPhotoNotes": source_data.get("Photo Notes") or source_data.get("Ecomm Photo Notes") or "",
        "pathToArt": source_data.get("Path to Art") or "",
        "itemName": product_name or upc or "Topco source row",
        "referenceData": source_data,
        "sourceSnapshotMatchMethod": "Source Lookup",
        "actionableReason": "activate_in_marks",
        "clearBlankSourceFields": True,
    }


def _topco_client_id():
    data = {"records": _client_records()}
    for record in _permitted_client_records(data.get("records", [])):
        if _client_name(record).strip().casefold() == "topco":
            return record.get("id", "")
    return ""


def _topco_activation_existing_product(client_id, source_row):
    source_data = source_row.get("sourceData") if isinstance(source_row, dict) else {}
    source_data = source_data if isinstance(source_data, dict) else {}
    source_row_number = source_row.get("sourceRowNumber") or source_row.get("rowNumber") or ""
    source_upc_key = _source_check_upc_key(source_data.get("UPC"))
    try:
        product_records = _list_all_records(C.PRODUCTS_TABLE)
    except requests.HTTPError:
        raise
    permitted_records = _filter_by_client_field(product_records, C.F_ITEM_CLIENT)
    same_client = [
        record for record in permitted_records
        if client_id in (record.get("fields", {}).get(C.F_ITEM_CLIENT, []) or [])
    ]
    for record in same_client:
        reference_data = _parse_reference_data(record.get("fields", {}).get(C.F_ITEM_REFERENCE_DATA, ""))
        snapshot = reference_data.get("_sourceSnapshot") if isinstance(reference_data, dict) else {}
        if not isinstance(snapshot, dict):
            continue
        if (
            snapshot.get("sourceRowNumber") == source_row_number
            and str(snapshot.get("source") or "").strip() == "TOPCO (MARKS) PROJECTS"
            and str(snapshot.get("sheetTab") or "").strip() == C.TOPCO_SOURCE_SHEET_TAB
        ):
            return record
    if source_upc_key:
        for record in same_client:
            fields = record.get("fields", {})
            product_upc_key = _source_check_upc_key(fields.get(C.F_ITEM_UPC) or fields.get(C.F_ITEM_IDENTIFIER))
            if product_upc_key and product_upc_key == source_upc_key:
                return record
    return None


def _source_check_product_indexes(products):
    by_upc = {}
    by_name = {}
    for product in products:
        upc_key = _source_check_upc_key(product.get("upc") or product.get("primaryMatchKey"))
        if upc_key:
            by_upc.setdefault(upc_key, []).append(product)
        name_key = _source_check_normalized_text(product.get("name"))
        if name_key:
            by_name.setdefault(name_key, []).append(product)
    return by_upc, by_name


def _source_check_match_product(source_data, by_upc, by_name):
    upc_key = _source_check_upc_key(source_data.get("UPC"))
    if upc_key:
        matches = by_upc.get(upc_key, [])
        if matches:
            return matches[0], "UPC", len(matches)
    name_key = _source_check_normalized_text(source_data.get("Product Name"))
    if name_key:
        matches = by_name.get(name_key, [])
        if matches:
            return matches[0], "Product Name", len(matches)
    return None, "", 0


def _topco_client_record(client_id):
    client_id = str(client_id or "").strip()
    if not client_id:
        return None
    try:
        record = _client_record(client_id)
    except requests.HTTPError:
        raise
    return record if _client_name(record).strip().casefold() == "topco" else None


# The first row in every merchandise history. Named for what happened to the physical
# item rather than to the database row, matching how the rest of Planning reads.
MERCHANDISE_CREATED_EVENT = "Merchandise received"


def _record_merchandise_history(merchandise_id, event, *, from_value=None, to_value=None):
    """Record one merchandise lifecycle event against the signed-in user.

    Deliberately simple: what happened, who did it, when. Failures are swallowed
    because losing an audit line must never block the action the user was taking.
    """
    if not merchandise_id:
        return
    user_id = _current_user_id()
    try:
        _create_history_event(
            event,
            merchandise_ids=[merchandise_id],
            user_ids=[user_id] if user_id else None,
            from_value=from_value,
            to_value=to_value,
        )
    except Exception:  # noqa: BLE001 - history is never worth failing a request over
        current_app.logger.exception("Could not record merchandise history event")


def _topco_source_suggestion_score(source_data, *, product_name="", upc=""):
    source_name = _source_check_normalized_text(source_data.get("Product Name"))
    source_upc = _source_check_upc_key(source_data.get("UPC"))
    name = _source_check_normalized_text(product_name)
    upc_key = _source_check_upc_key(upc)
    score = 0
    basis = []
    upc_score = 0
    name_score = 0
    if upc_key and source_upc:
        if upc_key == source_upc:
            upc_score = 100
        elif source_upc.startswith(upc_key) or upc_key.startswith(source_upc):
            upc_score = 78
        elif upc_key in source_upc or source_upc in upc_key:
            upc_score = 58
        if upc_score:
            score += upc_score
            basis.append("UPC")
    if name and source_name:
        if name == source_name:
            name_score = 70
            basis.append("Product Name")
        elif name in source_name or source_name in name:
            name_score = 42
            basis.append("Product Name")
        else:
            name_tokens = {token for token in name.split() if len(token) >= 3}
            source_tokens = {token for token in source_name.split() if len(token) >= 3}
            overlap = name_tokens & source_tokens
            if overlap:
                name_score = min(30, len(overlap) * 10)
                basis.append("Product Name")
        score += name_score
    if upc_score and name_score:
        score += 35
    # When the receiver supplied both clues, a row has to satisfy both. Scoring them
    # additively meant a name-only hit still surfaced alongside real candidates even
    # though its UPC shared nothing with what was typed. An exact UPC still stands on
    # its own, because a whole-value barcode hit is stronger than any name text.
    if upc_key and name and upc_score < 100 and not (upc_score and name_score):
        return 0, []
    return score, basis


def _topco_source_suggestions(*, client_id="", product_name="", upc="", limit=8):
    if client_id:
        client_record = _topco_client_record(client_id)
        if not client_record:
            return []
    product_name = str(product_name or "").strip()
    upc = str(upc or "").strip()
    if len(_match_compact(product_name)) < 3 and len(_source_check_upc_key(upc)) < 3:
        return []
    try:
        result_limit = int(limit)
    except (TypeError, ValueError):
        result_limit = 8
    result_limit = max(1, min(result_limit, 20))
    source_rows, _source_range = _fetch_topco_source_rows(C.TOPCO_SOURCE_MATCH_ROW_WINDOW)
    suggestions = []
    for row in source_rows:
        source_data = row.get("sourceData") or {}
        score, basis = _topco_source_suggestion_score(source_data, product_name=product_name, upc=upc)
        if not score:
            continue
        suggestions.append({
            **row,
            "score": score,
            "matchBasis": " + ".join(basis) or "Source row",
        })
    return sorted(
        suggestions,
        key=lambda row: (-row.get("score", 0), row.get("sourceRowNumber") or 0),
    )[:result_limit]


def _topco_source_row_by_number(source_row_number):
    source_rows, _source_range = _fetch_topco_source_rows(
        max(C.TOPCO_SOURCE_MATCH_ROW_WINDOW, source_row_number - C.TOPCO_SOURCE_HEADER_ROW)
    )
    source_row = next((row for row in source_rows if row.get("sourceRowNumber") == source_row_number), None)
    if not source_row:
        raise ValueError("That source row was not found in the current Topco source sheet read.")
    return source_row


def _source_snapshot_for_topco_product(record):
    reference_data = _parse_reference_data(record.get("fields", {}).get(C.F_ITEM_REFERENCE_DATA, ""))
    snapshot = reference_data.get("_sourceSnapshot") if isinstance(reference_data, dict) else {}
    if not isinstance(snapshot, dict):
        return None
    if str(snapshot.get("source") or "").strip() != "TOPCO (MARKS) PROJECTS":
        return None
    if str(snapshot.get("sheetTab") or "").strip() != C.TOPCO_SOURCE_SHEET_TAB:
        return None
    try:
        source_row_number = int(snapshot.get("sourceRowNumber") or 0)
    except (TypeError, ValueError):
        source_row_number = 0
    if source_row_number <= 0:
        return None
    return {**snapshot, "sourceRowNumber": source_row_number}


# The sheet is a public CSV export: free to fetch, and it offers no ETag or
# Last-Modified, so a conditional request is impossible. What costs is the
# Airtable work that follows. Fingerprinting the rows lets the sheet be polled
# as often as we like while Airtable is touched only when something changed -
# which, for a sheet the client rarely edits, is almost never.
_SOURCE_SHEET_FINGERPRINTS = {}


def _topco_source_sheet_fingerprint(limit):
    rows, _source_range = _fetch_topco_source_rows(limit, max_age_seconds=0)
    payload = json.dumps(rows, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _topco_source_sheet_unchanged(client_id, limit):
    """True when the sheet reads exactly as it did last time we looked."""
    try:
        digest = _topco_source_sheet_fingerprint(limit)
    except Exception:  # noqa: BLE001 - a sheet read failure must not skip the refresh
        return False
    previous = _SOURCE_SHEET_FINGERPRINTS.get(client_id)
    _SOURCE_SHEET_FINGERPRINTS[client_id] = digest
    return previous is not None and previous == digest


def _unchanged_refresh_result():
    return {
        "checked": 0,
        "updated": 0,
        "skipped": 0,
        "missingSourceRows": [],
        "sourceUnchanged": True,
        "summary": {"itemsCreated": 0, "itemsUpdated": 0, "rowsSkipped": 0, "errors": 0, "warnings": 0},
    }


def refresh_topco_source_linked_products(client_id, limit=100, enforce_permissions=True, force=False):
    try:
        refresh_limit = max(1, min(int(limit or 100), 500))
    except (TypeError, ValueError):
        refresh_limit = 100
    # Checked before the Products scan: an unchanged sheet costs no Airtable call.
    if not force and _topco_source_sheet_unchanged(client_id, refresh_limit):
        return _unchanged_refresh_result()
    product_records = _list_all_records(C.PRODUCTS_TABLE)
    permitted_records = _filter_by_client_field(product_records, C.F_ITEM_CLIENT) if enforce_permissions else product_records
    linked_products = []
    for record in permitted_records:
        fields = record.get("fields", {})
        if client_id not in (fields.get(C.F_ITEM_CLIENT, []) or []):
            continue
        snapshot = _source_snapshot_for_topco_product(record)
        if not snapshot:
            continue
        linked_products.append((record, snapshot))
        if len(linked_products) >= refresh_limit:
            break
    if not linked_products:
        return {
            "checked": 0,
            "updated": 0,
            "skipped": 0,
            "missingSourceRows": [],
            "summary": {"itemsCreated": 0, "itemsUpdated": 0, "rowsSkipped": 0, "errors": 0, "warnings": 0},
        }
    max_source_row = max(snapshot["sourceRowNumber"] for _record, snapshot in linked_products)
    source_limit = max(20, max_source_row - C.TOPCO_SOURCE_HEADER_ROW)
    source_rows, source_range = _fetch_topco_source_rows(source_limit)
    source_by_number = {row.get("sourceRowNumber"): row for row in source_rows}
    updated_count = 0
    missing_source_rows = []
    for record, snapshot in linked_products:
        source_row = source_by_number.get(snapshot["sourceRowNumber"])
        if not source_row:
            missing_source_rows.append(snapshot["sourceRowNumber"])
            continue
        refresh_row = _topco_source_row_to_import_row(source_row)
        refresh_row["sourceSnapshot"] = build_source_snapshot(
            source_row,
            "Source Refresh",
            "source_refresh",
        )
        refresh_row["actionableReason"] = "source_refresh"
        if not str((source_row.get("sourceData") or {}).get("Product Name") or "").strip():
            refresh_row["itemName"] = record.get("fields", {}).get(C.F_ITEM_NAME) or refresh_row["itemName"]
        fields = _item_fields_from_row(client_id, refresh_row)
        previous = airtable.get_record(C.PRODUCTS_TABLE, record["id"], by_field_id=False)
        updated = airtable.update_record(C.PRODUCTS_TABLE, record["id"], fields, by_field_id=False)
        _log_item_changes(record["id"], previous, updated, fields)
        updated_count += 1
    if not updated_count:
        return {
            "checked": len(linked_products),
            "updated": 0,
            "skipped": len(missing_source_rows),
            "missingSourceRows": sorted(set(missing_source_rows)),
            "sourceRange": source_range,
            "summary": {"itemsCreated": 0, "itemsUpdated": 0, "rowsSkipped": 0, "errors": 0, "warnings": 0},
        }
    summary = {
        "itemsCreated": 0,
        "itemsUpdated": updated_count,
        "rowsSkipped": len(missing_source_rows),
        "errors": 0,
        "warnings": 0,
    }
    return {
        "checked": len(linked_products),
        "updated": updated_count,
        "skipped": len(missing_source_rows),
        "missingSourceRows": sorted(set(missing_source_rows)),
        "sourceRange": source_range,
        "summary": summary,
    }


def _topco_source_match_metadata(source_row, match_method="Source Match"):
    source_data = source_row.get("sourceData") or {}
    return {
        "client": "Topco",
        "source": "TOPCO (MARKS) PROJECTS",
        "sheetTab": "Master Tracker 2026",
        "sourceRowNumber": source_row.get("sourceRowNumber") or source_row.get("rowNumber") or "",
        "sourceCheckedAt": datetime.now(timezone.utc).isoformat(),
        "matchMethod": match_method,
        "sourceIdentity": {
            "productName": source_data.get("Product Name") or "",
            "upc": source_data.get("UPC") or "",
        },
    }


def _merchandise_manual_info_with_source_match(entry_fields, source_row, match_method="Source Match"):
    manual_info = _manual_product_info_object(entry_fields.get(C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO, ""))
    manual_info["_sourceMatch"] = _topco_source_match_metadata(source_row, match_method=match_method)
    return _json_text(manual_info)


def _store_pending_topco_source_match(entry_id, entry, source_row, receipt=None):
    updated = _update_receipt_entry_record(entry_id, {
        C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO: _merchandise_manual_info_with_source_match(
            entry.get("fields", {}),
            source_row,
            match_method="Source Match",
        ),
        C.F_RECEIPT_ENTRY_MERCH_STATUS: "Received",
    })
    return _shape_verification_entry(updated, receipt)


def _activate_topco_source_product(client_id, source_row_number, source_row=None):
    source_row = source_row or _topco_source_row_by_number(source_row_number)

    existing_product = _topco_activation_existing_product(client_id, source_row)
    activation_row = _topco_source_row_to_import_row(source_row)
    if existing_product:
        activation_row["existingItemId"] = existing_product.get("id")
    plan = _build_intake_plan_from_mapped_rows(
        client_id,
        "Topco Source Lookup",
        [activation_row],
    )
    result = _execute_intake_plan(plan)
    row = (result.get("rows") or [{}])[0]
    product_id = row.get("existingItemId")
    if not product_id:
        raise RuntimeError("Source row activation did not return a Product.")
    product_record = airtable.get_record(C.PRODUCTS_TABLE, product_id, by_field_id=False)
    return product_record, result, row


def _link_merchandise_to_product(entry_id, item_id):
    entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    item = airtable.get_record(C.PRODUCTS_TABLE, item_id, by_field_id=False)
    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return None, None, None, _forbidden()
    item_client_ids = _as_list(item.get("fields", {}).get(C.F_ITEM_CLIENT, []))
    if not _client_ids_permitted(item_client_ids):
        return None, None, None, _forbidden()
    receipt_client_ids = _as_list(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])) if receipt else []
    if receipt_client_ids and item_client_ids and not (set(receipt_client_ids) & set(item_client_ids)):
        return None, None, None, err("Product does not belong to this Shipment client.", 403)
    entry_fields = entry.get("fields", {})
    updated = _update_receipt_entry_record(entry_id, {
        C.F_RECEIPT_ENTRY_ITEM: [item_id],
        C.F_RECEIPT_ENTRY_PLANNING_STATUS: entry_fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) or "New",
        **_merch_status_normalization_fields(entry_fields),
    })
    return updated, receipt, item, None


@api.get("/source-check/topco")
def source_check_topco():
    try:
        source_rows, source_range = _fetch_topco_source_rows(
            request.args.get("limit", "20"),
            max_age_seconds=0 if request.args.get("refresh") in ("1", "true", "yes") else None,
        )
    except requests.RequestException as exc:
        current_app.logger.exception("Source Check source sheet read failed")
        return err(f"Could not read source sheet: {exc}", 502)
    except ValueError:
        return err("Invalid Source Check row limit.", 400)
    try:
        product_records = _list_all_records(
            C.PRODUCTS_TABLE,
            params={"sort[0][field]": C.F_ITEM_NAME, "sort[0][direction]": "asc"},
        )
    except requests.HTTPError as exc:
        return airtable_err(exc)
    records = _filter_by_client_field(product_records, C.F_ITEM_CLIENT)
    clients_by_id = _clients_by_id()
    products = [_shape_item(record, clients_by_id=clients_by_id) for record in records]
    by_upc, by_name = _source_check_product_indexes(products)
    rows = []
    for source_row in source_rows:
        product, match_method, match_count = _source_check_match_product(source_row["sourceData"], by_upc, by_name)
        rows.append({
            **source_row,
            "product": product,
            "matchMethod": match_method,
            "matchCount": match_count,
        })
    return jsonify({
        "source": {
            "title": "TOPCO (MARKS) PROJECTS",
            "spreadsheetId": C.TOPCO_SOURCE_SHEET_ID,
            "sheetName": C.TOPCO_SOURCE_SHEET_TAB,
            "range": source_range,
            "headerRow": C.TOPCO_SOURCE_HEADER_ROW,
        },
        "checkedAt": _now_iso(),
        "rows": rows,
    })


@api.get("/source-check/topco/suggestions")
def source_check_topco_suggestions():
    client_id = (request.args.get("clientId") or "").strip()
    try:
        records = _topco_source_suggestions(
            client_id=client_id,
            product_name=request.args.get("productName") or request.args.get("name") or "",
            upc=request.args.get("upc") or request.args.get("identifier") or "",
            limit=request.args.get("limit", "8"),
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    except requests.RequestException as exc:
        current_app.logger.exception("Source Lookup suggestions source sheet read failed")
        return err(f"Could not read source sheet: {exc}", 502)
    return jsonify({"records": records})


@api.post("/source-check/topco/activate")
def activate_topco_source_row():
    body = request.get_json(silent=True) or {}
    try:
        source_row_number = int(body.get("sourceRowNumber") or 0)
    except (TypeError, ValueError):
        return err("Choose a source row to activate.", 400)
    if source_row_number <= 0:
        return err("Choose a source row to activate.", 400)

    client_id = str(body.get("clientId") or "").strip()
    if client_id:
        if not _client_permitted(client_id):
            return _forbidden()
        client_record = _client_record(client_id)
        if _client_name(client_record).strip().casefold() != "topco":
            return err("Source Lookup activation is currently available only for Topco.", 400)
    else:
        client_id = _topco_client_id()
        if not client_id:
            return err("Topco client is not available for this user.", 404)

    try:
        product_record, result, row = _activate_topco_source_product(client_id, source_row_number)
    except requests.RequestException as exc:
        current_app.logger.exception("Source Lookup activation source sheet read failed")
        return err(f"Could not read source sheet: {exc}", 502)
    except requests.HTTPError as error:
        return airtable_err(error)
    except ValueError as exc:
        return err(str(exc), 400)
    except RuntimeError as exc:
        return err(str(exc), 500)
    return jsonify({
        "activated": True,
        "action": row.get("action"),
        "sourceRowNumber": source_row_number,
        "summary": result.get("summary", {}),
        "record": _shape_item(product_record, clients_by_id=_clients_by_id(), issues_by_item_id=_issues_by_item_id()),
    })


@api.post("/source-check/topco/refresh-linked-products")
def refresh_topco_linked_source_products():
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    body = request.get_json(silent=True) or {}
    client_id = str(body.get("clientId") or "").strip()
    if client_id:
        if not _client_permitted(client_id):
            return _forbidden()
        client_record = _client_record(client_id)
        if _client_name(client_record).strip().casefold() != "topco":
            return err("Source refresh is currently available only for Topco.", 400)
    else:
        client_id = _topco_client_id()
        if not client_id:
            return err("Topco client is not available for this user.", 404)
    try:
        result = refresh_topco_source_linked_products(client_id, limit=body.get("limit") or 100, force=True)
    except requests.RequestException as exc:
        current_app.logger.exception("Topco linked Product source refresh failed")
        return err(f"Could not read source sheet: {exc}", 502)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({
        "refreshed": True,
        "clientId": client_id,
        **result,
    })


# ── Products ──────────────────────────────────────────────────────────────────

@api.get("/items")
@api.get("/products")
@api.get("/skus")
def list_items():
    params = {
        "sort[0][field]": C.F_ITEM_NAME,
        "sort[0][direction]": "asc",
    }
    data = airtable.list_records(C.PRODUCTS_TABLE, params=params, by_field_id=False)
    records = _filter_by_client_field(data.get("records", []), C.F_ITEM_CLIENT)
    clients_by_id = _clients_by_id()
    issues_by_item_id = _issues_by_item_id()
    try:
        production_summaries = _product_production_summaries(records)
    except (AssertionError, requests.RequestException):
        # Keep the Product data workspace usable when an optional relationship
        # table is unavailable or a narrow test/integration double omits it.
        production_summaries = {}
    records = [
        _shape_item(
            r,
            clients_by_id=clients_by_id,
            issues_by_item_id=issues_by_item_id,
            production_summary=production_summaries.get(r.get("id")),
        )
        for r in records
    ]
    return jsonify({"records": records})


@api.get("/items/<record_id>")
@api.get("/products/<record_id>")
@api.get("/skus/<record_id>")
def get_item(record_id):
    record = airtable.get_record(C.PRODUCTS_TABLE, record_id, by_field_id=False)
    if not _client_ids_permitted(record.get("fields", {}).get(C.F_ITEM_CLIENT, [])):
        return _forbidden()
    clients_by_id = _clients_by_id()
    issues_by_item_id = _issues_by_item_id()
    try:
        production_summary = _product_production_summaries([record]).get(record.get("id"))
    except (AssertionError, requests.RequestException):
        production_summary = None
    return jsonify({"record": _shape_item(record, clients_by_id=clients_by_id, issues_by_item_id=issues_by_item_id, required_to_shoot_full=True, production_summary=production_summary)})


@api.post("/items")
@api.post("/skus")
def create_item():
    body = request.get_json(silent=True) or {}
    client_id = body.get("clientId")
    if client_id and not _client_permitted(client_id):
        return _forbidden()
    identifier = (body.get("primaryMatchKey") or body.get("productId") or body.get("id") or body.get("gtinUpc") or "").strip()
    client_config = _client_config(client_id) if client_id else {}
    code_type = client_config.get("codeType") or body.get("codeType") or ""
    validation_error = _validate_item_identifier(identifier, code_type, _identifier_label(client_config))
    if validation_error:
        return err(validation_error)

    values = {**body, "name": body.get("name") or body.get("product") or identifier}
    try:
        # Through the merge like everything else: this endpoint created without
        # looking, which is how the same SKU acquired two records.
        data, outcome, _filled = merge_product(
            client_id, identifier, values, source=str(body.get("source") or "manual"),
        )
    except ValueError as error:
        return err(str(error))
    except requests.HTTPError as error:
        return airtable_err(error)

    if outcome == "created":
        _create_history_event("Item Created", item_ids=[data["id"]])
    shaped = _shape_item(data, clients_by_id=_clients_by_id(), issues_by_item_id=_issues_by_item_id())
    # 200 when it already existed, because nothing was created.
    return jsonify(shaped), (201 if outcome == "created" else 200)


@api.patch("/items/<record_id>")
@api.patch("/products/<record_id>")
@api.patch("/skus/<record_id>")
def update_item(record_id):
    body = request.get_json(silent=True) or {}
    previous = airtable.get_record(C.PRODUCTS_TABLE, record_id, by_field_id=False)
    if not _client_ids_permitted(previous.get("fields", {}).get(C.F_ITEM_CLIENT, [])):
        return _forbidden()
    fields = {}
    identifier = body.get("primaryMatchKey") or body.get("productId") or body.get("id") or body.get("gtinUpc")
    client_config = _client_config(body.get("clientId")) if body.get("clientId") else {}
    code_type = body.get("codeType") or client_config.get("codeType", "")
    if identifier is not None:
        validation_error = _validate_item_identifier(identifier.strip(), code_type, _identifier_label(client_config))
        if validation_error:
            return err(validation_error)
        fields[C.F_ITEM_IDENTIFIER] = identifier.strip()
    if body.get("clientId"):
        if not _client_permitted(body["clientId"]):
            return _forbidden()
        fields[C.F_ITEM_CLIENT] = [body["clientId"]]
    _apply_item_fields(fields, body)
    if not fields:
        return err("No updatable fields provided")

    try:
        data = airtable.update_record(
            C.PRODUCTS_TABLE,
            record_id,
            fields,
            by_field_id=False,
            typecast=any(field in PRODUCT_TYPECAST_FIELDS and fields[field] is not None for field in fields),
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    _log_item_changes(record_id, previous, data, fields)
    # Deliberately not touching the Creative Force feed. Editing a Product changes the
    # Product; releasing again is what tells Creative Force, and that is where the
    # producer gets warned.
    return jsonify(_shape_item(data, clients_by_id=_clients_by_id(), issues_by_item_id=_issues_by_item_id()))


@api.delete("/items/<record_id>")
@api.delete("/products/<record_id>")
@api.delete("/skus/<record_id>")
def delete_item(record_id):
    try:
        previous = airtable.get_record(C.PRODUCTS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _client_ids_permitted(previous.get("fields", {}).get(C.F_ITEM_CLIENT, [])):
        return _forbidden()
    try:
        airtable.delete_record(C.PRODUCTS_TABLE, record_id)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"deleted": True, "id": record_id})


def _clients_by_id():
    data = {"records": _client_records()}
    return {record["id"]: _shape_client(record) for record in _permitted_client_records(data.get("records", []))}


def _issues_by_item_id():
    data = airtable.list_records(C.ISSUES_TABLE, by_field_id=False)
    issues = {}
    for record in _filter_indirect_client_records(data.get("records", []), _client_ids_for_issue):
        shaped = _shape_issue(record)
        for item_id in shaped.get("itemIds", []):
            issues.setdefault(item_id, []).append(shaped)
    return issues


def _product_production_summaries(product_records):
    """Build Product progress as a read model from owned child records.

    This intentionally does not persist a Product status. Merchandise owns
    physical/intake state, Workstream Cards own photo work state, and THR3D
    Shipping Items own outbound movement.
    """
    product_ids = {record.get("id") for record in product_records if record.get("id")}
    if not product_ids:
        return {}
    merchandise_records = _list_all_records(C.MERCHANDISE_TABLE)
    workstream_records = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
    thr3d_records = _list_all_records(C.THR3D_SHIPPING_ITEMS_TABLE)
    grouped = {product_id: {"merchandise": [], "workstreams": [], "thr3d": []} for product_id in product_ids}
    for record in merchandise_records:
        linked = _as_list(record.get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM, []))
        for product_id in product_ids.intersection(linked):
            grouped[product_id]["merchandise"].append(record)
    for record in workstream_records:
        linked = _as_list(record.get("fields", {}).get(C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT, []))
        for product_id in product_ids.intersection(linked):
            grouped[product_id]["workstreams"].append(record)
    for record in thr3d_records:
        linked = _as_list(record.get("fields", {}).get(C.F_THR3D_SHIPPING_ITEM_EXPECTED_PRODUCT, []))
        for product_id in product_ids.intersection(linked):
            grouped[product_id]["thr3d"].append(record)
    return {product_id: _derive_product_production_summary(**children) for product_id, children in grouped.items()}


def _derive_product_production_summary(*, merchandise, workstreams, thr3d):
    merchandise_fields = [record.get("fields", {}) for record in merchandise]
    workstream_fields = [record.get("fields", {}) for record in workstreams]
    thr3d_fields = [record.get("fields", {}) for record in thr3d]
    work_units = len(workstream_fields) + len(thr3d_fields)
    physical_issue = any(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS) == "Issue" for fields in merchandise_fields)
    planning_statuses = {
        str(fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) or "").strip()
        for fields in merchandise_fields
    }
    # Preserve the existing summary vocabulary while sourcing the queue from
    # the single canonical Planning Status field.
    planning_labels = {
        {
            "New": "Needs Review",
            "Needs More Information": "Waiting on Information",
            "Awaiting Photo Release": "Awaiting Photo Release",
        }.get(value, value)
        for value in planning_statuses
    }
    card_statuses = {str(fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) or "").strip() for fields in workstream_fields}
    cf_statuses = {
        _parse_creative_force_sync(fields.get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, "")).get("status", "")
        for fields in workstream_fields
    }
    cf_statuses.discard("")
    shipping_statuses = {str(fields.get(C.F_THR3D_SHIPPING_ITEM_STATUS) or "").strip() for fields in thr3d_fields}

    if not merchandise:
        status = "No Merchandise"
    elif physical_issue:
        status = "Issue"
    elif "Needs Review" in planning_labels:
        status = "Needs Review"
    elif any(value in {"In Production", "Complete"} for value in cf_statuses):
        status = "Complete" if work_units and all(value == "Complete" for value in cf_statuses) and (not thr3d_fields or shipping_statuses == {"Shipped"}) else "In Production"
    elif work_units == 0:
        status = "Waiting on Information" if "Waiting on Information" in planning_labels else "Work Not Defined"
    elif {"Needs More Information", "Waiting on Information"} & card_statuses or "Waiting on Information" in planning_labels:
        status = "Waiting on Information"
    elif workstream_fields and all(value == "Awaiting Photo Release" for value in card_statuses):
        status = "Awaiting Photo Release"
    elif thr3d_fields and shipping_statuses == {"Shipped"} and not workstream_fields:
        status = "Complete"
    else:
        status = "Work Identified"

    return {
        "status": status,
        "merchandiseCount": len(merchandise_fields),
        "workstreamCount": len(workstream_fields),
        "thr3dShippingCount": len(thr3d_fields),
        "merchStatuses": sorted({str(fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS) or "").strip() for fields in merchandise_fields if fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS)}),
        "planningStatusLabels": sorted(planning_labels - {""}),
        "workstreamStatuses": sorted(card_statuses - {""}),
        "creativeForceStatuses": sorted(cf_statuses),
        "shippingStatuses": sorted(shipping_statuses - {""}),
    }


def _shape_item(r, *, clients_by_id=None, issues_by_item_id=None, required_to_shoot_full=False, production_summary=None):
    f = r.get("fields", {})
    code_type = f.get(C.F_ITEM_IDENTIFIER_TYPE, "")
    if isinstance(code_type, list):
        code_type = code_type[0] if code_type else ""
    client_ids = f.get(C.F_ITEM_CLIENT, [])
    client = (clients_by_id or {}).get(client_ids[0]) if client_ids else None
    if client and not code_type:
        code_type = client.get("codeType", "")
    item = {
        "id": r["id"],
        "name": f.get(C.F_ITEM_NAME, ""),
        "clientIds": client_ids,
        "clientName": client.get("name", "") if client else "",
        "productId": f.get(C.F_ITEM_IDENTIFIER, ""),
        "identifier": f.get(C.F_ITEM_IDENTIFIER, ""),
        "gtinUpc": f.get(C.F_ITEM_IDENTIFIER, ""),
        "primaryMatchKey": f.get(C.F_ITEM_IDENTIFIER, ""),
        "codeType": code_type,
        "identifierLabel": _identifier_label(client),
        "primaryMatchKeyLabel": _identifier_label(client),
        "upc": f.get(C.F_ITEM_UPC, "") or f.get(C.F_ITEM_IDENTIFIER, ""),
        "cvid": f.get(C.F_ITEM_CVID, ""),
        "brandPrefix": f.get(C.F_ITEM_BRAND_PREFIX, ""),
        "product": f.get(C.F_ITEM_PRODUCT, ""),
        "requestType": f.get(C.F_ITEM_REQUEST_TYPE, ""),
        "projectStatus": f.get(C.F_ITEM_PROJECT_STATUS, ""),
        "wkftJobNumber": f.get(C.F_ITEM_WKFT_JOB_NUMBER, ""),
        "mboxNumber": f.get(C.F_ITEM_MBOX_NUMBER, ""),
        "projectName": f.get(C.F_ITEM_PROJECT_NAME, ""),
        "productType": f.get(C.F_ITEM_PRODUCT_TYPE, ""),
        "fileNameDescription": f.get(C.F_ITEM_FILE_NAME_DESCRIPTION, ""),
        "preproOverlays": f.get(C.F_ITEM_PREPRO_OVERLAYS, ""),
        "ecommPhotoNotes": f.get(C.F_ITEM_ECOMM_PHOTO_NOTES, ""),
        "pathToArt": f.get(C.F_ITEM_PATH_TO_ART, ""),
        "itemJobNumber": f.get(C.F_ITEM_JOB_NUMBER, ""),
        "masterOrVariant": f.get(C.F_ITEM_MASTER_VARIANT, ""),
        "pickupJobNumber": f.get(C.F_ITEM_PICKUP_JOB_NUMBER, ""),
        "brand": f.get(C.F_ITEM_BRAND, ""),
        "category": f.get(C.F_ITEM_CATEGORY, ""),
        "artworkReceived": f.get(C.F_ITEM_ARTWORK_RECEIVED, False),
        "notes": f.get(C.F_ITEM_NOTES, ""),
        "referenceDataRaw": f.get(C.F_ITEM_REFERENCE_DATA, ""),
        "referenceData": _parse_reference_data(f.get(C.F_ITEM_REFERENCE_DATA, "")),
        "productionSummary": production_summary or {
            "status": "Not Calculated",
            "merchandiseCount": 0,
            "workstreamCount": 0,
            "thr3dShippingCount": 0,
            "merchStatuses": [],
            "planningStatusLabels": [],
            "workstreamStatuses": [],
            "creativeForceStatuses": [],
            "shippingStatuses": [],
        },
    }
    return item


# Every path that brings product data into the system - a Structure Form, a source
# sheet row, receiving, someone typing what a chat message told them - goes through
# one merge. Four paths that each created Products their own way is how a SKU ends
# up with two records that no later rule can reconcile.


def normalized_identifier(value):
    """The comparable form of an identifier.

    Digits only, with leading zeros dropped: the spreadsheet stores UPCs as numbers
    and strips them, so `036800120457` and `36800120457` are one product. Nothing
    else is relaxed. Dropping the trailing digit would merge five distinct CT
    cheeses, which differ in that digit alone.
    """
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits.lstrip("0")


def _product_identifier_index(client_id, records=None):
    """Every Product for a client, keyed by comparable identifier."""
    index = {}
    for record in (records if records is not None else _list_all_records(C.PRODUCTS_TABLE)):
        fields = record.get("fields", {})
        if client_id and client_id not in (fields.get(C.F_ITEM_CLIENT, []) or []):
            continue
        for candidate in (fields.get(C.F_ITEM_UPC), fields.get(C.F_ITEM_IDENTIFIER)):
            key = normalized_identifier(candidate)
            if key:
                index.setdefault(key, record)
    return index


def merge_product(client_id, identifier, values, *, source="", reference=None, records=None):
    """Create a Product or fill its gaps. The only way a Product comes to exist.

    Returns (record, outcome, filled) where outcome is "created", "filled" or
    "unchanged", and filled names the fields this call supplied.

    Existing values are never overwritten. A contribution can only answer a
    question the record has not answered yet, so a late Structure Form cannot
    quietly rewrite what receiving observed, and vice versa.
    """
    key = normalized_identifier(identifier)
    if not key:
        raise ValueError("A Product needs an identifier.")

    patch = {k: v for k, v in (values or {}).items() if str(v or "").strip()}
    existing = _product_identifier_index(client_id, records).get(key)

    if existing is None:
        fields = {C.F_ITEM_CLIENT: [client_id]} if client_id else {}
        _apply_item_fields(fields, patch)
        # Written to both identifier fields. Products made from a form carried a UPC
        # and no Identifier, and the sheet import indexes Identifier - so the import
        # could not see them and made a second record for the same SKU. Filling both
        # removes the disagreement at its source rather than at every lookup.
        raw = str(identifier).strip()
        for field in (C.F_ITEM_UPC, C.F_ITEM_IDENTIFIER):
            if not str(fields.get(field, "") or "").strip():
                fields[field] = raw
        provenance = _product_reference_data({}, source, sorted(patch), created=True,
                                             reference=reference)
        if provenance:
            fields[C.F_ITEM_REFERENCE_DATA] = provenance
        record = airtable.create_record(C.PRODUCTS_TABLE, fields, by_field_id=False, typecast=True)
        return record, "created", sorted(patch)

    current = existing.get("fields", {})
    gaps = {}
    for name, value in patch.items():
        target = {}
        _apply_item_fields(target, {name: value})
        for field in target:
            if not str(current.get(field, "") or "").strip():
                gaps[name] = value
    if not gaps:
        return existing, "unchanged", []

    fields = {}
    _apply_item_fields(fields, gaps)
    provenance = _product_reference_data(
        _parse_reference_data(current.get(C.F_ITEM_REFERENCE_DATA, "")),
        source, sorted(gaps), created=False, reference=reference,
    )
    if provenance:
        fields[C.F_ITEM_REFERENCE_DATA] = provenance
    record = airtable.update_record(C.PRODUCTS_TABLE, existing["id"], fields,
                                    by_field_id=False, typecast=True)
    return record, "filled", sorted(gaps)


def _product_reference_data(stored, source, filled, *, created, reference=None):
    """Reference Data noting which source answered which fields.

    Per field rather than per record: when a CVID turns out wrong, the question is
    where that one came from, not whether the Product had a form behind it.

    Returned for the caller to include in the same write. Recording provenance is
    not worth a second API call, and it is not worth failing a merge over either.
    """
    if not (source or reference):
        return ""
    try:
        merged = dict(stored or {})
        for key, value in (reference or {}).items():
            merged.setdefault(key, value)
        if source and filled:
            contributions = dict(merged.get("_contributions") or {})
            stamp = {"source": source, "at": _now_iso()}
            for name in filled:
                contributions.setdefault(name, stamp)
            merged["_contributions"] = contributions
            if created:
                merged.setdefault("_origin", stamp)
        return _reference_data_json(merged)
    except Exception:  # noqa: BLE001 - provenance never fails a merge
        current_app.logger.exception("Could not build Product provenance")
        return ""


def _apply_item_fields(fields, body):
    mapping = {
        "name": C.F_ITEM_NAME,
        "product": C.F_ITEM_PRODUCT,
        "brand": C.F_ITEM_BRAND,
        "category": C.F_ITEM_CATEGORY,
        "notes": C.F_ITEM_NOTES,
        "artworkReceived": C.F_ITEM_ARTWORK_RECEIVED,
    }
    for key, field in mapping.items():
        if key in body and body[key] is not None:
            fields[field] = body[key]
    if "itemJobNumber" in body and body["itemJobNumber"] is not None:
        fields[C.F_ITEM_JOB_NUMBER] = _normalize_item_job_number(body.get("itemJobNumber"))
    if "upc" in body and body["upc"] is not None:
        fields[C.F_ITEM_UPC] = str(body.get("upc") or "").strip()
    if "cvid" in body and body["cvid"] is not None:
        fields[C.F_ITEM_CVID] = str(body.get("cvid") or "").strip()
    if "brandPrefix" in body and body["brandPrefix"] is not None:
        fields[C.F_ITEM_BRAND_PREFIX] = str(body.get("brandPrefix") or "").strip()
    for key, field in {
        "requestType": C.F_ITEM_REQUEST_TYPE,
        "wkftJobNumber": C.F_ITEM_WKFT_JOB_NUMBER,
        "mboxNumber": C.F_ITEM_MBOX_NUMBER,
        "projectName": C.F_ITEM_PROJECT_NAME,
        "productType": C.F_ITEM_PRODUCT_TYPE,
        "preproOverlays": C.F_ITEM_PREPRO_OVERLAYS,
        "ecommPhotoNotes": C.F_ITEM_ECOMM_PHOTO_NOTES,
        "pathToArt": C.F_ITEM_PATH_TO_ART,
        "studioDestination": C.F_ITEM_STUDIO_DESTINATION,
        "vendor": C.F_ITEM_VENDOR,
    }.items():
        if key in body and body[key] is not None:
            value = body.get(key)
            if key == "productType":
                normalized = _normalize_product_type(value)
                if normalized:
                    fields[field] = normalized
            elif key == "requestType":
                normalized = _normalize_product_request_type(value)
                if normalized:
                    fields[field] = normalized
            else:
                fields[field] = str(value or "").strip()
    if "fileNameDescription" in body and body["fileNameDescription"] is not None:
        fields[C.F_ITEM_FILE_NAME_DESCRIPTION] = str(body.get("fileNameDescription") or "").strip()
    if "masterOrVariant" in body and body["masterOrVariant"] is not None:
        normalized = _normalize_master_or_variant(body.get("masterOrVariant"))
        if normalized:
            fields[C.F_ITEM_MASTER_VARIANT] = normalized
    if "pickupJobNumber" in body and body["pickupJobNumber"] is not None:
        fields[C.F_ITEM_PICKUP_JOB_NUMBER] = _normalize_item_job_number(body.get("pickupJobNumber"))
    if "referenceData" in body:
        fields[C.F_ITEM_REFERENCE_DATA] = _reference_data_json(body.get("referenceData"))
    elif "referenceDataRaw" in body:
            fields[C.F_ITEM_REFERENCE_DATA] = str(body.get("referenceDataRaw") or "")


def _log_item_changes(record_id, previous, current, changed_fields):
    pass


def _client_code_type(client_id):
    if not client_id:
        return ""
    data = {"records": _client_records()}
    for record in data.get("records", []):
        if record.get("id") == client_id:
            return record.get("fields", {}).get(C.F_CLIENT_IDENTIFIER_TYPE, "")
    return ""


def _client_config(client_id):
    if not client_id:
        return {}
    try:
        return _shape_client(airtable.get_record(C.CLIENTS_TABLE, client_id, by_field_id=False))
    except Exception:  # noqa: BLE001
        # Readiness and naming fall back to their defaults rather than failing the
        # request outright when the client record cannot be read.
        return {}


def _validate_item_identifier(identifier, code_type, label="Primary Match Key"):
    return _validate_identifier_value(identifier, code_type, label)


def _match_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _match_compact(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _item_match_score(item, query):
    query_text = _match_text(query)
    query_compact = _match_compact(query)
    if len(query_compact) < 3:
        return 0
    reference_data = item.get("referenceData") or {}
    reference_values = reference_data.values() if isinstance(reference_data, dict) else []
    fields = [
        item.get("identifier", ""),
        item.get("name", ""),
        item.get("product", ""),
        item.get("itemJobNumber", ""),
        item.get("masterOrVariant", ""),
        item.get("pickupJobNumber", ""),
        item.get("description", ""),
        item.get("brand", ""),
        item.get("category", ""),
        item.get("referenceDataRaw", ""),
        *[str(value) for value in reference_values],
    ]
    identifier = _match_compact(item.get("identifier", ""))
    item_job_number = _match_compact(item.get("itemJobNumber", ""))
    name_text = _match_text(item.get("name", ""))
    product_text = _match_text(item.get("product", ""))
    detail_text = _match_text(" ".join(fields))
    if query_compact and query_compact in {identifier, item_job_number}:
        return 120
    if query_compact and (identifier.startswith(query_compact) or item_job_number.startswith(query_compact)):
        return 104
    if query_compact and (query_compact in identifier or query_compact in item_job_number):
        return 92
    if query_text and query_text in {name_text, product_text}:
        return 84
    if query_text and (query_text in name_text or query_text in product_text):
        return 70
    query_tokens = set(query_text.split())
    detail_tokens = set(detail_text.split())
    overlap = query_tokens & detail_tokens
    if overlap:
        return 48 + min(18, len(overlap) * 6)
    # Prefix / partial-word matching: "smit" matches "smithfield", "000368" matches "000368abc"
    prefix_matches = sum(
        1 for qt in query_tokens
        if len(qt) >= 3 and any(dt.startswith(qt) for dt in detail_tokens)
    )
    if prefix_matches:
        return 36 + min(16, prefix_matches * 8)
    return 0


def _find_matching_skus(query, *, client_id="", include_item_id="", limit=8):
    cleaned = (query or "").strip()
    if len(_match_compact(cleaned)) < 3:
        return []

    data = airtable.list_records(C.PRODUCTS_TABLE, params={"sort[0][field]": C.F_ITEM_NAME, "sort[0][direction]": "asc"}, by_field_id=False)
    matches = []
    for record in _filter_by_client_field(data.get("records", []), C.F_ITEM_CLIENT):
        if client_id and client_id not in _as_list(record.get("fields", {}).get(C.F_ITEM_CLIENT, [])):
            continue
        sku = _shape_item(record)
        score = _item_match_score(sku, cleaned)
        if score:
            matches.append({
                **sku,
                "score": score,
                "jobNumber": sku.get("itemJobNumber") or "",
                # A Product with no merchandise linked is still expecting some. One
                # that already has merchandise is more likely a coincidental name
                # match than the thing in the receiver's hands.
                "awaitingMerchandise": not _as_list(record.get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM_MERCHANDISE, [])),
            })
    try:
        result_limit = int(limit)
    except (TypeError, ValueError):
        result_limit = 8
    result_limit = max(1, min(result_limit, 100))
    return sorted(
        matches,
        key=lambda sku: (
            not sku.get("awaitingMerchandise"),
            -sku["score"],
            sku.get("gtinUpc") or sku.get("name") or "",
        ),
    )[:result_limit]


# ── Issues ────────────────────────────────────────────────────────────────────

@api.get("/issues")
def list_issues():
    data = airtable.list_records(
        C.ISSUES_TABLE,
        params={"sort[0][field]": C.F_ISSUE_OPENED, "sort[0][direction]": "desc"},
        by_field_id=False,
    )
    records = _filter_indirect_client_records(data.get("records", []), _client_ids_for_issue)
    records = [_shape_issue(record) for record in records]
    return jsonify({"records": records})


@api.post("/issues")
def create_issue():
    body = request.get_json(silent=True) or {}
    issue = (body.get("issue") or body.get("name") or "").strip()
    if not issue:
        return err("issue is required")

    fields = {C.F_ISSUE_NAME: issue}
    item_ids = _as_list(body.get("itemId") or body.get("itemIds"))
    if item_ids and not _all_linked_records_permitted(C.PRODUCTS_TABLE, item_ids):
        return _forbidden()
    _set_link_field(fields, C.F_ISSUE_ITEM, body.get("itemId") or body.get("itemIds"))
    _set_link_field(fields, C.F_ISSUE_ASSIGNED, body.get("assignedId") or body.get("assignedIds"))
    for key, field in {
        "type": C.F_ISSUE_TYPE,
        "status": C.F_ISSUE_STATUS,
        "priority": C.F_ISSUE_PRIORITY,
        "opened": C.F_ISSUE_OPENED,
        "closed": C.F_ISSUE_CLOSED,
        "notes": C.F_ISSUE_NOTES,
    }.items():
        if key in body and body[key] not in (None, ""):
            fields[field] = body[key]

    data = airtable.create_record(C.ISSUES_TABLE, fields, by_field_id=False)
    issue_id = data["id"]
    issue_fields = data.get("fields", {})
    _create_history_event(
        "Issue Created",
        item_ids=issue_fields.get(C.F_ISSUE_ITEM),
        user_ids=issue_fields.get(C.F_ISSUE_ASSIGNED),
    )
    return jsonify(_shape_issue(data)), 201


@api.patch("/issues/<record_id>")
def update_issue(record_id):
    body = request.get_json(silent=True) or {}
    previous = airtable.get_record(C.ISSUES_TABLE, record_id, by_field_id=False)
    if not _client_ids_permitted(_client_ids_for_issue(previous)):
        return _forbidden()
    fields = {}
    item_ids = _as_list(body.get("itemId") or body.get("itemIds"))
    if item_ids and not _all_linked_records_permitted(C.PRODUCTS_TABLE, item_ids):
        return _forbidden()
    _set_link_field(fields, C.F_ISSUE_ITEM, body.get("itemId") or body.get("itemIds"))
    _set_link_field(fields, C.F_ISSUE_ASSIGNED, body.get("assignedId") or body.get("assignedIds"))
    for key, field in {
        "issue": C.F_ISSUE_NAME,
        "name": C.F_ISSUE_NAME,
        "type": C.F_ISSUE_TYPE,
        "status": C.F_ISSUE_STATUS,
        "priority": C.F_ISSUE_PRIORITY,
        "opened": C.F_ISSUE_OPENED,
        "closed": C.F_ISSUE_CLOSED,
        "notes": C.F_ISSUE_NOTES,
    }.items():
        if key in body and body[key] not in (None, ""):
            fields[field] = body[key]
    if not fields:
        return err("No updatable fields provided")

    data = airtable.update_record(C.ISSUES_TABLE, record_id, fields, by_field_id=False)
    old_status = _field_value(previous, C.F_ISSUE_STATUS)
    new_status = _field_value(data, C.F_ISSUE_STATUS)
    if C.F_ISSUE_STATUS in fields and old_status != "Resolved" and new_status == "Resolved":
        issue_fields = data.get("fields", {})
        _create_history_event(
            "Issue Resolved",
            item_ids=issue_fields.get(C.F_ISSUE_ITEM),
                user_ids=issue_fields.get(C.F_ISSUE_ASSIGNED),
            from_value=old_status,
            to_value="Resolved",
        )
    return jsonify(_shape_issue(data))


def _shape_issue(r):
    f = r.get("fields", {})
    return {
        "id": r["id"],
        "name": f.get(C.F_ISSUE_NAME, ""),
        "issue": f.get(C.F_ISSUE_NAME, ""),
        "itemIds": f.get(C.F_ISSUE_ITEM, []),
        "type": f.get(C.F_ISSUE_TYPE, ""),
        "status": f.get(C.F_ISSUE_STATUS, ""),
        "priority": f.get(C.F_ISSUE_PRIORITY, ""),
        "assignedIds": f.get(C.F_ISSUE_ASSIGNED, []),
        "opened": f.get(C.F_ISSUE_OPENED, ""),
        "closed": f.get(C.F_ISSUE_CLOSED, ""),
        "photos": [],
        "notes": f.get(C.F_ISSUE_NOTES, ""),
    }


# ── History ───────────────────────────────────────────────────────────────────

@api.get("/history")
def list_history():
    item_id = request.args.get("itemId")
    user_id = request.args.get("userId")
    limit = request.args.get("limit", "100")
    requested_limit = int(limit) if limit.isdigit() else 100
    params = {
        "sort[0][field]": C.F_HISTORY_DATE,
        "sort[0][direction]": "desc",
        "maxRecords": max(requested_limit * 10, requested_limit, 100),
    }
    data = airtable.list_records(
        C.HISTORY_TABLE,
        params=params,
        by_field_id=False,
    )
    records = _filter_indirect_client_records(data.get("records", []), _client_ids_for_history)
    records = [_shape_history(record) for record in records]
    if item_id or user_id:
        records = [
            record for record in records
            if (item_id and item_id in record.get("itemIds", []))
            or (user_id and user_id in record.get("userIds", []))
        ]
    records = records[:requested_limit]
    return jsonify({"records": records})


@api.post("/history")
def create_history():
    body = request.get_json(silent=True) or {}
    event = (body.get("event") or "").strip()
    if not event:
        return err("event is required")

    fields = {C.F_HISTORY_EVENT: event}
    item_ids = _as_list(body.get("itemId") or body.get("itemIds"))
    if item_ids and not _all_linked_records_permitted(C.PRODUCTS_TABLE, item_ids):
        return _forbidden()
    _set_link_field(fields, C.F_HISTORY_ITEM, body.get("itemId") or body.get("itemIds"))
    _set_link_field(fields, C.F_HISTORY_USER, body.get("userId") or body.get("userIds"))
    for key, field in {
        "date": C.F_HISTORY_DATE,
        "from": C.F_HISTORY_FROM,
        "to": C.F_HISTORY_TO,
    }.items():
        if key in body and body[key] not in (None, ""):
            fields[field] = body[key]

    data = airtable.create_record(C.HISTORY_TABLE, fields, by_field_id=False)
    return jsonify(_shape_history(data)), 201


def _shape_history(r):
    f = r.get("fields", {})
    return {
        "id": r["id"],
        "event": f.get(C.F_HISTORY_EVENT, ""),
        "itemIds": f.get(C.F_HISTORY_ITEM, []),
        "userIds": f.get(C.F_HISTORY_USER, []),
        "date": f.get(C.F_HISTORY_DATE, ""),
        "from": f.get(C.F_HISTORY_FROM, ""),
        "to": f.get(C.F_HISTORY_TO, ""),
    }


def _set_link_field(fields, field, value):
    if not value:
        return
    fields[field] = value if isinstance(value, list) else [value]


# ── Shipments ────────────────────────────────────────────────────────────────

@api.get("/receipts")
@api.get("/shipments")
@api.get("/receiving")
def list_receipts():
    try:
        data = airtable.list_records(
            C.SHIPMENTS_TABLE,
            params={"sort[0][field]": C.F_RECEIPT_RECEIVED, "sort[0][direction]": "desc"},
            by_field_id=False,
        )
    except requests.HTTPError as error:
        return airtable_err(error)

    records = _filter_receipts_by_access(data.get("records", []))
    client_id = (request.args.get("clientId") or "").strip()
    unassigned_client = (request.args.get("unassignedClient") or "").strip().lower() in {"1", "true", "yes"}
    if client_id:
        if not _client_permitted(client_id):
            return _forbidden()
        records = [record for record in records if client_id in _as_list(record.get("fields", {}).get(C.F_RECEIPT_CLIENT, []))]
    if unassigned_client:
        records = [record for record in records if not _as_list(record.get("fields", {}).get(C.F_RECEIPT_CLIENT, []))]
    entries_by_receipt = _receipt_entries_by_receipt_id([record["id"] for record in records])
    records = [_shape_receipt(record, entries_by_receipt=entries_by_receipt) for record in records]
    return jsonify({"records": records})


@api.delete("/shipments/<shipment_id>")
@api.delete("/receiving/<shipment_id>")
def delete_shipment(shipment_id):
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, shipment_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    entries_by_receipt = _receipt_entries_by_receipt_id([shipment_id])
    entries = entries_by_receipt.get(shipment_id, [])
    if entries:
        return jsonify({
            "error": "Remove merchandise from this Shipment before deleting it.",
            "entryCount": len(entries),
        }), 400
    photos = _shipment_photo_metadata_from_fields(receipt.get("fields", {}), include_urls=False)
    deleted_photos = []
    storage = None
    if photos:
        try:
            storage = _photo_storage()
        except (ReceivingPhotoConfigError, ReceivingPhotoValidationError):
            storage = None
    try:
        airtable.delete_record(C.SHIPMENTS_TABLE, shipment_id)
    except requests.HTTPError as error:
        return airtable_err(error)
    if storage:
        for photo in photos:
            object_key = photo.get("object_key") or photo.get("objectKey")
            if not object_key:
                continue
            try:
                storage.delete_photo(object_key)
                deleted_photos.append(object_key)
            except (ReceivingPhotoValidationError, ReceivingPhotoConfigError, ReceivingPhotoStorageError):
                pass
    return jsonify({"deleted": True, "id": shipment_id, "deletedPhotoKeys": deleted_photos})


# ── Verification ──────────────────────────────────────────────────────────────

def _list_merchandise_review_records():
    try:
        entries = _list_all_records(C.MERCHANDISE_TABLE)
        receipts = _list_all_records(C.SHIPMENTS_TABLE)
        issues = _list_all_records(C.ISSUES_TABLE)
    except requests.HTTPError as error:
        raise error

    receipts_by_id = {record["id"]: record for record in _filter_receipts_by_access(receipts)}
    issues_by_item = {}
    for issue in _filter_indirect_client_records(issues, _client_ids_for_issue):
        shaped_issue = _shape_issue(issue)
        for item_id in shaped_issue.get("itemIds", []):
            issues_by_item.setdefault(item_id, []).append(shaped_issue)
    records = []
    for entry in entries:
        if entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RELEASED):
            continue
        linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
        receipt = next((receipts_by_id.get(receipt_id) for receipt_id in linked_receipts if receipt_id in receipts_by_id), None)
        if linked_receipts and receipt is None:
            continue
        records.append(_shape_verification_entry(entry, receipt, issues_by_item_id=issues_by_item))
    records.sort(key=lambda record: (record.get("received") or "", record.get("name") or ""), reverse=True)
    return records


NON_INVENTORY_MERCH_STATUSES = {
    "disposed",
    "destroyed",
    "removed",
    "returned",
    "shipped",
    "shipped to thr3d",
    "sent to thr3d",
    "sent to thread",
}
NON_INVENTORY_PRODUCT_STATUSES = {"cancelled"}
MERCH_STATUS_VALUES = {"Received", "Issue", "Ready to Ship", "Shipped", "Disposed"}


def _normalized_merch_status(value, default="Received"):
    status = str(value or "").strip()
    if status in MERCH_STATUS_VALUES:
        return status
    legacy = status.lower()
    if legacy in {"matched", "validated", "needs match", "no clear match", ""}:
        return default
    if legacy in {"shipped to thr3d", "sent to thr3d", "sent to thread", "returned"}:
        return "Shipped"
    if legacy in {"destroyed", "removed"}:
        return "Disposed"
    return status or default


def _merch_status_normalization_fields(fields):
    current = fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS)
    normalized = _normalized_merch_status(current)
    return {C.F_RECEIPT_ENTRY_MERCH_STATUS: normalized} if normalized != current else {}


def _intake_decision_value(body, *keys):
    for key in keys:
        if key in body:
            return (body.get(key) or "").strip()
    return None


def _validate_intake_choice(value, allowed, label):
    if value is None:
        return None
    if value == "":
        return ""
    if value not in allowed:
        return err(f"{label} must be one of: {', '.join(allowed)}.")
    return value


DELIVERABLE_ALIASES = {
    "packaging": "Packaging",
    "packaging photo": "Packaging",
    "packaging photography": "Packaging",
    "ecomm": "Ecomm",
    "ecomm photo": "Ecomm",
    "ecommerce": "Ecomm",
    "ecommerce photo": "Ecomm",
    "ecommerce photography": "Ecomm",
    "gs1 ecomm": "Ecomm",
    "thr3d": "Thr3d",
    "3d": "Thr3d",
    "thread": "Thr3d",
}


def _strip_surrounding_quotes(value):
    text = str(value or "").strip()
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
        text = text[1:-1].strip()
    if text and all(character in {'"', "'"} for character in text):
        return ""
    return text


def _flatten_deliverable_values(value):
    if value in (None, ""):
        return []
    if isinstance(value, list):
        flattened = []
        for item in value:
            flattened.extend(_flatten_deliverable_values(item))
        return flattened
    if isinstance(value, dict):
        return _flatten_deliverable_values(value.get("name") or value.get("label") or value.get("value") or "")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if (text.startswith("[") and text.endswith("]")) or (text.startswith('"') and text.endswith('"')):
            try:
                parsed = json.loads(text)
            except (TypeError, ValueError):
                parsed = None
            if parsed is not None and parsed != text:
                return _flatten_deliverable_values(parsed)
        if "," in text:
            flattened = []
            for item in text.split(","):
                flattened.extend(_flatten_deliverable_values(item))
            return flattened
        return [_strip_surrounding_quotes(text)]
    return [_strip_surrounding_quotes(value)]


def _as_clean_string_list(value):
    cleaned = []
    for item in _flatten_deliverable_values(value):
        text = str(item or "").strip()
        if text and text not in cleaned:
            cleaned.append(text)
    return cleaned


def _normalize_deliverable(value):
    text = _strip_surrounding_quotes(value)
    if not text:
        return ""
    return DELIVERABLE_ALIASES.get(text.lower(), text)


def _deliverable_values(value):
    normalized = [_normalize_deliverable(item) for item in _as_clean_string_list(value)]
    return [item for index, item in enumerate(normalized) if item and item not in normalized[:index]]


def _validate_deliverables(value):
    selected = _deliverable_values(value)
    invalid = [item for item in selected if item not in C.DELIVERABLE_OPTIONS]
    if invalid:
        current_app.logger.warning("Rejected malformed Deliverables payload: raw=%r normalized=%r invalid=%r", value, selected, invalid)
        return err(f"Deliverables must be one or more of: {', '.join(C.DELIVERABLE_OPTIONS)}.")
    return selected


def _validate_planning_status_label(value):
    if value is None:
        return None
    if value not in C.PLANNING_STATUS_OPTIONS:
        return err(f"Planning status must be one of: {', '.join(C.PLANNING_STATUS_OPTIONS)}.")
    return value


def _positive_int(value, label):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return err(f"{label} must be a whole number greater than 0.")
    if number <= 0:
        return err(f"{label} must be a whole number greater than 0.")
    return number


def _json_text(value):
    if value in (None, "", {}, []):
        return ""
    if isinstance(value, str):
        return value.strip()
    return json.dumps(value, sort_keys=True)


def _manual_product_info_object(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return {str(key): val for key, val in value.items() if val not in (None, "")}
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except (TypeError, ValueError, json.JSONDecodeError):
            return {}
        if isinstance(parsed, dict):
            return {str(key): val for key, val in parsed.items() if val not in (None, "")}
    return {}


def _manual_product_from_fields(fields, fallback_entry_fields=None):
    fallback_entry_fields = fallback_entry_fields or {}
    manual = _manual_product_info_object(fields.get(C.F_WORKSTREAM_CARD_MANUAL_PRODUCT_INFO) or fields.get(C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO))
    fallback_name = fallback_entry_fields.get(C.F_RECEIPT_ENTRY_NAME) or fields.get(C.F_RECEIPT_ENTRY_NAME) or ""
    fallback_identifier = fallback_entry_fields.get(C.F_RECEIPT_ENTRY_SKU_ID) or fields.get(C.F_RECEIPT_ENTRY_SKU_ID) or ""
    fallback_description = fallback_entry_fields.get(C.F_RECEIPT_ENTRY_DESCRIPTION) or fields.get(C.F_RECEIPT_ENTRY_DESCRIPTION) or ""
    if fallback_name and not (manual.get("name") or manual.get("productName") or manual.get("product")):
        manual["name"] = fallback_name
    if fallback_identifier and not (manual.get("upc") or manual.get("primaryMatchKey") or manual.get("identifier") or manual.get("productId")):
        manual["upc"] = fallback_identifier
    if fallback_description and not (manual.get("description") or manual.get("productDescription")):
        manual["description"] = fallback_description
    return manual


def _manual_product_info_from_body(body):
    for key in ("manualProductInfo", "manual_product_info", "productInfo", "product_info"):
        if key in body:
            return _json_text(body.get(key))
    return ""


def _workstream_name(entry_fields, workstream_type):
    package_name = str(entry_fields.get(C.F_RECEIPT_ENTRY_NAME) or "Received Merch").strip()
    identifier = str(entry_fields.get(C.F_RECEIPT_ENTRY_SKU_ID) or "").strip()
    suffix = f" - {identifier}" if identifier else ""
    return f"{package_name}{suffix} - {workstream_type}"


def _thr3d_shipping_item_name(entry_fields):
    package_name = str(entry_fields.get(C.F_RECEIPT_ENTRY_NAME) or "Received Merch").strip()
    identifier = str(entry_fields.get(C.F_RECEIPT_ENTRY_SKU_ID) or "").strip()
    suffix = f" - {identifier}" if identifier else ""
    return f"{package_name}{suffix} - THR3D"


_WORKSTREAM_META_PREFIX = "<!-- marks-photo-workstream-meta "
_WORKSTREAM_META_SUFFIX = " -->"


def _workstream_notes_metadata(value):
    raw = str(value or "")
    start = raw.rfind(_WORKSTREAM_META_PREFIX)
    if start < 0 or not raw.endswith(_WORKSTREAM_META_SUFFIX):
        return raw, {}
    payload = raw[start + len(_WORKSTREAM_META_PREFIX):-len(_WORKSTREAM_META_SUFFIX)]
    try:
        metadata = json.loads(payload)
    except (TypeError, ValueError, json.JSONDecodeError):
        return raw, {}
    if not isinstance(metadata, dict):
        return raw, {}
    visible = raw[:start].rstrip()
    return visible, metadata


def _workstream_notes_with_metadata(value, metadata):
    visible, _ = _workstream_notes_metadata(value)
    if not metadata:
        return visible
    return f"{visible}\n\n{_WORKSTREAM_META_PREFIX}{json.dumps(metadata, separators=(',', ':'))}{_WORKSTREAM_META_SUFFIX}".strip()


def _shape_workstream_card(record):
    fields = record.get("fields", {})
    creative_force_sync = _parse_creative_force_sync(fields.get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, ""))
    visible_notes, _ = _workstream_notes_metadata(fields.get(C.F_WORKSTREAM_CARD_NOTES, ""))
    return {
        "id": record.get("id", ""),
        "name": fields.get(C.F_WORKSTREAM_CARD_NAME, ""),
        "receivedMerchIds": _as_list(fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])),
        "expectedProductIds": _as_list(fields.get(C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT, [])),
        "type": fields.get(C.F_WORKSTREAM_CARD_TYPE, ""),
        "planningStatus": _planning_status_for_fields(fields),
        "quantity": fields.get(C.F_WORKSTREAM_CARD_QUANTITY, 0),
        "manualProductInfo": fields.get(C.F_WORKSTREAM_CARD_MANUAL_PRODUCT_INFO, ""),
        "notes": visible_notes,
        "creativeForce": creative_force_sync,
        "creativeForceStatus": fields.get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_STATUS, ""),
        "creativeForceStep": fields.get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_STEP, ""),
        "released": bool(fields.get(C.F_WORKSTREAM_CARD_RELEASED, False)),
        "releasedAt": fields.get(C.F_WORKSTREAM_CARD_RELEASED_AT, ""),
        "releasedByIds": _as_list(fields.get(C.F_WORKSTREAM_CARD_RELEASED_BY, [])),
    }


def _parse_creative_force_sync(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {"syncError": "Malformed Creative Force sync data."}
    return parsed if isinstance(parsed, dict) else {"syncError": "Malformed Creative Force sync data."}


def _creative_force_status(action, raw_status):
    action_value = str(action or "").strip().lower()
    raw_value = str(raw_status or "").strip()
    value = raw_value.lower().replace("_", " ")
    compact_value = re.sub(r"[^a-z]+", "", value)
    if action_value in {"completed", "workunitcompleted"} or value in {"done", "completed"}:
        return "Complete"
    if action_value in {"rejected", "failed", "workunitreset"} or value in {"rejected", "failed", "error"}:
        return "Blocked"
    if action_value in {"started", "inprogress", "statuschanged", "workunitstatuschanged"} and (value in {"inprogress", "in progress", "shooting"} or compact_value == "inprogress"):
        return "In Production"
    if value in {"todo", "to do", "assigned", "readytowork", "ready to work", "scheduled"}:
        return "Scheduled"
    if value in {"backlog", "new", "created"}:
        return "Accepted"
    return raw_status or "Unknown"


def _creative_force_sync_from_payload(payload):
    action = payload.get("Action") or payload.get("action") or ""
    raw_status = payload.get("WorkUnitStatusName") or payload.get("workUnitStatusName") or ""
    raw_step_status = payload.get("StepStatusName") or payload.get("stepStatusName") or ""
    step_name = payload.get("StepName") or payload.get("stepName") or ""
    event_time = payload.get("EventDatetimeUtc") or payload.get("eventDatetimeUtc") or ""
    if isinstance(event_time, (int, float)):
        event_time = datetime.fromtimestamp(event_time / 1000, tz=timezone.utc).isoformat()
    return {
        "system": "creative_force",
        "workUnitId": payload.get("WorkUnitId") or payload.get("workUnitId") or "",
        "productId": payload.get("ProductId") or payload.get("productId") or "",
        "productCode": payload.get("ProductCode") or payload.get("productCode") or "",
        "productionTypeId": (payload.get("ProductionTypeId") or payload.get("productionTypeId")
                             or payload.get("ShootingTypeId") or payload.get("shootingTypeId") or ""),
        "productionTypeName": (payload.get("ProductionTypeName") or payload.get("productionTypeName")
                               or payload.get("ShootingTypeName") or payload.get("shootingTypeName") or ""),
        "taskId": payload.get("TaskId") or payload.get("taskId") or "",
        "workflowId": payload.get("WorkflowId") or payload.get("workflowId") or "",
        "workflowName": payload.get("WorkflowName") or payload.get("workflowName") or "",
        "jobCode": payload.get("JobCode") or payload.get("jobCode") or "",
        "stepId": payload.get("StepId") or payload.get("stepId") or "",
        "eventGroup": payload.get("EventGroupName") or payload.get("eventGroupName") or "",
        "status": _creative_force_status(action, raw_status),
        "statusRaw": raw_status,
        "stepName": step_name,
        "stepStatusRaw": raw_step_status,
        "action": action,
        "lastReportedAt": event_time,
        "lastSyncedAt": _now_iso(),
        "payloadId": payload.get("PayloadId") or payload.get("payloadId") or "",
    }


def _creative_force_product_code_for_card(record):
    fields = record.get("fields", {})
    product_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT, [])) or [""])[0]
    if not product_id:
        return ""
    product_record = airtable.get_record(C.PRODUCTS_TABLE, product_id, by_field_id=False)
    product_fields = product_record.get("fields", {})
    client_id = (_as_list(product_fields.get(C.F_ITEM_CLIENT, [])) or [""])[0]
    if not client_id:
        return ""
    client_record = airtable.get_record(C.CLIENTS_TABLE, client_id, by_field_id=False)
    client_fields = client_record.get("fields", {})
    requirements, _ = _parse_photo_production_requirements(
        client_fields.get(C.F_CLIENT_PHOTO_PRODUCTION_REQUIREMENTS, ""),
        client_fields.get(C.F_CLIENT_NAME, ""),
    )
    workstream_type = fields.get(C.F_WORKSTREAM_CARD_TYPE, "")
    workstream = (requirements.get("workstreams") or {}).get(workstream_type) or {}
    product_code_field = (workstream.get("creativeForce") or {}).get("productCodeField") or ""
    if not product_code_field:
        return ""
    product = {
        "name": product_fields.get(C.F_ITEM_NAME, ""),
        "cvid": product_fields.get(C.F_ITEM_CVID, ""),
        "upc": product_fields.get(C.F_ITEM_UPC, ""),
        "jobNumber": product_fields.get(C.F_ITEM_WKFT_JOB_NUMBER, ""),
        "brandPrefix": product_fields.get(C.F_ITEM_BRAND_PREFIX, ""),
    }
    return str(_photo_product_value(product, product_code_field) or "").strip()


CREATIVE_FORCE_DONE_STATUSES = {"done", "completed", "complete", "approved"}


def _creative_force_client_is_ours(payload):
    """Creative Force sends every client's events to the same endpoint.

    Without this, an event is only ignored because its Product Code happens not to
    match one of our cards — which would stop being true the moment two clients share
    a code. Configure by ClientId; the name is a fallback for readability.
    """
    allowed_ids = [value.casefold() for value in C.CREATIVE_FORCE_CLIENT_IDS]
    allowed_names = [value.casefold() for value in C.CREATIVE_FORCE_CLIENT_NAMES]
    if not allowed_ids and not allowed_names:
        return True, ""
    client_id = str(payload.get("ClientId") or payload.get("clientId") or "").strip()
    client_name = str(payload.get("ClientName") or payload.get("clientName") or "").strip()
    if client_id.casefold() in allowed_ids or client_name.casefold() in allowed_names:
        return True, ""
    return False, client_name or client_id or "unknown client"


def _creative_force_steps_after_event(existing_steps, sync):
    """Per-step state keyed by Creative Force's StepId, which orders the workflow."""
    steps = {str(key): dict(value) for key, value in (existing_steps or {}).items()}
    step_id = str(sync.get("stepId") or "").strip()
    step_name = str(sync.get("stepName") or "").strip()
    if not step_id and not step_name:
        return steps
    key = step_id or step_name
    steps[key] = {
        "name": step_name or steps.get(key, {}).get("name", ""),
        "status": sync.get("stepStatusRaw") or "",
        "reportedAt": sync.get("lastReportedAt") or sync.get("lastSyncedAt") or "",
    }
    return steps


def _creative_force_current_step(steps):
    """The step Creative Force reported most recently.

    StepId is not workflow order: a Photo Review of id 15 can finish before an
    External Post Production of id 7 begins. Nor can a step be judged finished by
    its own status, because Creative Force never reports a completion for a step
    it has moved past — those sit at In Progress indefinitely. What it does report
    is each transition as it happens, so the newest report is the current step.

    Timestamps carry microseconds, so ties are only produced by one action firing
    the whole chain at once: a reset. Work resumes from the first step, which
    CREATIVE_FORCE_STEP_ORDER names because Creative Force does not encode its
    ordering in StepId. A step missing from that list sorts after the named ones,
    by StepId, so an unconfigured workflow still resolves deterministically.
    """
    if not steps:
        return {}

    order = [name.casefold() for name in C.CREATIVE_FORCE_STEP_ORDER]

    def workflow_position(item):
        key, step = item
        name = str(step.get("name") or "").strip().casefold()
        try:
            step_id = int(key)
        except (TypeError, ValueError):
            step_id = 0
        return (order.index(name), 0) if name in order else (len(order), step_id)

    def reported(item):
        return str(item[1].get("reportedAt") or "")

    latest = max((reported(item) for item in steps.items()), default="")
    newest = [item for item in steps.items() if reported(item) == latest]
    if len(newest) == 1:
        return newest[0][1]
    return sorted(newest, key=workflow_position)[0][1]


def _creative_force_event_is_derived(sync, existing):
    """True when an event belongs to a workflow other than the item's main one.

    Creative Force sends an all-zero WorkflowId for derived workflows as well as
    the main one, so only the name separates them. `CREATIVE_FORCE_MAIN_WORKFLOW_NAME`
    names the main workflow when it is known; otherwise Creative Force's own default
    naming for derived workflows is used.

    Events without a workflow name are judged against the main work unit a named
    event has already identified. Before anything is known nothing is rejected,
    or the card would stay blank waiting for a named event to arrive first.
    """
    workflow_name = str(sync.get("workflowName") or "").strip().casefold()
    if workflow_name:
        configured_main = str(C.CREATIVE_FORCE_MAIN_WORKFLOW_NAME or "").strip().casefold()
        if configured_main:
            return workflow_name != configured_main
        return workflow_name.startswith("derived workflow")
    main_work_unit = str(existing.get("mainWorkUnitId") or "").strip()
    if not main_work_unit:
        return False
    return str(sync.get("workUnitId") or "").strip() != main_work_unit


def _find_creative_force_card(records, sync):
    """Match a first CF event by its handoff identity, then use WorkUnitId thereafter."""
    work_unit_id = str(sync.get("workUnitId") or "").strip()
    if work_unit_id:
        linked = next((record for record in records if _parse_creative_force_sync(
            record.get("fields", {}).get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, "")
        ).get("workUnitId") == work_unit_id), None)
        if linked:
            return linked

    product_code = str(sync.get("productCode") or "").strip().casefold()
    production_type = str(sync.get("productionTypeName") or "").strip().casefold()
    if not product_code or not production_type:
        return None

    # The feed's Source Key is the authoritative card correlation key. CF does
    # not echo it in Work Unit events, so recover it by Product Code first.
    feed_matches = [record for record in _list_all_records(C.CREATIVE_FORCE_PRODUCT_FEED_TABLE)
                    if str(record.get("fields", {}).get(C.F_CF_FEED_PRODUCT_CODE) or "").strip().casefold() == product_code]
    source_keys = [str(record.get("fields", {}).get(C.F_CF_FEED_SOURCE_KEY) or "").strip()
                   for record in feed_matches]
    source_keys = [key for key in source_keys if key]
    if len(source_keys) == 1:
        source_match = next((record for record in records if record.get("id") == source_keys[0]), None)
        if source_match:
            return source_match

    matches = []
    code_matches = []
    for record in records:
        fields = record.get("fields", {})
        try:
            configured_code = _creative_force_product_code_for_card(record).casefold()
        except requests.HTTPError:
            continue
        if configured_code == product_code:
            code_matches.append(record)
            if str(fields.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip().casefold() == production_type:
                matches.append(record)
                if len(matches) > 1:
                    return None
    if len(matches) == 1:
        return matches[0]
    return code_matches[0] if len(code_matches) == 1 else None


def _workstream_card_context(record):
    fields = record.get("fields", {})
    merchandise_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])) or [""])[0]
    entry, receipt, access_error = _permitted_merchandise_or_error(merchandise_id) if merchandise_id else (None, None, None)
    if access_error:
        return None, None, None, access_error
    product_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT, []))
                  or _as_list((entry or {}).get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM, []))
                  or [""])[0]
    product_record = None
    if product_id:
        try:
            product_record = airtable.get_record(C.PRODUCTS_TABLE, product_id, by_field_id=False)
        except requests.HTTPError as error:
            return None, None, None, airtable_err(error)
    clients = _clients_by_id()
    client_id = (
        _as_list((entry or {}).get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
        or _as_list((receipt or {}).get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
        or [""]
    )[0]
    return entry, product_record, clients.get(client_id), None


def _creative_force_handoff(record):
    fields = record.get("fields", {})
    workstream_type = fields.get(C.F_WORKSTREAM_CARD_TYPE, "")
    entry, product_record, client, access_error = _workstream_card_context(record)
    if access_error:
        return access_error
    product = _shape_item(product_record, clients_by_id={client["id"]: client} if client else {}) if product_record else _manual_product_from_fields(fields, entry.get("fields", {}) if entry else {})
    requirements = (client or {}).get("photoProductionRequirements") or _empty_photo_production_requirements()
    config = (requirements.get("workstreams") or {}).get(workstream_type) or {}
    status = _photo_production_status(workstream_type, product, client)
    creative_force = config.get("creativeForce") or {}
    product_code_field = creative_force.get("productCodeField") or ""
    product_code = _photo_product_value(product, product_code_field) if product_code_field else ""
    category_field = creative_force.get("categoryField") or "clientName"
    category = creative_force.get("categoryValue", "") if category_field == "custom" else (
        client or {}).get("name", "") if category_field == "clientName" else _photo_product_value(product, category_field)
    payload = {
        "client": {"id": (client or {}).get("id", ""), "name": (client or {}).get("name", "")},
        "workstream": {"id": record.get("id", ""), "type": workstream_type, "quantity": fields.get(C.F_WORKSTREAM_CARD_QUANTITY, 0)},
        "product": {
            "marksPhotoId": (product_record or {}).get("id", ""),
            "name": product.get("name", ""),
            "upc": product.get("upc", ""),
            "cvid": product.get("cvid", ""),
            "jobNumber": _photo_product_value(product, "jobNumber"),
            # Only what this client's workstream configuration asks for. The
            # feed table is built from the same list, so projecting every known
            # field instead writes columns that do not exist and fails the
            # release outright.
            "requiredFields": {
                key: _photo_product_value(product, key)
                for key in _configured_product_field_keys(client, workstream_type)
            },
        },
        "creativeForce": {
            "productCode": str(product_code or "").strip(),
            "productCodeField": product_code_field,
            "category": str(category or "").strip(),
            "categoryField": category_field,
            "productionType": workstream_type,
        },
    }
    return {
        "ready": bool(status.get("ready")),
        "missing": [*status.get("productData", {}).get("missing", []), *status.get("fileNaming", {}).get("missing", []), *status.get("creativeForce", {}).get("missing", [])],
        "validation": status,
        "payload": payload,
    }


def _configured_product_field_keys(client, workstream_type):
    """The product fields this client requires for one workstream."""
    workstreams = ((client or {}).get("photoProductionRequirements") or {}).get("workstreams") or {}
    config = workstreams.get(workstream_type) or {}
    return [
        key for key in (config.get("requiredProductFields") or [])
        if key in PHOTO_PRODUCTION_REQUIREMENT_FIELDS
    ]


def _creative_force_feed_fields(record, handoff):
    payload = handoff.get("payload", {})
    product = payload.get("product", {})
    client = payload.get("client", {})
    creative_force = payload.get("creativeForce", {})
    fields = {
        C.F_CF_FEED_PRODUCT: product.get("name", ""),
        C.F_CF_FEED_CLIENT: client.get("name", ""),
        C.F_CF_FEED_PRODUCT_CODE: creative_force.get("productCode", ""),
        C.F_CF_FEED_CATEGORY: creative_force.get("category", ""),
        C.F_CF_FEED_PRODUCTION_TYPE: creative_force.get("productionType", ""),
        C.F_CF_FEED_SOURCE_KEY: record.get("id", ""),
    }
    for key, value in (product.get("requiredFields") or {}).items():
        label = C.CREATIVE_FORCE_FEED_PRODUCT_FIELDS.get(key)
        if label and value not in (None, ""):
            fields[label] = value
    return fields


def _creative_force_product_feed_preview():
    rows = []
    for card in _list_all_records(C.WORKSTREAM_CARDS_TABLE):
        card_fields = card.get("fields", {})
        if card_fields.get(C.F_WORKSTREAM_CARD_TYPE) not in C.WORKSTREAM_TYPE_OPTIONS:
            continue
        handoff = _creative_force_handoff(card)
        if not isinstance(handoff, dict):
            continue
        if card_fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) != PLANNING_STATUS_LABELS["awaiting-photo-release"]:
            continue
        feed_fields = _creative_force_feed_fields(card, handoff)
        rows.append({
            "sourceKey": card.get("id", ""),
            "workstreamType": card_fields.get(C.F_WORKSTREAM_CARD_TYPE, ""),
            "fields": feed_fields,
        })
    return rows


def _workstream_cards_for_merchandise(entry_id):
    """Photo cards belonging to one merchandise record."""
    cards = []
    for record in _list_all_records(C.WORKSTREAM_CARDS_TABLE):
        linked = _as_list(record.get("fields", {}).get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, []))
        if entry_id in linked:
            cards.append(record)
    return cards


def _sync_creative_force_product_feed_cards(cards):
    """Upsert photo-release cards into the CF feed as part of the release transition."""
    existing = _list_all_records(C.CREATIVE_FORCE_PRODUCT_FEED_TABLE)
    existing_by_key = {
        str(record.get("fields", {}).get(C.F_CF_FEED_SOURCE_KEY) or "").strip(): record
        for record in existing
    }
    synced = []
    for card in cards:
        card_fields = card.get("fields", {})
        if card_fields.get(C.F_WORKSTREAM_CARD_TYPE) not in C.WORKSTREAM_TYPE_OPTIONS:
            continue
        handoff = _creative_force_handoff(card)
        if not isinstance(handoff, dict):
            continue
        if card_fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) != PLANNING_STATUS_LABELS["awaiting-photo-release"]:
            continue
        feed_fields = _creative_force_feed_fields(card, handoff)
        source_key = feed_fields[C.F_CF_FEED_SOURCE_KEY]
        if source_key in existing_by_key:
            saved = airtable.update_record(
                C.CREATIVE_FORCE_PRODUCT_FEED_TABLE,
                existing_by_key[source_key]["id"],
                feed_fields,
                by_field_id=False,
                typecast=True,
            )
            action = "updated"
        else:
            saved = airtable.create_record(
                C.CREATIVE_FORCE_PRODUCT_FEED_TABLE,
                feed_fields,
                by_field_id=False,
                typecast=True,
            )
            action = "created"
        synced.append({"sourceKey": source_key, "action": action, "record": saved})
    return synced


def _populate_creative_force_feed_for_ready_cards(cards):
    """Keep the CF table current without exposing a second manual sync workflow."""
    _ensure_creative_force_feed_schema()
    return _sync_creative_force_product_feed_cards(cards)


@api.get("/integrations/creative-force/product-feed/preview")
def preview_creative_force_product_feed():
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    try:
        rows = _creative_force_product_feed_preview()
        existing = _list_all_records(C.CREATIVE_FORCE_PRODUCT_FEED_TABLE)
        table_provisioned = True
    except requests.HTTPError as error:
        response = getattr(error, "response", None)
        if getattr(response, "status_code", None) in {403, 404}:
            existing = []
            table_provisioned = False
        else:
            return airtable_err(error)
    return jsonify({
        "table": C.CREATIVE_FORCE_PRODUCT_FEED_TABLE,
        "tableProvisioned": table_provisioned,
        "rows": rows,
        "counts": {"ready": len(rows), "existing": len(existing)},
    })


def _forward_creative_force_event(body, signature, forwarded_from):
    """Relay one authentic event to a second instance, in the background.

    Sent verbatim with its signature, so the receiver validates exactly what
    Creative Force sent. Never raised and never awaited: the event is already
    accepted here, and a sleeping laptop must not cost production anything.

    The X-CF-Forwarded header stops a relay from being relayed onward, which
    would otherwise loop if both ends were ever configured to forward.
    """
    url = str(C.CREATIVE_FORCE_FORWARD_URL or "").strip()
    if not url or forwarded_from:
        return

    def send():
        try:
            requests.post(
                url,
                data=body,
                headers={
                    "Content-Type": "application/json",
                    "X-CF-Signature": signature,
                    "X-CF-Forwarded": "1",
                },
                timeout=8,
            )
        except Exception:  # noqa: BLE001 - a relay failure is not production's problem
            pass

    threading.Thread(target=send, daemon=True).start()


@api.post("/integrations/creative-force/webhook")
def creative_force_webhook():
    secret = str(C.CREATIVE_FORCE_WEBHOOK_SECRET or "")
    signature = request.headers.get("X-CF-Signature", "")
    if not secret:
        return err("Creative Force webhook is not configured.", 503)
    expected = hmac.new(secret.encode("utf-8"), request.get_data(), hashlib.sha256).hexdigest()
    if not signature or not hmac.compare_digest(expected, signature):
        return err("Invalid Creative Force webhook signature.", 401)
    # Relayed here rather than at each exit: the event is authentic, and every
    # branch below is a legitimate outcome a second instance should also see.
    _forward_creative_force_event(
        request.get_data(), signature, request.headers.get("X-CF-Forwarded", ""),
    )
    payload = request.get_json(silent=True) or {}
    sync = _creative_force_sync_from_payload(payload)
    received_at = _now_iso()
    ours, other_client = _creative_force_client_is_ours(payload)
    if not ours:
        _record_creative_force_webhook({
            "receivedAt": received_at, "payload": payload, "sync": sync,
            "accepted": True, "ignored": f"Event belongs to {other_client}.",
        })
        return jsonify({"accepted": True, "ignored": "other-client", "client": other_client})
    work_unit_id = sync.get("workUnitId")
    if not work_unit_id:
        _record_creative_force_webhook({
            "receivedAt": received_at,
            "payload": payload,
            "sync": sync,
            "accepted": False,
            "reason": "Missing WorkUnitId.",
        })
        return jsonify({"accepted": False, "reason": "Missing WorkUnitId."})
    records = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
    target = _find_creative_force_card(records, sync)
    if not target:
        _record_creative_force_webhook({
            "receivedAt": received_at,
            "payload": payload,
            "sync": sync,
            "accepted": False,
            "reason": "No unique matching workstream card.",
            "workUnitId": work_unit_id,
        })
        return jsonify({"accepted": False, "reason": "No unique matching workstream card.", "workUnitId": work_unit_id})
    existing = _parse_creative_force_sync(target.get("fields", {}).get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, ""))
    if sync.get("payloadId") and sync.get("payloadId") == existing.get("payloadId"):
        _record_creative_force_webhook({
            "receivedAt": received_at,
            "payload": payload,
            "sync": sync,
            "accepted": True,
            "duplicate": True,
            "workUnitId": work_unit_id,
            "workstreamCardId": target.get("id", ""),
        })
        return jsonify({"accepted": True, "duplicate": True, "workUnitId": work_unit_id})
    if _creative_force_event_is_derived(sync, existing):
        _record_creative_force_webhook({
            "receivedAt": received_at, "payload": payload, "sync": sync,
            "accepted": True, "ignored": "Event belongs to a derived workflow.",
            "workUnitId": work_unit_id, "workstreamCardId": target.get("id", ""),
        })
        return jsonify({"accepted": True, "ignored": "derived-workflow", "workUnitId": work_unit_id})

    if str(sync.get("workflowName") or "").strip():
        sync["mainWorkUnitId"] = sync.get("workUnitId") or ""
    same_work_unit = str(existing.get("workUnitId") or "").strip() == str(sync.get("workUnitId") or "").strip()
    carried = existing if same_work_unit else {}
    merged = {**carried, **{key: value for key, value in sync.items() if value not in ("", None)}}
    merged["steps"] = _creative_force_steps_after_event(carried.get("steps"), sync)
    current = _creative_force_current_step(merged["steps"])
    if current:
        merged["stepName"] = current.get("name") or merged.get("stepName", "")
        merged["stepStatusRaw"] = current.get("status") or ""
        merged["stepReportedAt"] = current.get("reportedAt") or ""
    if not same_work_unit and existing.get("mainWorkUnitId"):
        merged.setdefault("mainWorkUnitId", existing["mainWorkUnitId"])
    updated = airtable.update_record(
        C.WORKSTREAM_CARDS_TABLE,
        target["id"],
        {
            C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC: json.dumps(merged, sort_keys=True),
            C.F_WORKSTREAM_CARD_CREATIVE_FORCE_STATUS: (
                merged.get("stepStatusRaw") or merged.get("statusRaw", "")
            ),
            C.F_WORKSTREAM_CARD_CREATIVE_FORCE_STEP: merged.get("stepName", ""),
        },
        by_field_id=False,
        typecast=True,
    )
    _record_creative_force_webhook({
        "receivedAt": received_at,
        "payload": payload,
        "sync": sync,
        "accepted": True,
        "workUnitId": work_unit_id,
        "workstreamCardId": target.get("id", ""),
        "creativeForceStatus": merged.get("stepStatusRaw") or merged.get("statusRaw", ""),
        "workUnitStatus": merged.get("statusRaw", ""),
        "creativeForceStep": merged.get("stepName", ""),
    })
    return jsonify({"accepted": True, "workUnitId": work_unit_id, "record": _shape_workstream_card(updated)})


@api.get("/integrations/creative-force/webhook/diagnostics")
def creative_force_webhook_diagnostics():
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    return jsonify({
        "webhook": _CREATIVE_FORCE_LAST_WEBHOOK,
        "recent": [
            {
                "receivedAt": entry.get("receivedAt", ""),
                "action": (entry.get("payload") or {}).get("Action", ""),
                "eventGroup": (entry.get("payload") or {}).get("EventGroupName", ""),
                "stepName": (entry.get("sync") or {}).get("stepName", ""),
                "stepStatus": (entry.get("sync") or {}).get("stepStatusRaw", ""),
                "workUnitStatus": (entry.get("sync") or {}).get("statusRaw", ""),
                "workflowName": (entry.get("sync") or {}).get("workflowName", ""),
                "outcome": (
                    entry.get("ignored") or entry.get("reason")
                    or ("Duplicate" if entry.get("duplicate") else None)
                    or ("Written" if entry.get("accepted") else "Rejected")
                ),
            }
            for entry in _CREATIVE_FORCE_RECENT_WEBHOOKS
        ],
    })


@api.get("/workstream-cards/<record_id>/creative-force-handoff")
def creative_force_handoff(record_id):
    try:
        record = airtable.get_record(C.WORKSTREAM_CARDS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    merchandise_id = (_as_list(record.get("fields", {}).get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])) or [""])[0]
    if merchandise_id:
        _, _, access_error = _permitted_merchandise_or_error(merchandise_id)
        if access_error:
            return access_error
    result = _creative_force_handoff(record)
    if isinstance(result, tuple):
        return result
    return jsonify(result)


@api.patch("/workstream-cards/<record_id>/creative-force-link")
def link_creative_force_work_unit(record_id):
    admin_error = _require_admin()
    if admin_error:
        return admin_error
    body = request.get_json(silent=True) or {}
    work_unit_id = str(body.get("workUnitId") or "").strip()
    if not work_unit_id:
        return err("workUnitId is required.")
    try:
        record = airtable.get_record(C.WORKSTREAM_CARDS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    handoff = _creative_force_handoff(record)
    if isinstance(handoff, tuple):
        return handoff
    if not handoff.get("ready"):
        return jsonify({"error": "Creative Force handoff is not ready.", "missing": handoff.get("missing", []), "handoff": handoff}), 400
    sync = _parse_creative_force_sync(record.get("fields", {}).get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, ""))
    sync.update({
        "system": "creative_force",
        "workUnitId": work_unit_id,
        "productId": body.get("productId") or sync.get("productId", ""),
        "productCode": handoff["payload"]["creativeForce"]["productCode"],
        "productionTypeName": handoff["payload"]["creativeForce"]["productionType"],
        "linkedAt": _now_iso(),
        "syncError": "",
    })
    try:
        updated = airtable.update_record(C.WORKSTREAM_CARDS_TABLE, record_id, {
            C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC: json.dumps(sync, sort_keys=True),
        }, by_field_id=False, typecast=True)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"record": _shape_workstream_card(updated), "handoff": handoff})


def _product_view_for_requirements(product_fields):
    """Raw Airtable Product fields in the shape the requirement resolver reads."""
    fields = product_fields or {}
    return {
        "name": fields.get(C.F_ITEM_NAME, ""),
        "upc": fields.get(C.F_ITEM_UPC, ""),
        "identifier": fields.get(C.F_ITEM_IDENTIFIER, ""),
        "cvid": fields.get(C.F_ITEM_CVID, ""),
        "itemJobNumber": fields.get(C.F_ITEM_JOB_NUMBER, ""),
        "brandPrefix": fields.get(C.F_ITEM_BRAND_PREFIX, ""),
        "fileNameDescription": fields.get(C.F_ITEM_FILE_NAME_DESCRIPTION, ""),
        "productType": fields.get(C.F_ITEM_PRODUCT_TYPE, ""),
        "ecommPhotoNotes": fields.get(C.F_ITEM_ECOMM_PHOTO_NOTES, ""),
        "pathToArt": fields.get(C.F_ITEM_PATH_TO_ART, ""),
        "referenceData": _parse_reference_data(fields.get(C.F_ITEM_REFERENCE_DATA, "")),
    }


def _photo_product_value(product, key):
    product = product or {}
    reference_data = product.get("referenceData") or {}
    direct_values = {
        "productName": product.get("name") or product.get("product") or product.get("productName"),
        "upc": product.get("upc") or product.get("primaryMatchKey") or product.get("identifier"),
        "cvid": product.get("cvid"),
        "jobNumber": product.get("itemJobNumber") or product.get("jobNumber") or product.get("wkftJobNumber") or product.get("pickupJobNumber"),
        "brandPrefix": product.get("brandPrefix"),
        "productType": product.get("productType"),
        "ecommPhotoNotes": product.get("ecommPhotoNotes"),
        "pathToArt": product.get("pathToArt"),
        "fileNameDescription": product.get("fileNameDescription"),
    }
    if key in direct_values and direct_values[key] not in (None, ""):
        return direct_values[key]
    if key == "upc":
        return product.get("productId") or product.get("skuId") or ""
    if key == "brandPrefix":
        return product.get("brand") or ""
    if key == "fileNameDescription":
        candidates = {re.sub(r"[^a-z0-9]+", "", str(name).lower()): value for name, value in reference_data.items()}
        for candidate in [
            "filenamedescriptionmadebycoordinatornotfilenames",
            "filenamedescriptioncvidnumber",
            "filenamedescription",
            "proddescrip",
            "productdescription",
        ]:
            if candidates.get(candidate) not in (None, ""):
                return candidates[candidate]
        return product.get("fileNameDescription") or ""
    return ""


def _photo_production_status(workstream_type, product, client):
    requirements = (client or {}).get("photoProductionRequirements") or _empty_photo_production_requirements()
    config = (requirements.get("workstreams") or {}).get(workstream_type) or {}
    required_fields = config.get("requiredProductFields") or []
    checks = []
    for key in required_fields:
        value = _photo_product_value(product, key)
        present = _photo_production_value_is_valid(key, value)
        checks.append({
            "key": key,
            "label": PHOTO_PRODUCTION_REQUIREMENT_FIELDS.get(key, key),
            "present": present,
        })
    naming = config.get("naming") or {}
    naming_checks = []
    for key in naming.get("tokens") or []:
        if key == "view":
            value = ", ".join(naming.get("views") or [])
        else:
            value = _photo_product_value(product, key)
        naming_checks.append({
            "key": key,
            "label": PHOTO_PRODUCTION_REQUIREMENT_FIELDS.get(key, "View" if key == "view" else key),
            "present": bool(str(value or "").strip()),
        })
    missing = [check["label"] for check in checks if not check["present"]]
    missing_naming = [check["label"] for check in naming_checks if not check["present"]]
    creative_force_config = config.get("creativeForce")
    creative_force_checks = []
    if creative_force_config is not None:
        product_code_field = creative_force_config.get("productCodeField") or ""
        category_field = creative_force_config.get("categoryField") or creative_force_config.get("categorySource") or "clientName"
        product_code_value = _photo_product_value(product, product_code_field) if product_code_field else ""
        category_value = (
            creative_force_config.get("categoryValue", "")
            if category_field == "custom"
            else (client or {}).get("name", "")
            if category_field == "clientName"
            else _photo_product_value(product, category_field)
        )
        creative_force_checks = [
            {"key": "productCodeField", "label": "Product Code mapping", "present": bool(product_code_field)},
            {"key": "productCode", "label": "Creative Force Product Code", "present": bool(str(product_code_value or "").strip())},
            {"key": "category", "label": "Creative Force Category", "present": bool(str(category_value or "").strip())},
        ]
    creative_force_missing = [check["label"] for check in creative_force_checks if not check["present"]]
    creative_force_ready = not creative_force_missing
    if creative_force_config is None:
        category_value = ""
    return {
        "workstreamType": workstream_type,
        "ready": not missing and not missing_naming and creative_force_ready,
        "productData": {"ready": not missing, "checks": checks, "missing": missing},
        "fileNaming": {
            "ready": not missing_naming,
            "template": naming.get("template", ""),
            "checks": naming_checks,
            "missing": missing_naming,
            "views": naming.get("views") or [],
        },
        "creativeForce": {
            "ready": creative_force_ready,
            "productCodeField": (creative_force_config or {}).get("productCodeField", "") if creative_force_config is not None else "",
            "categoryField": (creative_force_config or {}).get("categoryField", "clientName") if creative_force_config is not None else "clientName",
            "categoryValue": category_value,
            "checks": creative_force_checks,
            "missing": creative_force_missing,
        },
    }


def _photo_production_value_is_valid(key, value):
    text = str(value or "").strip()
    if not text:
        return False
    return True


PLANNING_STATUS_VALUES = ("new", "needs-more-information", "awaiting-photo-release")
WORKSTREAM_CARD_PLANNING_STATUS_VALUES = ("needs-more-information", "awaiting-photo-release")
PLANNING_STATUS_LABELS = {
    "new": "New",
    "needs-more-information": "Needs More Information",
    "awaiting-photo-release": "Awaiting Photo Release",
}


def _normalized_planning_status(value):
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.lower().replace("_", "-")
    aliases = {
        "needs more information": "needs-more-information",
        "awaiting photo release": "awaiting-photo-release",
    }
    return aliases.get(normalized, normalized)


def _planning_status_for_fields(fields=None):
    """Return the single API-owned Planning status used by the board.

    Airtable still stores the older intake and child-card fields for compatibility,
    but callers should use this normalized value for queue placement.
    """
    fields = fields or {}
    explicit_planning_status = _normalized_planning_status(
        fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS)
        or fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS, "")
    )
    if explicit_planning_status in PLANNING_STATUS_VALUES:
        return explicit_planning_status
    if explicit_planning_status:
        # A stored value outside the canonical set has no board placement.
        return ""
    # No stored Planning Status at all. A newly received item stays in New Merch
    # until a PM accepts it; anything already accepted belongs in the review lane.
    # Nothing else is inferred here: board placement follows the stored field and,
    # for parent merchandise, whether child work exists.
    return "needs-more-information" if fields.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED, False) else "new"


def _shape_thr3d_shipping_item(record):
    fields = record.get("fields", {})
    return {
        "id": record.get("id", ""),
        "name": fields.get(C.F_THR3D_SHIPPING_ITEM_NAME, ""),
        "receivedMerchIds": _as_list(fields.get(C.F_THR3D_SHIPPING_ITEM_RECEIVED_MERCH, [])),
        "expectedProductIds": _as_list(fields.get(C.F_THR3D_SHIPPING_ITEM_EXPECTED_PRODUCT, [])),
        "quantityToShip": fields.get(C.F_THR3D_SHIPPING_ITEM_QUANTITY, 0),
        "shippingStatus": fields.get(C.F_THR3D_SHIPPING_ITEM_STATUS, ""),
        "outboundShipmentIds": _as_list(fields.get(C.F_THR3D_SHIPPING_ITEM_OUTBOUND_SHIPMENT, [])),
        "manualProductInfo": fields.get(C.F_THR3D_SHIPPING_ITEM_MANUAL_PRODUCT_INFO, ""),
        "notes": fields.get(C.F_THR3D_SHIPPING_ITEM_NOTES, ""),
    }


def _workstream_cards_for_planning():
    records = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
    shaped_records = []
    clients_by_id = _clients_by_id()
    for record in records:
        fields = record.get("fields", {})
        merchandise_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])) or [""])[0]
        if not merchandise_id:
            continue
        entry, receipt, access_error = _permitted_merchandise_or_error(merchandise_id)
        if access_error:
            continue
        product_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT, [])) or _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM, [])) or [""])[0]
        product_record = None
        if product_id:
            try:
                product_record = airtable.get_record(C.PRODUCTS_TABLE, product_id, by_field_id=False)
            except requests.HTTPError:
                product_record = None
        client_id = (
            _as_list(entry.get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
            or _as_list(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
            or [""]
        )[0]
        client = clients_by_id.get(client_id)
        expected_product = _shape_item(product_record, clients_by_id=clients_by_id) if product_record else None
        product_for_photo = expected_product or _manual_product_from_fields(fields, entry.get("fields", {}))
        photo_status = _photo_production_status(fields.get(C.F_WORKSTREAM_CARD_TYPE, ""), product_for_photo, client)
        shaped_records.append({
            **_shape_workstream_card(record),
            "planningStatus": _planning_status_for_fields(
                {**entry.get("fields", {}), **fields},
            ),
            "receivedMerch": _shape_verification_entry(entry, receipt, item_record=product_record),
            "expectedProduct": expected_product,
            "photoProduction": photo_status,
        })
    shaped_records.sort(key=lambda item: (item.get("status") or "", item.get("type") or "", item.get("name") or ""))
    return shaped_records


def _thr3d_shipping_items_for_shipments(records=None, *, include_shipped=False, parent_cache=None):
    records = records if records is not None else _list_all_records(C.THR3D_SHIPPING_ITEMS_TABLE)
    parent_cache = parent_cache if parent_cache is not None else {}
    shaped_records = []
    for record in records:
        fields = record.get("fields", {})
        is_shipped = str(fields.get(C.F_THR3D_SHIPPING_ITEM_STATUS) or "").strip().lower() == "shipped"
        if is_shipped != include_shipped:
            continue
        merchandise_id = (_as_list(fields.get(C.F_THR3D_SHIPPING_ITEM_RECEIVED_MERCH, [])) or [""])[0]
        parent = None
        if merchandise_id:
            if merchandise_id not in parent_cache:
                parent_cache[merchandise_id] = _permitted_merchandise_or_error(merchandise_id)
            entry, receipt, access_error = parent_cache[merchandise_id]
            if access_error:
                continue
            parent = _shape_verification_entry(entry, receipt)
        shaped_records.append({
            **_shape_thr3d_shipping_item(record),
            "receivedMerch": parent,
        })
    shaped_records.sort(key=lambda item: (item.get("shippingStatus") or "", item.get("name") or ""))
    return shaped_records


def _mark_thr3d_shipping_item_shipped(record_id, body):
    try:
        shipping_item = airtable.get_record(C.THR3D_SHIPPING_ITEMS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    fields = shipping_item.get("fields", {})
    merchandise_id = (_as_list(fields.get(C.F_THR3D_SHIPPING_ITEM_RECEIVED_MERCH, [])) or [""])[0]
    if not merchandise_id:
        return err("THR3D Shipping Item must be linked to Received Merch.", 400)
    entry, receipt, access_error = _permitted_merchandise_or_error(merchandise_id)
    if access_error:
        return access_error
    try:
        carrier = _normalize_receipt_carrier(body.get("carrier"))
    except ValueError as error:
        return err(str(error))
    tracking = (body.get("tracking") or "").strip()
    if not carrier:
        return err("Carrier is required.")
    if not tracking:
        return err("Tracking is required.")

    item_name = fields.get(C.F_THR3D_SHIPPING_ITEM_NAME) or entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_NAME) or record_id
    shipment_fields = {
        C.F_RECEIPT_NAME: f"THR3D outbound - {item_name}",
        C.F_RECEIPT_CARRIER: carrier,
        C.F_RECEIPT_TRACKING: tracking,
        C.F_RECEIPT_BOX_QUANTITY: 1,
        C.F_RECEIPT_RECEIVED: _now_iso(),
        C.F_RECEIPT_NOTES: f"Outbound THR3D shipment for {item_name}.",
    }
    receipt_fields = receipt.get("fields", {}) if receipt else {}
    client_ids = _as_list(receipt_fields.get(C.F_RECEIPT_CLIENT, []))
    if client_ids:
        shipment_fields[C.F_RECEIPT_CLIENT] = client_ids
    current_user_id = _current_user_id()
    if current_user_id:
        shipment_fields[C.F_RECEIPT_RECEIVER] = [current_user_id]

    try:
        outbound_shipment = airtable.create_record(C.SHIPMENTS_TABLE, shipment_fields, by_field_id=False, typecast=True)
        updated_item = airtable.update_record(
            C.THR3D_SHIPPING_ITEMS_TABLE,
            record_id,
            {
                C.F_THR3D_SHIPPING_ITEM_STATUS: "Shipped",
                C.F_THR3D_SHIPPING_ITEM_OUTBOUND_SHIPMENT: [outbound_shipment["id"]],
            },
            by_field_id=False,
            typecast=True,
        )
        parent_fields = entry.get("fields", {})
        try:
            quantity_to_ship = int(fields.get(C.F_THR3D_SHIPPING_ITEM_QUANTITY) or 0)
            parent_quantity = int(parent_fields.get(C.F_RECEIPT_ENTRY_QUANTITY) or 0)
        except (TypeError, ValueError):
            quantity_to_ship = 0
            parent_quantity = 0
        updated_entry = entry
        if quantity_to_ship and parent_quantity and quantity_to_ship >= parent_quantity:
            updated_entry = _update_receipt_entry_record(merchandise_id, {C.F_RECEIPT_ENTRY_MERCH_STATUS: "Shipped"})
    except requests.HTTPError as error:
        return airtable_err(error)

    return jsonify({
        "record": {
            **_shape_thr3d_shipping_item(updated_item),
            "receivedMerch": _shape_verification_entry(updated_entry, receipt),
            "outboundShipment": _shape_receipt(outbound_shipment, entries_by_receipt={outbound_shipment["id"]: []}),
        }
    })


def _confirm_assign_payload(body, entry, item_record=None, card_planning_status=None):
    entry_fields = entry.get("fields", {})
    card_status = card_planning_status or PLANNING_STATUS_LABELS["needs-more-information"]
    manual_product_info = _manual_product_info_from_body(body)
    expected_product_ids = []
    expected_product_id = str(body.get("expectedProductId") or body.get("productId") or body.get("itemId") or "").strip()
    if expected_product_id:
        expected_product_ids = [expected_product_id]
    elif item_record:
        expected_product_ids = [item_record.get("id")]

    workstreams = []
    seen_types = set()
    for raw in body.get("workstreams") or body.get("workstreamCards") or []:
        workstream_type = str((raw or {}).get("type") or (raw or {}).get("workstreamType") or "").strip()
        if workstream_type not in C.WORKSTREAM_TYPE_OPTIONS:
            return err(f"Workstream Type must be one of: {', '.join(C.WORKSTREAM_TYPE_OPTIONS)}.")
        if workstream_type in seen_types:
            return err(f"{workstream_type} can only be assigned once.")
        quantity = _positive_int((raw or {}).get("quantity", entry_fields.get(C.F_RECEIPT_ENTRY_QUANTITY, 1)), f"{workstream_type} Quantity")
        if isinstance(quantity, tuple):
            return quantity
        seen_types.add(workstream_type)
        workstreams.append({
            C.F_WORKSTREAM_CARD_NAME: _workstream_name(entry_fields, workstream_type),
            C.F_WORKSTREAM_CARD_RECEIVED_MERCH: [entry["id"]],
            C.F_WORKSTREAM_CARD_TYPE: workstream_type,
            C.F_WORKSTREAM_CARD_PLANNING_STATUS: card_status,
            C.F_WORKSTREAM_CARD_QUANTITY: quantity,
            **({C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT: expected_product_ids} if expected_product_ids else {}),
            **({C.F_WORKSTREAM_CARD_MANUAL_PRODUCT_INFO: manual_product_info} if manual_product_info else {}),
        })

    thr3d_raw = body.get("thr3d") or body.get("thr3dShippingItem")
    thr3d_fields = None
    if isinstance(thr3d_raw, dict) and thr3d_raw.get("quantity") not in (None, "", 0, "0"):
        quantity = _positive_int(thr3d_raw.get("quantity"), "THR3D Quantity")
        if isinstance(quantity, tuple):
            return quantity
        thr3d_fields = {
            C.F_THR3D_SHIPPING_ITEM_NAME: _thr3d_shipping_item_name(entry_fields),
            C.F_THR3D_SHIPPING_ITEM_RECEIVED_MERCH: [entry["id"]],
            C.F_THR3D_SHIPPING_ITEM_QUANTITY: quantity,
            C.F_THR3D_SHIPPING_ITEM_STATUS: "Needs Shipment",
            **({C.F_THR3D_SHIPPING_ITEM_EXPECTED_PRODUCT: expected_product_ids} if expected_product_ids else {}),
            **({C.F_THR3D_SHIPPING_ITEM_MANUAL_PRODUCT_INFO: manual_product_info} if manual_product_info else {}),
        }

    if thr3d_fields and any(fields[C.F_WORKSTREAM_CARD_TYPE] == "Ecomm" for fields in workstreams):
        return err("Ecomm and THR3D are alternate GS1 paths. Choose one of them, not both.", 400)
    if thr3d_fields:
        packaging_quantity = sum(
            fields[C.F_WORKSTREAM_CARD_QUANTITY]
            for fields in workstreams
            if fields[C.F_WORKSTREAM_CARD_TYPE] == "Packaging"
        )
        if packaging_quantity:
            try:
                parent_quantity = int(entry_fields.get(C.F_RECEIPT_ENTRY_QUANTITY) or 0)
            except (TypeError, ValueError):
                parent_quantity = 0
            total_assigned = packaging_quantity + thr3d_fields[C.F_THR3D_SHIPPING_ITEM_QUANTITY]
            if parent_quantity and total_assigned != parent_quantity:
                return err("Packaging and THR3D quantities must add up to the Received Merch quantity.", 400)
    if not workstreams and not thr3d_fields:
        return err("Choose at least one Workstream or THR3D Shipping Item.", 400)
    return workstreams, thr3d_fields, manual_product_info, expected_product_ids


def _required_to_shoot_requirement(key, label, ready, missing):
    return {
        "key": key,
        "label": label,
        "ready": bool(ready),
        "satisfied": bool(ready),
        "missing": "" if ready else missing,
        "detail": "Complete" if ready else missing,
        "tone": "green" if ready else "red",
    }


def _reference_data_value(product_fields, terms):
    reference_data = _parse_reference_data(product_fields.get(C.F_ITEM_REFERENCE_DATA, ""))
    candidates = []
    for key, value in reference_data.items():
        key_text = str(key or "").lower()
        if any(term in key_text for term in terms):
            candidates.append(value)
    return next((str(value or "").strip() for value in candidates if str(value or "").strip()), "")


def _client_config_for_entry(entry_fields, receipt=None):
    client_ids = _as_list((receipt or {}).get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
    if not client_ids:
        client_ids = _as_list(entry_fields.get(C.F_RECEIPT_CLIENT, []))
    return _client_config(client_ids[0]) if client_ids else {}


def _client_product_requirements(client_config, deliverables):
    """Product fields the client requires for the chosen photo deliverables.

    The client settings page already owns this list per workstream; the readiness
    gate reads it rather than keeping a second, hard-coded copy that can disagree
    with what the modal shows.
    """
    requirements = (client_config or {}).get("photoProductionRequirements") or {}
    workstreams = requirements.get("workstreams") or {}
    keys = []
    for deliverable in deliverables:
        for key in (workstreams.get(deliverable) or {}).get("requiredProductFields") or []:
            if key not in keys:
                keys.append(key)
    return keys


def _evaluate_required_to_shoot_from_fields(entry_fields, product_fields=None, client_config=None):
    product_fields = product_fields or {}
    item_ids = _as_list(entry_fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
    merchandise_verified = bool(entry_fields.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED, False))
    product_name = str(product_fields.get(C.F_ITEM_NAME, "") or "").strip()
    identifier = str(product_fields.get(C.F_ITEM_IDENTIFIER, "") or "").strip()
    identifier_label = (client_config or {}).get("identifierLabel") or "UPC / Product ID"
    deliverables = _deliverable_values(entry_fields.get(C.F_RECEIPT_ENTRY_DELIVERABLES, ""))
    requirements = [
        _required_to_shoot_requirement("merchandise-verified", "Merchandise Verified", merchandise_verified, "Verify the physical merchandise."),
        _required_to_shoot_requirement("deliverables", "Deliverables", bool(deliverables), "Select at least one Deliverable."),
    ]
    photo_deliverables = [value for value in deliverables if value in {"Packaging", "Ecomm"}]
    if deliverables == ["Thr3d"]:
        requirements = [
            _required_to_shoot_requirement("client", "Client", bool(_as_list(entry_fields.get(C.F_RECEIPT_CLIENT, []))), "Select a Client."),
            _required_to_shoot_requirement("merchandise-photo", "Merchandise Photo", _has_merchandise_photo(entry_fields), "Add at least one merchandise photo."),
            _required_to_shoot_requirement("quantity", "Quantity", _quantity_is_present(entry_fields.get(C.F_RECEIPT_ENTRY_QUANTITY)), "Add Quantity."),
            _required_to_shoot_requirement("deliverables", "Deliverables", True, "Select Thr3d."),
        ]
        complete_count = len([item for item in requirements if item["ready"]])
        missing = [item["label"] for item in requirements if not item["ready"]]
        return {
            "ready": complete_count == len(requirements),
            "complete": complete_count == len(requirements),
            "completeCount": complete_count,
            "totalCount": len(requirements),
            "summary": f"{complete_count} of {len(requirements)} Complete",
            "missing": missing,
            "requirements": requirements,
        }
    if photo_deliverables:
        requirements.extend([
            _required_to_shoot_requirement("product-linked", "Product Linked", bool(item_ids and product_fields), "Link a Product."),
            _required_to_shoot_requirement("product-name", "Product Name", bool(product_name), "Add Product Name."),
            _required_to_shoot_requirement("identifier", identifier_label, bool(identifier), f"Add the Product {identifier_label}."),
        ])
    # Which Product fields are required is the client's configuration and nothing
    # else. A client that has asked for no fields blocks on no fields.
    product_view = _product_view_for_requirements(product_fields)
    for key in _client_product_requirements(client_config, photo_deliverables):
        if key in {"productName", "upc"}:
            continue  # identity, already covered above and supplied by matching
        label = PHOTO_PRODUCTION_REQUIREMENT_FIELDS.get(key, key)
        present = _photo_production_value_is_valid(key, _photo_product_value(product_view, key))
        requirements.append(_required_to_shoot_requirement(key, label, present, f"Add {label}."))
    complete_count = len([item for item in requirements if item["ready"]])
    missing = [item["label"] for item in requirements if not item["ready"]]
    return {
        "ready": complete_count == len(requirements),
        "complete": complete_count == len(requirements),
        "completeCount": complete_count,
        "totalCount": len(requirements),
        "summary": f"{complete_count} of {len(requirements)} Complete",
        "missing": missing,
        "requirements": requirements,
    }


def _evaluate_required_to_shoot(shaped, linked_item=None):
    entry_fields = {
        C.F_RECEIPT_ENTRY_ITEM: shaped.get("itemIds", []),
        C.F_RECEIPT_ENTRY_NAME: shaped.get("productName", ""),
        C.F_RECEIPT_ENTRY_SKU_ID: shaped.get("skuId", ""),
        C.F_RECEIPT_ENTRY_DELIVERABLES: shaped.get("deliverables", []),
        C.F_RECEIPT_ENTRY_MERCH_VERIFIED: shaped.get("merchandiseVerified", False),
        C.F_RECEIPT_ENTRY_QUANTITY: shaped.get("quantity", 0),
        C.F_RECEIPT_CLIENT: shaped.get("clientIds", []),
        C.F_RECEIPT_ENTRY_PHOTO_METADATA: shaped.get("photoMetadata", []),
        C.F_RECEIPT_ENTRY_PHOTOS: shaped.get("itemPhotos", []) or shaped.get("photos", []),
    }
    product_fields = {}
    if linked_item:
        product_fields = {
            C.F_ITEM_NAME: linked_item.get("name") or "",
            C.F_ITEM_IDENTIFIER: linked_item.get("identifier") or linked_item.get("productId") or "",
            C.F_ITEM_ARTWORK_RECEIVED: linked_item.get("artworkReceived", False),
            C.F_ITEM_REFERENCE_DATA: json.dumps(linked_item.get("referenceData") or {}, sort_keys=True),
        }
    return _evaluate_required_to_shoot_from_fields(entry_fields, product_fields)


def _quantity_is_present(value):
    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def _has_merchandise_photo(fields):
    if _photo_metadata_from_entry(fields, include_urls=False):
        return True
    return bool(_as_list(fields.get(C.F_RECEIPT_ENTRY_PHOTOS, [])))


def _intake_decision_fields_from_body(body, existing_fields=None):
    deliverables_supplied = "deliverables" in body
    deliverables_raw = body.get("deliverables")
    fields = {}

    if any(key in body for key in ("productName", "product_name", "packageName", "package_name", "name")):
        product_name = (
            body.get("productName")
            or body.get("product_name")
            or body.get("packageName")
            or body.get("package_name")
            or body.get("name")
            or ""
        ).strip()
        fields[C.F_RECEIPT_ENTRY_NAME] = product_name or "Unnamed Product"
    if any(key in body for key in ("skuId", "sku_id", "observedIdentifier", "identifier")):
        sku_id = (
            body.get("skuId")
            or body.get("sku_id")
            or body.get("observedIdentifier")
            or body.get("identifier")
            or ""
        ).strip()
        fields[C.F_RECEIPT_ENTRY_SKU_ID] = sku_id

    if deliverables_supplied:
        deliverables = _validate_deliverables(deliverables_raw)
        if isinstance(deliverables, tuple):
            return deliverables
        fields[C.F_RECEIPT_ENTRY_DELIVERABLES] = deliverables
    manual_product_info = _manual_product_info_from_body(body)
    if any(key in body for key in ("manualProductInfo", "manual_product_info", "productInfo", "product_info")):
        fields[C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO] = manual_product_info

    return fields

def _parse_airtable_datetime(value):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = f"{text[:-1]}+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _days_here_from_received(received, now=None):
    parsed = _parse_airtable_datetime(received)
    if not parsed:
        return None
    now = now or _now_utc()
    if now.tzinfo is None:
        now = now.replace(tzinfo=timezone.utc)
    return max(0, (now.astimezone(timezone.utc).date() - parsed.date()).days)


def _time_here_label(days_here):
    if days_here is None:
        return "Unknown"
    if days_here == 0:
        return "Today"
    if days_here < 14 or days_here > 30:
        return f"{days_here} day{'s' if days_here != 1 else ''}"
    weeks = max(2, days_here // 7)
    return f"{weeks} weeks"


def _age_group_for_days(days_here):
    if days_here is None:
        return "unknown"
    if days_here <= 7:
        return "0-7"
    if days_here <= 14:
        return "8-14"
    if days_here <= 30:
        return "15-30"
    return "30-plus"


def _is_merchandise_physically_present(merch_status):
    merch = str(merch_status or "").strip().lower()
    return merch not in NON_INVENTORY_MERCH_STATUSES


def _derive_merchandise_inventory_status(entry, linked_product=None, client=None, days_here=None):
    merch_status = _normalized_merch_status(entry.get("merchStatus"))
    if merch_status == "Issue":
        return "Issue"

    dispo_days = (client or {}).get("dispoDays")
    try:
        dispo_days = int(dispo_days)
    except (TypeError, ValueError):
        dispo_days = None
    if dispo_days and days_here is not None and days_here >= dispo_days:
        return "Disposition Due"

    return merch_status or "Received"


def _shape_merchandise_inventory_entry(entry, *, receipts_by_id, products_by_id, clients_by_id, locations_by_id, now=None):
    shaped = _shape_receipt_entry(entry)
    receipt = next(
        (receipts_by_id.get(receipt_id) for receipt_id in _as_list(shaped.get("receiptIds")) if receipts_by_id.get(receipt_id)),
        None,
    )
    shaped = _with_shipment_photos(shaped, receipt)
    receipt_fields = receipt.get("fields", {}) if receipt else {}
    product_record = next(
        (products_by_id.get(product_id) for product_id in _as_list(shaped.get("itemIds")) if products_by_id.get(product_id)),
        None,
    )
    linked_product = _shape_item(product_record, clients_by_id=clients_by_id, issues_by_item_id={}) if product_record else None
    client_ids = _as_list(receipt_fields.get(C.F_RECEIPT_CLIENT, [])) or _as_list((linked_product or {}).get("clientIds", []))
    client = clients_by_id.get(client_ids[0]) if client_ids else None
    location_id = (_as_list(shaped.get("locationIds")) or _as_list(receipt_fields.get(C.F_RECEIPT_LOCATION, [])) or [""])[0]
    location = locations_by_id.get(location_id, {})
    received = receipt_fields.get(C.F_RECEIPT_RECEIVED, "")
    days_here = _days_here_from_received(received, now=now)
    status = _derive_merchandise_inventory_status(shaped, linked_product, client, days_here)
    return {
        **shaped,
        "packageName": shaped.get("productName", ""),
        "barcodeOrIdNumber": shaped.get("skuId", ""),
        "client": client.get("name", "") if client else "",
        "clientIds": client_ids,
        "matchedProduct": {
            "id": linked_product.get("id"),
            "name": linked_product.get("name") or linked_product.get("product") or "",
            "identifier": linked_product.get("identifier", ""),
        } if linked_product else None,
        "shipment": {
            "id": receipt.get("id") if receipt else "",
            "name": receipt_fields.get(C.F_RECEIPT_NAME, "") if receipt else "",
            "carrier": receipt_fields.get(C.F_RECEIPT_CARRIER, "") if receipt else "",
            "tracking": receipt_fields.get(C.F_RECEIPT_TRACKING, "") if receipt else "",
            "received": received,
        },
        "storageLocation": location.get("name", "") if location else "",
        "locationId": location_id,
        "dateReceived": received,
        "received": received,
        "daysHere": days_here,
        "timeHere": _time_here_label(days_here),
        "ageGroup": _age_group_for_days(days_here),
        "status": status,
        "inventoryStatus": status,
    }


def _list_merchandise_inventory_records():
    entries = _list_all_records(C.MERCHANDISE_TABLE)
    receipts = _filter_receipts_by_access(_list_all_records(C.SHIPMENTS_TABLE))
    products = _filter_by_client_field(_list_all_records(C.PRODUCTS_TABLE), C.F_ITEM_CLIENT)
    clients = _permitted_client_records(_list_all_records(C.CLIENTS_TABLE))
    locations = _filter_locations(_list_all_records(C.LOCATIONS_TABLE))

    receipts_by_id = {record["id"]: record for record in receipts}
    products_by_id = {record["id"]: record for record in products}
    clients_by_id = {record["id"]: _shape_client(record) for record in clients}
    locations_by_id = {record["id"]: _shape_location(record) for record in locations}
    now = _now_utc()
    records = []
    for entry in entries:
        shaped = _shape_receipt_entry(entry)
        linked_receipt_ids = _as_list(shaped.get("receiptIds"))
        if linked_receipt_ids and not any(receipt_id in receipts_by_id for receipt_id in linked_receipt_ids):
            continue
        linked_product = products_by_id.get((_as_list(shaped.get("itemIds")) or [""])[0])
        if shaped.get("itemIds") and not linked_product:
            continue
        if not _is_merchandise_physically_present(shaped.get("merchStatus")):
            continue
        records.append(_shape_merchandise_inventory_entry(
            entry,
            receipts_by_id=receipts_by_id,
            products_by_id=products_by_id,
            clients_by_id=clients_by_id,
            locations_by_id=locations_by_id,
            now=now,
        ))
    records.sort(key=lambda record: (
        record.get("daysHere") is None,
        -(record.get("daysHere") or -1),
        record.get("packageName") or "",
    ))
    return records


@api.get("/merchandise")
def list_merchandise_inventory():
    try:
        records = _list_merchandise_inventory_records()
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": records})


@api.get("/verification/entries")
@api.get("/merchandise/review")
def list_verification_entries():
    try:
        records = _list_merchandise_review_records()
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": records})


def _permitted_merchandise_or_error(entry_id):
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return None, None, airtable_err(error)
    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return None, None, _forbidden()
    return entry, receipt, None


def _comment_user_ids(records):
    user_ids = set()
    for record in records:
        for user_id in _as_list(record.get("fields", {}).get(C.F_COMMENT_USER, [])):
            if user_id:
                user_ids.add(user_id)
    return user_ids


def _users_by_id(user_ids):
    users = {}
    for user_id in user_ids:
        try:
            users[user_id] = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
        except requests.HTTPError:
            continue
    return users


def _shape_comment(record, users_by_id=None):
    fields = record.get("fields", {})
    user_id = (_as_list(fields.get(C.F_COMMENT_USER, [])) or [""])[0]
    user_record = (users_by_id or {}).get(user_id)
    author = _shape_user(user_record) if user_record else {
        "id": user_id,
        "name": "",
        "displayName": "",
        "role": "",
        "avatar": "",
    }
    display_name = _user_display_name(author)
    created_at = fields.get(C.F_COMMENT_CREATED_AT) or record.get("createdTime") or ""
    return {
        "id": record.get("id", ""),
        "body": fields.get(C.F_COMMENT_BODY, ""),
        "merchandiseId": (_as_list(fields.get(C.F_COMMENT_MERCHANDISE, [])) or [""])[0],
        "createdAt": created_at,
        "author": {
            "id": user_id,
            "name": display_name,
            "displayName": display_name,
            "role": author.get("role", ""),
            "avatar": author.get("avatar", ""),
            "initials": author.get("avatar") or _user_initials(author),
        },
    }


def _comment_sort_key(record):
    fields = record.get("fields", {})
    return fields.get(C.F_COMMENT_CREATED_AT) or record.get("createdTime") or ""


def _list_comments_for_merchandise(entry_id):
    records = _list_all_records(C.COMMENTS_TABLE)
    return [
        record for record in sorted(records, key=_comment_sort_key)
        if entry_id in _as_list(record.get("fields", {}).get(C.F_COMMENT_MERCHANDISE, []))
    ]


def _comment_reads_for_user(user_id):
    """Read the signed-in user's comment read-through map.

    Stored as app-owned JSON on the Users record so unread state follows the person
    across devices. Malformed JSON degrades to an empty map rather than failing the
    request: losing read state shows extra unread badges, which is recoverable, while
    a hard error would block the Planning board from loading at all.
    """
    if not user_id:
        return {}
    try:
        record = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
    except requests.HTTPError:
        return {}
    raw = (record.get("fields", {}) or {}).get(C.F_USER_COMMENT_READS) or ""
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError):
        current_app.logger.warning("Malformed Comment Reads JSON for user %s", user_id)
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): str(value) for key, value in parsed.items() if key and value}


@api.get("/comment-reads")
def list_comment_reads():
    user_id = _current_user_id()
    if not user_id:
        return err("Sign in to read comment state.", 401)
    return jsonify({"reads": _comment_reads_for_user(user_id)})


@api.post("/comment-reads/<entry_id>")
def mark_comment_read(entry_id):
    user_id = _current_user_id()
    if not user_id:
        return err("Sign in to update comment state.", 401)
    _, _, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    reads = _comment_reads_for_user(user_id)
    reads[entry_id] = datetime.now(timezone.utc).isoformat()
    try:
        airtable.update_record(
            C.USERS_TABLE,
            user_id,
            {C.F_USER_COMMENT_READS: json.dumps(reads, sort_keys=True)},
            by_field_id=False,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"reads": reads})


@api.get("/merchandise/<entry_id>/history")
def list_merchandise_history(entry_id):
    """Lifecycle events for one merchandise record, newest first."""
    entry, receipt, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    history_ids = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_HISTORY, []))
    records = []
    for history_id in history_ids:
        try:
            records.append(airtable.get_record(C.HISTORY_TABLE, history_id, by_field_id=False))
        except requests.HTTPError:
            continue
    user_ids = {
        user_id
        for record in records
        for user_id in _as_list(record.get("fields", {}).get(C.F_HISTORY_USER, []))
        if user_id
    }
    users_by_id = _users_by_id(user_ids)
    events = []
    for record in records:
        fields = record.get("fields", {})
        user_id = (_as_list(fields.get(C.F_HISTORY_USER, [])) or [""])[0]
        user_record = users_by_id.get(user_id)
        events.append({
            "id": record.get("id", ""),
            "action": fields.get(C.F_HISTORY_EVENT, ""),
            "actor": _user_display_name(user_record) if user_record else "System",
            "createdAt": fields.get(C.F_HISTORY_DATE) or record.get("createdTime") or "",
            "from": fields.get(C.F_HISTORY_FROM, ""),
            "to": fields.get(C.F_HISTORY_TO, ""),
        })
    events.sort(key=lambda event: event.get("createdAt") or "", reverse=True)

    # Every item's history begins with its arrival. Records created before merchandise
    # events were recorded have no such row, and the moment is already known from the
    # shipment, so it is derived rather than backfilled as an invented audit line.
    if not any(event["action"] == MERCHANDISE_CREATED_EVENT for event in events):
        receipt_fields = (receipt or {}).get("fields", {})
        created_at = receipt_fields.get(C.F_RECEIPT_RECEIVED) or entry.get("createdTime") or ""
        if created_at:
            events.append({
                "id": f"{entry_id}:created",
                "action": MERCHANDISE_CREATED_EVENT,
                "actor": "",
                "createdAt": created_at,
                "from": "",
                "to": "",
            })
    return jsonify({"records": events})


@api.get("/merchandise/<entry_id>/comments")
def list_merchandise_comments(entry_id):
    _, _, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    try:
        records = _list_comments_for_merchandise(entry_id)
        users = _users_by_id(_comment_user_ids(records))
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": [_shape_comment(record, users) for record in records]})


@api.post("/merchandise/<entry_id>/comments")
def create_merchandise_comment(entry_id):
    body = request.get_json(silent=True) or {}
    comment = (body.get("comment") or body.get("body") or "").strip()
    if not comment:
        return err("Comment is required.")
    user_id = _current_user_id()
    if not user_id:
        return err("Authenticated user is required to comment.", 401)
    _, _, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    try:
        user_record = airtable.get_record(C.USERS_TABLE, user_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _user_display_name(user_record):
        return err("Authenticated user must have a name before commenting.", 400)
    try:
        record = airtable.create_record(
            C.COMMENTS_TABLE,
            {
                C.F_COMMENT_BODY: comment,
                C.F_COMMENT_MERCHANDISE: [entry_id],
                C.F_COMMENT_USER: [user_id],
            },
            by_field_id=False,
            typecast=True,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"comment": _shape_comment(record, {user_id: user_record})}), 201


@api.get("/verification/items")
@api.get("/merchandise/products")
def verification_items():
    query = request.args.get("q", "")
    client_id = (request.args.get("clientId") or "").strip()
    include_item_id = (request.args.get("includeItemId") or "").strip()
    limit = request.args.get("limit", "8")
    if client_id and not _client_permitted(client_id):
        return _forbidden()
    try:
        matches = _find_matching_skus(query, client_id=client_id, include_item_id=include_item_id, limit=limit)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": matches})


@api.get("/workstream-cards")
def list_workstream_cards():
    try:
        records = _workstream_cards_for_planning()
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": records})


@api.post("/workstream-cards")
def create_workstream_card():
    body = request.get_json(silent=True) or {}
    merchandise_id = str(body.get("merchandiseId") or body.get("receivedMerchandiseId") or "").strip()
    workstream_type = str(body.get("workstreamType") or body.get("type") or "").strip()
    if not merchandise_id:
        return err("Received Merch is required.", 400)
    if workstream_type not in C.WORKSTREAM_TYPE_OPTIONS:
        return err(f"Workstream Type must be one of: {', '.join(C.WORKSTREAM_TYPE_OPTIONS)}.", 400)
    entry, _receipt, access_error = _permitted_merchandise_or_error(merchandise_id)
    if access_error:
        return access_error
    try:
        siblings = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
    except requests.HTTPError as error:
        return airtable_err(error)
    existing_types = []
    sibling_ready_for_photo = False
    for sibling in siblings:
        sibling_fields = sibling.get("fields", {})
        if merchandise_id not in _as_list(sibling_fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])):
            continue
        sibling_type = str(sibling_fields.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip()
        if sibling_type and sibling_type not in existing_types:
            existing_types.append(sibling_type)
        if str(sibling_fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS) or "").strip().lower() == "awaiting photo release":
            sibling_ready_for_photo = True
    if workstream_type in existing_types:
        return err(f"{workstream_type} already exists for this merchandise.", 409)

    entry_fields = entry.get("fields", {})
    expected_product_ids = _as_list(entry_fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
    manual_product_info = entry_fields.get(C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO, "")
    quantity = _positive_int(entry_fields.get(C.F_RECEIPT_ENTRY_QUANTITY, 1), "Quantity")
    if isinstance(quantity, tuple):
        return quantity
    fields = {
        C.F_WORKSTREAM_CARD_NAME: _workstream_name(entry_fields, workstream_type),
        C.F_WORKSTREAM_CARD_RECEIVED_MERCH: [merchandise_id],
        C.F_WORKSTREAM_CARD_TYPE: workstream_type,
        C.F_WORKSTREAM_CARD_PLANNING_STATUS: PLANNING_STATUS_LABELS["needs-more-information"],
        C.F_WORKSTREAM_CARD_QUANTITY: quantity,
        **({C.F_WORKSTREAM_CARD_EXPECTED_PRODUCT: expected_product_ids} if expected_product_ids else {}),
        **({C.F_WORKSTREAM_CARD_MANUAL_PRODUCT_INFO: manual_product_info} if manual_product_info else {}),
    }
    try:
        created = airtable.create_record(C.WORKSTREAM_CARDS_TABLE, fields, by_field_id=False, typecast=True)
    except requests.HTTPError as error:
        return airtable_err(error)
    remaining_types = existing_types + [workstream_type]
    parent_is_ready_for_photo = str(entry_fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) or "").strip().lower() == "awaiting photo release"
    parent_update = {
        C.F_RECEIPT_ENTRY_DELIVERABLES: remaining_types,
    }
    if not (sibling_ready_for_photo or parent_is_ready_for_photo):
        parent_update.update({
            C.F_RECEIPT_ENTRY_PLANNING_STATUS: PLANNING_STATUS_LABELS["needs-more-information"],
            **_merch_status_normalization_fields(entry_fields),
        })
    try:
        _update_receipt_entry_record(merchandise_id, parent_update)
    except requests.HTTPError:
        return jsonify({"record": _shape_workstream_card(created), "warning": "Workstream added, but the parent merchandise deliverables could not be updated."}), 201
    return jsonify({"record": _shape_workstream_card(created), "remainingDeliverables": remaining_types}), 201


@api.patch("/workstream-cards/<record_id>")
def update_workstream_card(record_id):
    body = request.get_json(silent=True) or {}
    planning_status = _normalized_planning_status(body.get("planningStatus") or "")
    # A card is created only after merchandise is accepted and deliverables are
    # known, so it is born at Needs More Information. New belongs to the parent
    # merchandise and is not a reachable state for child work.
    if planning_status and planning_status not in WORKSTREAM_CARD_PLANNING_STATUS_VALUES:
        return err(
            f"planningStatus must be one of: {', '.join(WORKSTREAM_CARD_PLANNING_STATUS_VALUES)}.",
            400,
        )
    if "status" in body:
        return err("Status is no longer used for Workstream Cards; use planningStatus.", 400)
    try:
        current = airtable.get_record(C.WORKSTREAM_CARDS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    merchandise_id = (_as_list(current.get("fields", {}).get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])) or [""])[0]
    _entry = None
    if merchandise_id:
        _entry, _receipt, access_error = _permitted_merchandise_or_error(merchandise_id)
        if access_error:
            return access_error
    current_fields = current.get("fields", {})
    current_planning_status = _normalized_planning_status(
        current_fields.get(C.F_WORKSTREAM_CARD_PLANNING_STATUS, "")
    )
    effective_planning_status = planning_status or current_planning_status or "needs-more-information"
    workstream_type = body.get("workstreamType", body.get("type", current_fields.get(C.F_WORKSTREAM_CARD_TYPE, "")))
    workstream_type = str(workstream_type or "").strip()
    if workstream_type not in C.WORKSTREAM_TYPE_OPTIONS:
        return err(f"Workstream Type must be one of: {', '.join(C.WORKSTREAM_TYPE_OPTIONS)}.", 400)
    update_fields = {
        C.F_WORKSTREAM_CARD_TYPE: workstream_type,
        C.F_WORKSTREAM_CARD_PLANNING_STATUS: PLANNING_STATUS_LABELS[effective_planning_status],
    }
    if any(key in body for key in ("manualProductInfo", "manual_product_info", "productInfo", "product_info")):
        update_fields[C.F_WORKSTREAM_CARD_MANUAL_PRODUCT_INFO] = _manual_product_info_from_body(body)
    if workstream_type != current_fields.get(C.F_WORKSTREAM_CARD_TYPE, "") and _entry:
        update_fields[C.F_WORKSTREAM_CARD_NAME] = _workstream_name(_entry.get("fields", {}), workstream_type)
    try:
        updated = airtable.update_record(
            C.WORKSTREAM_CARDS_TABLE,
            record_id,
            update_fields,
            by_field_id=False,
            typecast=True,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"record": _shape_workstream_card(updated)})


@api.delete("/workstream-cards/<record_id>")
def delete_workstream_card(record_id):
    try:
        current = airtable.get_record(C.WORKSTREAM_CARDS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    fields = current.get("fields", {})
    creative_force = _parse_creative_force_sync(fields.get(C.F_WORKSTREAM_CARD_CREATIVE_FORCE_SYNC, ""))
    if creative_force.get("status") in {"Scheduled", "In Production", "Complete"}:
        return err("This workstream has already been handed to Creative Force and cannot be removed.", 409)
    merchandise_id = (_as_list(fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])) or [""])[0]
    if merchandise_id:
        entry, _receipt, access_error = _permitted_merchandise_or_error(merchandise_id)
        if access_error:
            return access_error
    else:
        entry = None
    remaining_types = []
    if merchandise_id:
        try:
            siblings = _list_all_records(C.WORKSTREAM_CARDS_TABLE)
        except requests.HTTPError as error:
            return airtable_err(error)
        for sibling in siblings:
            if sibling.get("id") == record_id:
                continue
            sibling_fields = sibling.get("fields", {})
            if merchandise_id in _as_list(sibling_fields.get(C.F_WORKSTREAM_CARD_RECEIVED_MERCH, [])):
                sibling_type = str(sibling_fields.get(C.F_WORKSTREAM_CARD_TYPE) or "").strip()
                if sibling_type and sibling_type not in remaining_types:
                    remaining_types.append(sibling_type)
    try:
        airtable.delete_record(C.WORKSTREAM_CARDS_TABLE, record_id)
    except requests.HTTPError as error:
        return airtable_err(error)

    parent_warning = ""
    if entry is not None:
        update_fields = {
            C.F_RECEIPT_ENTRY_DELIVERABLES: remaining_types,
            **_merch_status_normalization_fields(entry.get("fields", {})),
        }
        if not remaining_types:
            # Only removing the last card returns the parent to the board. While
            # sibling cards remain the parent stays off-board, so rewriting its
            # Planning Status here would regress accepted merchandise back to New.
            update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["needs-more-information"]
            update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED] = False
        try:
            updated_entry = _update_receipt_entry_record(merchandise_id, update_fields)
        except requests.HTTPError:
            parent_warning = "The photo card was removed, but the parent merchandise queue could not be updated. Refresh and review it."
            updated_entry = None
    response = {"deleted": record_id, "remainingDeliverables": remaining_types, "parentMerchandiseId": merchandise_id}
    if updated_entry is not None:
        response["parentMerchandise"] = _shape_verification_entry(updated_entry, _receipt)
    if parent_warning:
        response["warning"] = parent_warning
    return jsonify(response)


@api.get("/thr3d-shipping-items")
def list_thr3d_shipping_items():
    try:
        all_records = _list_all_records(C.THR3D_SHIPPING_ITEMS_TABLE)
        parent_cache = {}
        records = _thr3d_shipping_items_for_shipments(all_records, parent_cache=parent_cache)
        shipped = _thr3d_shipping_items_for_shipments(all_records, include_shipped=True, parent_cache=parent_cache)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"records": records, "shipped": shipped})


@api.post("/thr3d-shipping-items/<record_id>/ship")
def ship_thr3d_shipping_item(record_id):
    body = request.get_json(silent=True) or {}
    return _mark_thr3d_shipping_item_shipped(record_id, body)


@api.post("/merchandise/<entry_id>/confirm-assign")
@api.post("/merchandise/review/<entry_id>/confirm-assign")
def confirm_assign_merchandise(entry_id):
    body = request.get_json(silent=True) or {}
    entry, receipt, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error

    expected_product_id = str(body.get("expectedProductId") or body.get("productId") or body.get("itemId") or "").strip()
    item_record = None
    if expected_product_id:
        try:
            item_record = airtable.get_record(C.PRODUCTS_TABLE, expected_product_id, by_field_id=False)
        except requests.HTTPError as error:
            return airtable_err(error)
        item_client_ids = _as_list(item_record.get("fields", {}).get(C.F_ITEM_CLIENT, []))
        if not _client_ids_permitted(item_client_ids):
            return _forbidden()
        receipt_client_ids = _as_list(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])) if receipt else []
        if receipt_client_ids and item_client_ids and not (set(receipt_client_ids) & set(item_client_ids)):
            return err("Product does not belong to this Shipment client.", 403)

    requested_status = _normalized_planning_status(body.get("planningStatus", ""))
    card_planning_status = None
    if requested_status == "awaiting-photo-release":
        deliverable_names = [
            str((raw or {}).get("type") or (raw or {}).get("workstreamType") or "").strip()
            for raw in (body.get("workstreams") or body.get("workstreamCards") or [])
        ]
        effective_fields = {
            **entry.get("fields", {}),
            C.F_RECEIPT_ENTRY_DELIVERABLES: [name for name in deliverable_names if name],
        }
        if expected_product_id:
            effective_fields[C.F_RECEIPT_ENTRY_ITEM] = [expected_product_id]
        readiness = _evaluate_required_to_shoot_from_fields(
            effective_fields,
            item_record.get("fields", {}) if item_record else {},
            client_config=_client_config_for_entry(entry.get("fields", {}), receipt),
        )
        if not readiness["ready"]:
            return err(f"Cannot move to Awaiting Photo Release.\nMissing: {', '.join(readiness['missing'])}", 400)
        if expected_product_id and _blocking_merchandise_issues(_issues_by_item_id().get(expected_product_id, [])):
            return err("Cannot move to Awaiting Photo Release.\nMissing: Resolved Merchandise Issues", 400)
        card_planning_status = PLANNING_STATUS_LABELS["awaiting-photo-release"]
    elif requested_status and requested_status not in PLANNING_STATUS_VALUES:
        return err("planningStatus must be one of: Needs More Information, Awaiting Photo Release.")

    parsed = _confirm_assign_payload(body, entry, item_record=item_record, card_planning_status=card_planning_status)
    if isinstance(parsed, tuple) and len(parsed) == 2:
        return parsed
    workstream_fields, thr3d_fields, manual_product_info, expected_product_ids = parsed

    try:
        workstream_cards = [
            airtable.create_record(C.WORKSTREAM_CARDS_TABLE, fields, by_field_id=False, typecast=True)
            for fields in workstream_fields
        ]
        thr3d_items = []
        if thr3d_fields:
            thr3d_items.append(airtable.create_record(C.THR3D_SHIPPING_ITEMS_TABLE, thr3d_fields, by_field_id=False, typecast=True))

        update_fields = {
            # The parent leaves the board because child work now exists, not because
            # of a status write. It is accepted merchandise, so it must not be sent
            # back to New: that is where it would land if the child work is removed.
            C.F_RECEIPT_ENTRY_PLANNING_STATUS: PLANNING_STATUS_LABELS["needs-more-information"],
            **_merch_status_normalization_fields(entry.get("fields", {})),
        }
        if expected_product_ids:
            update_fields[C.F_RECEIPT_ENTRY_ITEM] = expected_product_ids
        if manual_product_info:
            update_fields[C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO] = manual_product_info
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)

    return jsonify({
        "merchandise": _shape_verification_entry(updated, receipt, item_record=item_record),
        "workstreamCards": [_shape_workstream_card(record) for record in workstream_cards],
        "thr3dShippingItems": [_shape_thr3d_shipping_item(record) for record in thr3d_items],
    }), 201


@api.patch("/merchandise/<entry_id>/intake-decisions")
@api.patch("/merchandise/review/<entry_id>/intake-decisions")
def update_merchandise_intake_decisions(entry_id):
    body = request.get_json(silent=True) or {}
    deliverables_supplied = "deliverables" in body
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    fields_or_error = _intake_decision_fields_from_body(body, entry.get("fields", {}))
    if isinstance(fields_or_error, tuple):
        return fields_or_error
    if not fields_or_error:
        return jsonify(_shape_verification_entry(entry, receipt))

    if deliverables_supplied:
        current_app.logger.info(
            "Intake Deliverables save prepared: %s",
            json.dumps({
                "route": request.path,
                "method": request.method,
                "table": C.MERCHANDISE_TABLE,
                "record_id": entry_id,
                "field": C.F_RECEIPT_ENTRY_DELIVERABLES,
                "request_deliverables": body.get("deliverables"),
                "airtable_fields": fields_or_error,
                "user_id": _current_user_id(),
            }, default=str),
        )
    try:
        updated = _update_receipt_entry_record(entry_id, fields_or_error)
    except requests.HTTPError as error:
        if deliverables_supplied:
            response = getattr(error, "response", None)
            current_app.logger.error(
                "Intake Deliverables save failed: %s",
                json.dumps({
                    "route": request.path,
                    "method": request.method,
                    "table": C.MERCHANDISE_TABLE,
                    "record_id": entry_id,
                    "field": C.F_RECEIPT_ENTRY_DELIVERABLES,
                    "airtable_fields": fields_or_error,
                    "status": getattr(response, "status_code", None),
                    "response": getattr(response, "text", ""),
                    "user_id": _current_user_id(),
                }, default=str),
            )
        return airtable_err(error)
    if deliverables_supplied:
        current_app.logger.info(
            "Intake Deliverables save succeeded: %s",
            json.dumps({
                "route": request.path,
                "method": request.method,
                "table": C.MERCHANDISE_TABLE,
                "record_id": entry_id,
                "field": C.F_RECEIPT_ENTRY_DELIVERABLES,
                "saved_deliverables": updated.get("fields", {}).get(C.F_RECEIPT_ENTRY_DELIVERABLES, []),
                "user_id": _current_user_id(),
            }, default=str),
        )
    return jsonify(_shape_verification_entry(updated, receipt))


@api.patch("/merchandise/<entry_id>/intake-state")
@api.patch("/merchandise/review/<entry_id>/intake-state")
def update_merchandise_intake_state(entry_id):
    body = request.get_json(silent=True) or {}
    stage = (body.get("stage") or body.get("currentStage") or body.get("currentGate") or "").strip()
    planning_status = str(body.get("planningStatus") or "").strip().lower()
    if planning_status:
        planning_stage_map = {
            "new": "new-review",
            "needs-more-information": "waiting-info",
            "awaiting-photo-release": "ready-production",
        }
        if planning_status not in planning_stage_map:
            return err(f"planningStatus must be one of: {', '.join(PLANNING_STATUS_VALUES)}.", 400)
        stage = planning_stage_map[planning_status]
    allowed_stages = {"new-review", "waiting-info", "send-thr3d", "waiting-activation", "ready-production"}
    explicit_status = _intake_decision_value(body, "planningStatusLabel")
    if explicit_status is None and stage not in allowed_stages:
        return err("stage must be one of: new-review, waiting-info, send-thr3d, waiting-activation, ready-production.")
    planning_label_value = _validate_planning_status_label(explicit_status)
    if isinstance(planning_label_value, tuple):
        return planning_label_value

    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    fields = entry.get("fields", {})
    linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    update_fields = {}
    if "deliverables" in body:
        decision_fields = _intake_decision_fields_from_body(body, fields)
        if isinstance(decision_fields, tuple):
            return decision_fields
        update_fields.update(decision_fields)
    effective_fields = {**fields, **update_fields}
    if planning_label_value:
        label_to_planning_slug = {
            "new": "new",
            "needs more information": "needs-more-information",
            "awaiting photo release": "awaiting-photo-release",
        }
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS[
            label_to_planning_slug.get(str(planning_label_value).strip().lower(), "needs-more-information")
        ]
        update_fields.update(_merch_status_normalization_fields(fields))
    elif stage == "new-review":
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["new"]
        update_fields.update(_merch_status_normalization_fields(fields))
    elif stage == "waiting-info":
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS[
            planning_status if planning_status == "needs-more-information" else "needs-more-information"
        ]
        update_fields.update(_merch_status_normalization_fields(fields))
        if not fields.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED, False):
            update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED] = True
            update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT] = _now_iso()
            verifier_id = _current_user_id()
            if verifier_id:
                update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY] = _current_user_display_name()
    elif stage == "send-thr3d":
        deliverables = _deliverable_values(effective_fields.get(C.F_RECEIPT_ENTRY_DELIVERABLES, []))
        if "Thr3d" not in deliverables:
            deliverables.append("Thr3d")
        if any(value in {"Packaging", "Ecomm"} for value in deliverables):
            return err("Mixed photo + Thr3d Merchandise must complete the photo path before Thr3d shipping.", 400)
        effective_fields_with_receipt = {
            **effective_fields,
            C.F_RECEIPT_ENTRY_DELIVERABLES: deliverables,
            C.F_RECEIPT_CLIENT: receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, []) if receipt else [],
        }
        requiredToShoot = _evaluate_required_to_shoot_from_fields(effective_fields_with_receipt, {})
        if not requiredToShoot["ready"]:
            return err(f"Cannot move to Thr3d Shipment.\nMissing: {', '.join(requiredToShoot['missing'])}", 400)
        update_fields.update(_intake_decision_fields_from_body({"deliverables": deliverables}, effective_fields))
        # THR3D work is never photographed, so it must not claim a photo-release
        # status. It leaves Planning because a THR3D shipping item exists, and the
        # physical hand-off is expressed by Merch Status.
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["needs-more-information"]
        update_fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = "Ready to Ship"
    elif stage == "waiting-activation":
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["needs-more-information"]
        update_fields.update(_merch_status_normalization_fields(fields))
    elif stage == "ready-production":
        item_ids = _as_list(fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
        item_record = None
        if item_ids:
            try:
                item_record = airtable.get_record(C.PRODUCTS_TABLE, item_ids[0], by_field_id=False)
            except requests.HTTPError as error:
                return airtable_err(error)
        requiredToShoot = _evaluate_required_to_shoot_from_fields(
            effective_fields,
            item_record.get("fields", {}) if item_record else {},
            client_config=_client_config_for_entry(fields, receipt),
        )
        if not requiredToShoot["ready"]:
            return err(f"Cannot move to Awaiting Photo Release.\nMissing: {', '.join(requiredToShoot['missing'])}", 400)
        issues = _issues_by_item_id().get(item_ids[0], [])
        if _blocking_merchandise_issues(issues):
            return err("Cannot move to Awaiting Photo Release.\nMissing: Resolved Merchandise Issues", 400)
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["awaiting-photo-release"]
        update_fields.update(_merch_status_normalization_fields(fields))

    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    next_status = update_fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS)
    previous_status = fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS)
    if next_status and next_status != previous_status:
        # Accepting merchandise is the transition off New and is the event people
        # most want attributed. Anything else reads as a plain queue move.
        accepted = str(previous_status or "").strip() in ("", "New")
        _record_merchandise_history(
            entry_id,
            "Merchandise accepted" if accepted else "Planning status changed",
            from_value=previous_status or "New",
            to_value=next_status,
        )
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/merchandise/<entry_id>/release")
@api.post("/merchandise/review/<entry_id>/release")
def release_merchandise_to_production(entry_id):
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    fields = entry.get("fields", {})
    linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    item_ids = _as_list(fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
    item_record = None
    if item_ids:
        try:
            item_record = airtable.get_record(C.PRODUCTS_TABLE, item_ids[0], by_field_id=False)
        except requests.HTTPError as error:
            return airtable_err(error)

    product_fields = item_record.get("fields", {}) if item_record else {}
    requiredToShoot = _evaluate_required_to_shoot_from_fields(
        fields, product_fields, client_config=_client_config_for_entry(fields, receipt)
    )
    if not requiredToShoot["ready"]:
        missing_text = ", ".join(requiredToShoot["missing"])
        return jsonify({
            "error": f"Cannot release to photo. Missing: {missing_text}",
            "missing": requiredToShoot["missing"],
            "requiredToShoot": requiredToShoot,
        }), 400

    # A release is of one workstream. Releasing Ecomm must not hand Packaging to
    # Creative Force, and must not mark it released on the board.
    body = request.get_json(silent=True) or {}
    workstream_type = str(body.get("workstreamType") or "").strip()
    try:
        cards = _workstream_cards_for_merchandise(entry_id)
    except requests.HTTPError as error:
        return airtable_err(error)
    if workstream_type:
        cards = [
            card for card in cards
            if str(card.get("fields", {}).get(C.F_WORKSTREAM_CARD_TYPE, "")).strip() == workstream_type
        ]
    # Releasing again is a no-op for the workstream asked for. Asked at this level
    # rather than the arrival's, or a released Ecomm would block Packaging forever.
    # An arrival with no cards at all still releases: there is nothing to have done.
    pending = [card for card in cards if not card.get("fields", {}).get(C.F_WORKSTREAM_CARD_RELEASED)]
    already_released = (cards and not pending) or (not cards and fields.get(C.F_RECEIPT_ENTRY_RELEASED))
    if already_released:
        return jsonify(_shape_verification_entry(entry, receipt, item_record=item_record))
    cards = pending
    try:
        synced = _populate_creative_force_feed_for_ready_cards(cards) if cards else []
    except requests.HTTPError as error:
        return airtable_err(error)

    released_at = _now_iso()
    user_id = _current_user_id()
    card_fields = {
        C.F_WORKSTREAM_CARD_RELEASED: True,
        C.F_WORKSTREAM_CARD_RELEASED_AT: released_at,
    }
    if user_id:
        card_fields[C.F_WORKSTREAM_CARD_RELEASED_BY] = [user_id]
    try:
        for card in cards:
            airtable.update_record(C.WORKSTREAM_CARDS_TABLE, card["id"], card_fields, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    update_fields = {
        C.F_RECEIPT_ENTRY_RELEASED_AT: released_at,
    }
    if user_id:
        update_fields[C.F_RECEIPT_ENTRY_RELEASED_BY] = [user_id]
    # The arrival counts as released once every workstream on it has been.
    remaining = [
        card for card in _workstream_cards_for_merchandise(entry_id)
        if not card.get("fields", {}).get(C.F_WORKSTREAM_CARD_RELEASED)
    ]
    if not remaining:
        update_fields[C.F_RECEIPT_ENTRY_RELEASED] = True

    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    _record_merchandise_history(entry_id, "Released to photo")
    payload = _shape_verification_entry(updated, receipt, item_record=item_record)
    payload["creativeForceFeed"] = [
        {"sourceKey": entry["sourceKey"], "action": entry["action"]} for entry in synced
    ]
    return jsonify(payload)


@api.post("/merchandise/<entry_id>/verify")
@api.post("/merchandise/review/<entry_id>/verify")
def verify_merchandise(entry_id):
    """Record an explicit human verification of the physical merchandise.

    Verification asserts the observed package name and identifier are correct;
    callers may pass corrected values, which are persisted alongside the stamp.
    """
    body = request.get_json(silent=True) or {}
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    fields = entry.get("fields", {})
    linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    update_fields = {}
    if "packageName" in body or "package_name" in body:
        update_fields[C.F_RECEIPT_ENTRY_NAME] = _intake_decision_value(body, "packageName", "package_name") or ""
    if "identifier" in body or "skuId" in body:
        update_fields[C.F_RECEIPT_ENTRY_SKU_ID] = _intake_decision_value(body, "identifier", "skuId") or ""

    effective_fields = {**fields, **update_fields}
    package_name = str(effective_fields.get(C.F_RECEIPT_ENTRY_NAME, "") or "").strip()
    package_identifier = str(effective_fields.get(C.F_RECEIPT_ENTRY_SKU_ID, "") or "").strip()
    if not package_name or not package_identifier:
        missing = [
            label for label, present in (
                ("Product Name on Package", package_name),
                ("UPC / Product ID", package_identifier),
            ) if not present
        ]
        return err(f"Add {', '.join(missing)} before verifying the merchandise.", 400)

    update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED] = True
    update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT] = _now_iso()
    user_id = _current_user_id()
    if user_id:
        update_fields[C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY] = _current_user_display_name()
    # Clearing any prior physical issue is intentional: re-confirming the item resolves it.
    if fields.get(C.F_RECEIPT_ENTRY_MERCH_STATUS) == "Issue":
        update_fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = "Received"

    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/merchandise/<entry_id>/unverify")
@api.post("/merchandise/review/<entry_id>/unverify")
def unverify_merchandise(entry_id):
    """Clear a prior physical-verification stamp so the item returns to Needs Verification."""
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    fields = entry.get("fields", {})
    linked_receipts = _as_list(fields.get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    update_fields = {
        C.F_RECEIPT_ENTRY_MERCH_VERIFIED: False,
        C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT: "",
        C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY: "",
    }
    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/verification/entries/<entry_id>/match")
@api.post("/merchandise/<entry_id>/match")
@api.post("/merchandise/review/<entry_id>/match")
def match_verification_entry(entry_id):
    body = request.get_json(silent=True) or {}
    item_id = (body.get("itemId") or "").strip()
    if not item_id:
        return err("itemId is required")

    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
        item = airtable.get_record(C.PRODUCTS_TABLE, item_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()
    item_client_ids = _as_list(item.get("fields", {}).get(C.F_ITEM_CLIENT, []))
    if not _client_ids_permitted(item_client_ids):
        return _forbidden()
    receipt_client_ids = _as_list(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])) if receipt else []
    if receipt_client_ids and item_client_ids and not (set(receipt_client_ids) & set(item_client_ids)):
        return err("Product does not belong to this Shipment client.", 403)
    entry_fields = entry.get("fields", {})

    try:
        updated = _update_receipt_entry_record(entry_id, {
            C.F_RECEIPT_ENTRY_ITEM: [item_id],
            C.F_RECEIPT_ENTRY_PLANNING_STATUS: entry_fields.get(C.F_RECEIPT_ENTRY_PLANNING_STATUS) or "New",
            **_merch_status_normalization_fields(entry_fields),
        })
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt, item_record=item))


# A scan links on its own only when it is unambiguous: a full-length UPC resolving to
# exactly one candidate. Anything shorter, or matching more than one row, is a
# judgement call and belongs to a PM in Planning rather than a receiver holding a box.
# Defined server-side so source-sheet and local-Product clients answer it identically.
AUTO_MATCH_MIN_UPC_DIGITS = 8


def _auto_match_source_candidates(upc_key):
    """Source rows whose UPC equals this one exactly.

    Exact means the whole normalized UPC is equal, never a prefix or substring. Prefix
    matching is precisely what makes the suggestion list ambiguous.
    """
    source_rows, _range = _fetch_topco_source_rows(C.TOPCO_SOURCE_MATCH_ROW_WINDOW)
    return [
        row for row in source_rows
        if _source_check_upc_key((row.get("sourceData") or {}).get("UPC")) == upc_key
    ]


def _auto_match_product_candidates(upc_key, client_id):
    products = []
    for record in _list_all_records(C.PRODUCTS_TABLE):
        fields = record.get("fields", {})
        if _source_check_upc_key(fields.get(C.F_ITEM_UPC)) != upc_key:
            continue
        product_clients = _as_list(fields.get(C.F_ITEM_CLIENT, []))
        if client_id and product_clients and client_id not in product_clients:
            continue
        products.append(record)
    return products


def _resolve_unambiguous_upc(upc_key, client_id):
    """Resolve a UPC to exactly one candidate, or nothing.

    Returns (kind, candidate, reason, count). The count lets callers say how many rows
    a scan hit, which is the difference between "unknown barcode" and "your sheet has
    duplicates". Shared by the resolve and link endpoints so a scan means the same
    thing whether the merchandise exists yet or not.
    """
    if len(upc_key) < AUTO_MATCH_MIN_UPC_DIGITS:
        return None, None, "too-short", 0
    source_client = _topco_client_record(client_id) if client_id else None
    if source_client:
        candidates = _auto_match_source_candidates(upc_key)
        kind = "source"
    else:
        candidates = _auto_match_product_candidates(upc_key, client_id)
        kind = "product"
    if not candidates:
        return kind, None, "no-match", 0
    if len(candidates) > 1:
        return kind, None, "ambiguous", len(candidates)
    return kind, candidates[0], "", 1


@api.get("/products/resolve-upc")
def resolve_upc():
    """Resolve a scanned UPC without linking anything.

    Receiving stages merchandise before it has a record, so it needs the answer before
    there is anything to link to. Read-only: it never activates a source row.
    """
    upc_key = _source_check_upc_key(request.args.get("upc") or request.args.get("identifier") or "")
    client_id = (request.args.get("clientId") or "").strip()
    try:
        kind, candidate, reason, candidate_count = _resolve_unambiguous_upc(upc_key, client_id)
    except requests.RequestException:
        current_app.logger.exception("Resolve UPC source sheet read failed")
        return jsonify({"resolved": False, "reason": "source-unavailable"})
    except requests.HTTPError as error:
        return airtable_err(error)
    if not candidate:
        payload = {"resolved": False, "reason": reason}
        if candidate_count > 1:
            payload["candidateCount"] = candidate_count
        return jsonify(payload)
    if kind == "source":
        # The source row is returned unactivated. Receiving stages it and activation
        # happens on save, matching how a manually chosen row already behaves.
        return jsonify({"resolved": True, "via": "source", "sourceRow": candidate})
    return jsonify({
        "resolved": True,
        "via": "product",
        "product": _shape_item(candidate, clients_by_id=_clients_by_id()),
    })


@api.post("/merchandise/<entry_id>/auto-match")
def auto_match_merchandise(entry_id):
    """Link Merchandise only when a scanned UPC resolves to exactly one candidate.

    "No match" and "more than one" are ordinary outcomes, not errors: they leave the
    merchandise unmatched so Planning can resolve it.
    """
    body = request.get_json(silent=True) or {}
    upc_key = _source_check_upc_key(body.get("upc") or body.get("identifier") or "")
    if len(upc_key) < AUTO_MATCH_MIN_UPC_DIGITS:
        return jsonify({"matched": False, "reason": "too-short"})

    entry, receipt, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    client_ids = _as_list((receipt or {}).get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
    client_id = client_ids[0] if client_ids else ""

    try:
        kind, candidate, reason, candidate_count = _resolve_unambiguous_upc(upc_key, client_id)
    except requests.RequestException:
        current_app.logger.exception("Auto-match source sheet read failed")
        return jsonify({"matched": False, "reason": "source-unavailable"})
    except requests.HTTPError as error:
        return airtable_err(error)

    if not candidate:
        payload = {"matched": False, "reason": reason}
        if candidate_count > 1:
            payload["candidateCount"] = candidate_count
        return jsonify(payload)

    try:
        if kind == "source":
            source_row_number = candidate.get("sourceRowNumber")
            product_record, _result, _row = _activate_topco_source_product(
                client_id, source_row_number, source_row=candidate,
            )
            product_id = product_record.get("id")
            if not product_id:
                return jsonify({"matched": False, "reason": "no-match"})
            updated, linked_receipt, linked_product, link_error = _link_merchandise_to_product(entry_id, product_id)
            if link_error:
                return link_error
            return jsonify({
                "matched": True,
                "via": "source",
                "sourceRowNumber": source_row_number,
                "merchandise": _shape_verification_entry(
                    updated, linked_receipt or receipt, item_record=linked_product or product_record,
                ),
            })
        updated, linked_receipt, linked_product, link_error = _link_merchandise_to_product(
            entry_id, candidate.get("id", ""),
        )
        if link_error:
            return link_error
    except requests.RequestException:
        current_app.logger.exception("Auto-match source sheet read failed")
        return jsonify({"matched": False, "reason": "source-unavailable"})
    except requests.HTTPError as error:
        return airtable_err(error)
    except (ValueError, RuntimeError):
        current_app.logger.exception("Auto-match activation failed")
        return jsonify({"matched": False, "reason": "no-match"})
    return jsonify({
        "matched": True,
        "via": "product",
        "merchandise": _shape_verification_entry(
            updated, linked_receipt or receipt, item_record=linked_product,
        ),
    })


@api.post("/merchandise/<entry_id>/activate-source-row")
@api.post("/merchandise/review/<entry_id>/activate-source-row")
def activate_source_row_for_merchandise(entry_id):
    body = request.get_json(silent=True) or {}
    try:
        source_row_number = int(body.get("sourceRowNumber") or 0)
    except (TypeError, ValueError):
        return err("Choose a source row to activate.", 400)
    if source_row_number <= 0:
        return err("Choose a source row to activate.", 400)

    entry, receipt, access_error = _permitted_merchandise_or_error(entry_id)
    if access_error:
        return access_error
    receipt_fields = receipt.get("fields", {}) if receipt else {}
    client_ids = _as_list(receipt_fields.get(C.F_RECEIPT_CLIENT, []))
    client_id = client_ids[0] if client_ids else ""
    if not client_id:
        return err("This Merchandise needs a Shipment client before source activation.", 400)

    try:
        client_record = _topco_client_record(client_id)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not client_record:
        return err("Source row activation is currently available only for Topco Shipments.", 400)

    try:
        source_row = _topco_source_row_by_number(source_row_number)
        product_record, result, row = _activate_topco_source_product(client_id, source_row_number, source_row=source_row)
        product_id = product_record.get("id")
        if not product_id:
            raise RuntimeError("Source row activation did not return a Product.")
        updated, linked_receipt, linked_product, link_error = _link_merchandise_to_product(entry_id, product_id)
        if link_error:
            return link_error
    except requests.RequestException as exc:
        current_app.logger.exception("Shipment source row activation source sheet read failed")
        return err(f"Could not read source sheet: {exc}", 502)
    except requests.HTTPError as error:
        return airtable_err(error)
    except ValueError as exc:
        return err(str(exc), 400)
    except RuntimeError as exc:
        return err(str(exc), 500)

    shaped_product = _shape_item(
        linked_product or product_record,
        clients_by_id=_clients_by_id(),
        issues_by_item_id=_issues_by_item_id(),
    )
    return jsonify({
        "activated": True,
        "action": row.get("action"),
        "sourceRowNumber": source_row_number,
        "summary": result.get("summary", {}),
        "product": shaped_product,
        "merchandise": _shape_verification_entry(
            updated,
            linked_receipt or receipt,
            item_record=linked_product or product_record,
        ),
    })


@api.post("/verification/entries/<entry_id>/validate")
@api.post("/merchandise/<entry_id>/validate")
@api.post("/merchandise/review/<entry_id>/validate")
def validate_verification_entry(entry_id):
    body = request.get_json(silent=True) or {}
    status = (body.get("status") or "").strip()
    if status not in {"Validated", "Received", "Issue", "Ready to Ship", "Shipped", "Disposed"}:
        return err("status must be one of: Validated, Received, Issue, Ready to Ship, Shipped, Disposed")

    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    fields = entry.get("fields", {})
    item_ids = _as_list(fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
    if status == "Validated":
        if not item_ids:
            return err("A Product must be linked before Merchandise can be validated.", 400)
        issues = _issues_by_item_id().get(item_ids[0], [])
        if _blocking_merchandise_issues(issues):
            return err("Resolve blocking Merchandise Issues before validation.", 400)

    update_fields = {}
    if status == "Validated":
        update_fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS["awaiting-photo-release"]
        update_fields.update(_merch_status_normalization_fields(fields))
    else:
        update_fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = status

    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/merchandise/review/<entry_id>/remove-match")
def remove_merchandise_review_match(entry_id):
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    try:
        # The link only. Whether the merchandise arrived as described, and how far
        # it has been reviewed, are different facts: unlinking a Product says
        # nothing about either. Resetting them sent an accepted item back to New,
        # which took its Deliverables step away and offered acceptance again.
        updated = _update_receipt_entry_record(entry_id, {C.F_RECEIPT_ENTRY_ITEM: []})
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/merchandise/review/<entry_id>/waiting-product-data")
def mark_merchandise_waiting_for_product_data(entry_id):
    body = request.get_json(silent=True) or {}
    note = (body.get("note") or body.get("notes") or "").strip()
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    update_fields = {
        C.F_RECEIPT_ENTRY_ITEM: [],
        C.F_RECEIPT_ENTRY_PLANNING_STATUS: "Needs More Information",
        C.F_RECEIPT_ENTRY_MERCH_STATUS: "Received",
    }
    if note:
        existing_notes = entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_NOTES, "") or ""
        update_fields[C.F_RECEIPT_ENTRY_NOTES] = f"{existing_notes}\n{note}".strip()

    try:
        updated = _update_receipt_entry_record(entry_id, update_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_verification_entry(updated, receipt))


@api.post("/merchandise/review/<entry_id>/issue")
def create_merchandise_review_issue(entry_id):
    body = request.get_json(silent=True) or {}
    issue_type = (body.get("type") or "Unknown Item").strip()
    description = (body.get("description") or body.get("issue") or "Merchandise issue").strip()
    notes = (body.get("notes") or "").strip()

    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)

    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    receipt = _first_permitted_receipt(linked_receipts)
    if linked_receipts and receipt is None:
        return _forbidden()

    entry_fields = entry.get("fields", {})
    item_ids = _as_list(entry_fields.get(C.F_RECEIPT_ENTRY_ITEM, []))
    if item_ids and not _all_linked_records_permitted(C.PRODUCTS_TABLE, item_ids):
        return _forbidden()

    issue_fields = {
        C.F_ISSUE_NAME: description,
        C.F_ISSUE_TYPE: issue_type,
        C.F_ISSUE_STATUS: body.get("status") or "Open",
        C.F_ISSUE_PRIORITY: body.get("priority") or "Normal",
        C.F_ISSUE_OPENED: body.get("opened") or _now_iso(),
        C.F_ISSUE_NOTES: notes,
    }
    if item_ids:
        issue_fields[C.F_ISSUE_ITEM] = item_ids
    photo_metadata = _photo_metadata_from_entry(entry_fields)
    if photo_metadata:
        image_keys = [item.get("object_key") for item in photo_metadata if item.get("object_key")]
        if image_keys:
            image_note = "R2 image references:\n" + "\n".join(image_keys)
            issue_fields[C.F_ISSUE_NOTES] = f"{notes}\n\n{image_note}".strip()

    try:
        issue = airtable.create_record(C.ISSUES_TABLE, issue_fields, by_field_id=False)
        updated = _update_receipt_entry_record(entry_id, {
            C.F_RECEIPT_ENTRY_MERCH_STATUS: "Issue",
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED: False,
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT: "",
            C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY: "",
        })
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({
        "issue": _shape_issue(issue),
        "merchandise": _shape_verification_entry(updated, receipt, issues_by_item_id={item_ids[0]: [_shape_issue(issue)]} if item_ids else {}),
    }), 201


@api.post("/receiving/photos")
def upload_receiving_photos():
    receipt_id = (request.form.get("receiptId") or "").strip()
    receipt_entry_id = (request.form.get("receiptEntryId") or request.form.get("entryId") or "").strip()
    if receipt_id and receipt_entry_id:
        return _upload_receiving_entry_photos(receipt_id, receipt_entry_id)

    return err("Create the received item before uploading permanent photos.", 409)


@api.get("/receiving/photo-storage/status")
def receiving_photo_storage_status():
    storage = _photo_storage()
    required = {
        "R2_ACCOUNT_ID": bool(C.R2_ACCOUNT_ID),
        "R2_ACCESS_KEY_ID": bool(C.R2_ACCESS_KEY_ID),
        "R2_SECRET_ACCESS_KEY": bool(C.R2_SECRET_ACCESS_KEY),
        "R2_BUCKET_NAME": bool(C.R2_BUCKET_NAME),
        "R2_PUBLIC_BASE_URL": bool(C.R2_PUBLIC_BASE_URL),
    }
    return jsonify({
        "mode": storage.mode,
        "bucketName": C.R2_BUCKET_NAME or "",
        "publicBaseUrlConfigured": bool(C.R2_PUBLIC_BASE_URL),
        "requiredVariables": required,
        "ready": storage.mode == "r2" and all(required.values()),
    })


@api.get("/shipments/<shipment_id>/photos")
def list_shipment_photos(shipment_id):
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, shipment_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    return jsonify({"photos": _shipment_photo_metadata_from_fields(receipt.get("fields", {}))})


@api.post("/shipments/<shipment_id>/photos")
def upload_shipment_photos(shipment_id):
    files = request.files.getlist("photos")
    if not files:
        return err("Add at least one photo.")
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, shipment_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()

    storage = _photo_storage()
    current_metadata = _shipment_photo_metadata_from_fields(receipt.get("fields", {}))
    existing_count = len([photo for photo in current_metadata if photo.get("active", True)])
    uploaded_photos = []
    user_id = _current_user_id()
    for uploaded in files:
        if not uploaded or not uploaded.filename:
            continue
        try:
            photo = storage.upload_shipment_photo(
                uploaded,
                shipment_id,
                sort_order=existing_count + len(uploaded_photos) + 1,
                uploaded_by=user_id,
            )
            uploaded_photos.append(photo)
        except ReceivingPhotoValidationError as error:
            return err(str(error))
        except ReceivingPhotoConfigError as error:
            return err(str(error), 500)
        except ReceivingPhotoCollisionError:
            return err("Photo could not be uploaded without overwriting an existing object.", 409)
        except ReceivingPhotoStorageError:
            return err("Photo could not be uploaded.", 502)
    if not uploaded_photos:
        return err("Add at least one photo.")

    metadata_payload = current_metadata + uploaded_photos
    try:
        current_notes = receipt.get("fields", {}).get(C.F_RECEIPT_NOTES, "")
        updated = airtable.update_record(
            C.SHIPMENTS_TABLE,
            shipment_id,
            {C.F_RECEIPT_NOTES: _shipment_notes_with_photo_metadata(current_notes, metadata_payload)},
            by_field_id=False,
        )
    except requests.HTTPError as error:
        for photo in uploaded_photos:
            try:
                storage.delete_photo(photo.get("object_key"))
            except ReceivingPhotoStorageError:
                pass
        return airtable_err(error)
    return jsonify({
        "photos": uploaded_photos,
        "shipment": _shape_receipt(updated, entries_by_receipt=_receipt_entries_by_receipt_id([shipment_id])),
    })


@api.delete("/shipments/<shipment_id>/photos/<photo_id>")
def delete_shipment_photo(shipment_id, photo_id):
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, shipment_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    current_metadata = _shipment_photo_metadata_from_fields(receipt.get("fields", {}))
    target = next((photo for photo in current_metadata if photo.get("photo_id") == photo_id), None)
    if not target:
        return err("Shipment photo not found.", 404)
    updated_metadata = [photo for photo in current_metadata if photo.get("photo_id") != photo_id]
    try:
        current_notes = receipt.get("fields", {}).get(C.F_RECEIPT_NOTES, "")
        updated = airtable.update_record(
            C.SHIPMENTS_TABLE,
            shipment_id,
            {C.F_RECEIPT_NOTES: _shipment_notes_with_photo_metadata(current_notes, updated_metadata)},
            by_field_id=False,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    try:
        _photo_storage().delete_photo(target.get("object_key"))
    except (ReceivingPhotoValidationError, ReceivingPhotoConfigError, ReceivingPhotoStorageError):
        pass
    return jsonify({
        "deleted": True,
        "photoId": photo_id,
        "objectKey": target.get("object_key", ""),
        "shipment": _shape_receipt(updated, entries_by_receipt=_receipt_entries_by_receipt_id([shipment_id])),
    })


@api.get("/receiving/photos/<path:filename>")
def receiving_photo(filename):
    return err("Receiving photos are stored in Cloudflare R2.", 410)


@api.post("/receiving/<receipt_id>/entries/<receipt_entry_id>/photos")
def upload_receiving_entry_photos(receipt_id, receipt_entry_id):
    return _upload_receiving_entry_photos(receipt_id, receipt_entry_id)


@api.delete("/receiving/photos")
def delete_receiving_photo():
    body = request.get_json(silent=True) or {}
    object_key = (body.get("objectKey") or body.get("object_key") or "").strip()
    try:
        result = _photo_storage().delete_photo(object_key)
    except ReceivingPhotoValidationError as error:
        return err(str(error))
    except ReceivingPhotoConfigError as error:
        return err(str(error), 500)
    except ReceivingPhotoStorageError:
        return err("Photo could not be deleted.", 502)
    return jsonify(result)


@api.delete("/receiving/<record_id>/entries/<entry_id>/photos")
def delete_receiving_entry_photo(record_id, entry_id):
    """Delete a single photo from a receipt entry: removes from R2 and updates Airtable metadata."""
    body = request.get_json(silent=True) or {}
    object_key = (body.get("objectKey") or body.get("object_key") or "").strip()
    if not object_key:
        return err("objectKey is required.")
    try:
        entry = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    # Verify the entry belongs to this receipt
    linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    if record_id not in linked_receipts:
        return _forbidden()
    # Delete from R2
    try:
        _photo_storage().delete_photo(object_key)
    except ReceivingPhotoValidationError as error:
        return err(str(error))
    except ReceivingPhotoConfigError as error:
        return err(str(error), 500)
    except ReceivingPhotoStorageError:
        return err("Photo could not be deleted.", 502)
    # Update Airtable metadata to remove this photo
    current_metadata = _photo_metadata_from_entry(entry.get("fields", {}))
    updated_metadata = [
        p for p in current_metadata
        if (p.get("object_key") or p.get("objectKey") or "") != object_key
    ]
    try:
        updated = airtable.update_record(
            C.MERCHANDISE_TABLE,
            entry_id,
            {C.F_RECEIPT_ENTRY_PHOTO_METADATA: json.dumps(updated_metadata)},
            by_field_id=False,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    shaped = _shape_receipt_entry_with_linked_product(updated)
    shaped["photoMetadata"] = updated_metadata
    return jsonify({"deleted": True, "objectKey": object_key, "entry": shaped})


def _local_received_datetime(value):
    raw = (value or "").strip()
    parsed = None
    if raw:
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
    if parsed is None:
        return datetime.now().astimezone()
    if parsed.tzinfo is not None:
        return parsed.astimezone()
    return parsed


def _delivery_folder_base(client_name, received):
    client_segment = sanitize_path_segment(client_name or "Unknown", "Unknown")
    local_received = _local_received_datetime(received)
    return f"{client_segment}-{local_received.strftime('%Y-%m-%d-%H-%M')}"


def _receipt_client_name(receipt):
    fields = receipt.get("fields", {}) if receipt else {}
    client_ids = _as_list(fields.get(C.F_RECEIPT_CLIENT, []))
    if not client_ids:
        return "Unknown"
    try:
        return _client_name(_client_record(client_ids[0])) or "Unknown"
    except requests.HTTPError:
        return "Unknown"


def _delivery_folder_for_receipt(receipt):
    fields = receipt.get("fields", {}) if receipt else {}
    base = _delivery_folder_base(_receipt_client_name(receipt), fields.get(C.F_RECEIPT_RECEIVED, ""))
    receipt_id = receipt.get("id", "")
    try:
        receipts = _list_all_records(C.SHIPMENTS_TABLE)
    except requests.HTTPError:
        return base

    collisions = []
    for candidate in receipts:
        candidate_base = _delivery_folder_base(
            _receipt_client_name(candidate),
            candidate.get("fields", {}).get(C.F_RECEIPT_RECEIVED, ""),
        )
        if candidate_base == base:
            collisions.append(candidate)
    collisions.sort(key=lambda item: (
        _local_received_datetime(item.get("fields", {}).get(C.F_RECEIPT_RECEIVED, "")).isoformat(),
        item.get("id", ""),
    ))
    for index, candidate in enumerate(collisions, start=1):
        if candidate.get("id") == receipt_id:
            return base if index == 1 else f"{base}-{index}"
    return base


def _sequence_from_object_key(object_key, delivery_folder):
    pattern = rf"^receiving/{re.escape(delivery_folder)}/{re.escape(delivery_folder)}-(\d+)\.[A-Za-z0-9]+$"
    match = re.match(pattern, str(object_key or ""))
    return int(match.group(1)) if match else None


def _existing_receipt_photo_metadata(receipt_id, current_entry_record):
    metadata = []
    try:
        entries = _list_all_records(C.MERCHANDISE_TABLE)
    except requests.HTTPError:
        entries = [current_entry_record]
    seen_current = False
    for entry in entries:
        if entry.get("id") == current_entry_record.get("id"):
            seen_current = True
        linked_receipts = _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
        if receipt_id in linked_receipts:
            metadata.extend(_photo_metadata_from_entry(entry.get("fields", {})))
    if not seen_current:
        metadata.extend(_photo_metadata_from_entry(current_entry_record.get("fields", {})))
    return metadata


def _next_delivery_photo_sequence(delivery_folder, existing_metadata, existing_object_keys):
    used = set()
    for item in existing_metadata or []:
        sequence = _sequence_from_object_key(item.get("object_key"), delivery_folder)
        if sequence:
            used.add(sequence)
    for key in existing_object_keys or []:
        sequence = _sequence_from_object_key(key, delivery_folder)
        if sequence:
            used.add(sequence)
    return max(used, default=0) + 1


def _upload_receiving_entry_photos(receipt_id, receipt_entry_id):
    files = request.files.getlist("photos")
    if not files:
        return err("Add at least one photo.")
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, receipt_id, by_field_id=False)
        entry_record = airtable.get_record(C.MERCHANDISE_TABLE, receipt_entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    linked_receipts = _as_list(entry_record.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    if receipt_id not in linked_receipts:
        return err("Merchandise does not belong to this Shipment.", 404)

    storage = _photo_storage()
    uploaded_photos = []
    delivery_folder = _delivery_folder_for_receipt(receipt)
    existing_metadata = _existing_receipt_photo_metadata(receipt_id, entry_record)
    try:
        existing_object_keys = storage.list_object_keys(f"receiving/{delivery_folder}/")
    except (ReceivingPhotoConfigError, ReceivingPhotoValidationError) as error:
        return err(str(error), 500 if isinstance(error, ReceivingPhotoConfigError) else 400)
    except ReceivingPhotoStorageError:
        existing_object_keys = []
    existing_keys = set(existing_object_keys)
    next_sequence = _next_delivery_photo_sequence(delivery_folder, existing_metadata, existing_keys)
    for uploaded in files:
        if not uploaded or not uploaded.filename:
            continue
        try:
            photo = storage.upload_photo(
                uploaded,
                receipt_id,
                receipt_entry_id,
                delivery_folder=delivery_folder,
                sequence_number=next_sequence,
                existing_keys=existing_keys,
            )
            uploaded_photos.append(photo)
            existing_keys.add(photo["object_key"])
            next_sequence = _next_delivery_photo_sequence(delivery_folder, existing_metadata + uploaded_photos, existing_keys)
        except ReceivingPhotoValidationError as error:
            return err(str(error))
        except ReceivingPhotoConfigError as error:
            return err(str(error), 500)
        except ReceivingPhotoStorageError:
            return err("Photo could not be uploaded.", 502)
    if not uploaded_photos:
        return err("Add at least one photo.")

    current_fields = entry_record.get("fields", {})
    metadata_payload = _photo_metadata_from_entry(current_fields) + uploaded_photos
    try:
        updated = airtable.update_record(
            C.MERCHANDISE_TABLE,
            receipt_entry_id,
            {
                C.F_RECEIPT_ENTRY_PHOTO_METADATA: json.dumps(_canonical_photo_manifest(metadata_payload)),
            },
            by_field_id=False,
        )
    except requests.HTTPError as error:
        return airtable_err(error)
    shaped = _shape_receipt_entry_with_linked_product(updated)
    shaped["photoMetadata"] = metadata_payload
    return jsonify({"photos": uploaded_photos, "entry": shaped})


@api.get("/receiving/<record_id>")
def get_receiving_session(record_id):
    try:
        record = airtable.get_record(C.SHIPMENTS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(record.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    entries_by_receipt = _receipt_entries_by_receipt_id([record_id])
    return jsonify(_shape_receipt(record, entries_by_receipt=entries_by_receipt))


@api.patch("/receiving/<record_id>")
def update_receiving_session(record_id):
    body = request.get_json(silent=True) or {}
    try:
        current = airtable.get_record(C.SHIPMENTS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(current.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    fields_or_error = _receipt_update_fields_from_body(body)
    if isinstance(fields_or_error, tuple):
        return fields_or_error
    if not fields_or_error:
        entries_by_receipt = _receipt_entries_by_receipt_id([record_id])
        return jsonify(_shape_receipt(current, entries_by_receipt=entries_by_receipt))
    if C.F_RECEIPT_NOTES in fields_or_error:
        fields_or_error[C.F_RECEIPT_NOTES] = _shipment_notes_with_photo_metadata(
            fields_or_error[C.F_RECEIPT_NOTES],
            _shipment_photo_metadata_from_fields(current.get("fields", {}), include_urls=False),
        )
    try:
        updated = airtable.update_record(C.SHIPMENTS_TABLE, record_id, fields_or_error, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    entries_by_receipt = _receipt_entries_by_receipt_id([record_id])
    return jsonify(_shape_receipt(updated, entries_by_receipt=entries_by_receipt))


@api.post("/receiving/sessions")
def create_receiving_session():
    body = request.get_json(silent=True) or {}
    fields_or_error = _receipt_fields_from_body(body)
    if isinstance(fields_or_error, tuple):
        return fields_or_error
    try:
        data = airtable.create_record(C.SHIPMENTS_TABLE, fields_or_error, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    receipt_fields = data.get("fields", {})
    _create_history_event(
        "Receiving Logged",
        user_ids=receipt_fields.get(C.F_RECEIPT_RECEIVER),
    )
    return jsonify(_shape_receipt(data, entries_by_receipt={data["id"]: []})), 201


@api.post("/receiving/<record_id>/entries")
def create_receiving_entry(record_id):
    body = request.get_json(silent=True) or {}
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, record_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    errors = _validate_receipt_entries([body])
    if errors:
        return err(errors[0])
    receipt_name = receipt.get("fields", {}).get(C.F_RECEIPT_NAME) or record_id
    existing_entries = _receipt_entries_by_receipt_id([record_id]).get(record_id, [])
    try:
        entry_fields = _receipt_entry_fields(body, record_id, len(existing_entries) + 1, receipt_name)
        match_fields = _receipt_entry_match_fields(body, receipt)
        if isinstance(match_fields, tuple):
            return match_fields
        entry_fields.update(match_fields)
        entry_data = _create_receipt_entry_record(entry_fields)
    except requests.HTTPError as error:
        return airtable_err(error)
    _record_merchandise_history(entry_data.get("id"), MERCHANDISE_CREATED_EVENT)
    return jsonify(_shape_receipt_entry_with_linked_product(entry_data)), 201


@api.patch("/receiving/<record_id>/entries/<entry_id>")
def update_receiving_entry(record_id, entry_id):
    body = request.get_json(silent=True) or {}
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, record_id, by_field_id=False)
        entry_record = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    linked_receipts = _as_list(entry_record.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    if record_id not in linked_receipts:
        return err("Merchandise does not belong to this Shipment.", 404)
    fields_or_error = _receipt_entry_update_fields_from_body(body)
    if isinstance(fields_or_error, tuple):
        return fields_or_error
    intake_fields = _intake_decision_fields_from_body(body, entry_record.get("fields", {}))
    if isinstance(intake_fields, tuple):
        return intake_fields
    fields_or_error.update(intake_fields)
    item_ids_for_match = _as_list(body.get("itemIds") or body.get("itemId"))
    should_update_match = bool(item_ids_for_match) or "matchStatus" in body
    if should_update_match:
        match_fields = _receipt_entry_match_fields(body, receipt)
        if isinstance(match_fields, tuple):
            return match_fields
        fields_or_error.update(match_fields)
    if not fields_or_error:
        return jsonify(_shape_receipt_entry_with_linked_product(entry_record))
    try:
        updated = _update_receipt_entry_record(entry_id, fields_or_error)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify(_shape_receipt_entry_with_linked_product(updated))


@api.delete("/receiving/<record_id>/entries/<entry_id>")
def delete_receiving_entry(record_id, entry_id):
    try:
        receipt = airtable.get_record(C.SHIPMENTS_TABLE, record_id, by_field_id=False)
        entry_record = airtable.get_record(C.MERCHANDISE_TABLE, entry_id, by_field_id=False)
    except requests.HTTPError as error:
        return airtable_err(error)
    if not _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
        return _forbidden()
    linked_receipts = _as_list(entry_record.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, []))
    if record_id not in linked_receipts:
        return err("Received item does not belong to this delivery.", 404)
    try:
        airtable.delete_record(C.MERCHANDISE_TABLE, entry_id)
    except requests.HTTPError as error:
        return airtable_err(error)
    return jsonify({"deleted": True, "id": entry_id})


@api.post("/receipts")
@api.post("/receiving")
def create_receipt():
    body = request.get_json(silent=True) or {}
    entries = body.get("entries") or []

    if not isinstance(entries, list) or not entries:
        return err("Add at least one merchandise entry.")
    entry_errors = _validate_receipt_entries(entries)
    if entry_errors:
        return err(entry_errors[0])
    fields_or_error = _receipt_fields_from_body(body)
    if isinstance(fields_or_error, tuple):
        return fields_or_error

    created_entry_ids = []
    try:
        data = airtable.create_record(C.SHIPMENTS_TABLE, fields_or_error, by_field_id=False)
        receipt_name = data.get("fields", {}).get(C.F_RECEIPT_NAME) or data["id"]
        shaped_entries = []
        for index, entry in enumerate(entries, start=1):
            entry_fields = _receipt_entry_fields(entry, data["id"], index, receipt_name)
            match_fields = _receipt_entry_match_fields(entry, data)
            if isinstance(match_fields, tuple):
                return match_fields
            entry_fields.update(match_fields)
            entry_data = _create_receipt_entry_record(entry_fields)
            _record_merchandise_history(entry_data.get("id"), MERCHANDISE_CREATED_EVENT)
            created_entry_ids.append(entry_data["id"])
            shaped_entries.append(_shape_receipt_entry_with_linked_product(entry_data))
    except requests.HTTPError as error:
        if created_entry_ids:
            try:
                airtable.delete_records(C.MERCHANDISE_TABLE, created_entry_ids)
            except requests.HTTPError:
                pass
        if "data" in locals() and data.get("id"):
            try:
                airtable.delete_record(C.SHIPMENTS_TABLE, data["id"])
            except requests.HTTPError:
                pass
        return airtable_err(error)

    receipt_fields = data.get("fields", {})
    _create_history_event(
        "Receiving Logged",
        user_ids=receipt_fields.get(C.F_RECEIPT_RECEIVER),
    )
    entries_by_receipt = {data["id"]: shaped_entries}
    notified, notice = _notify_shipment_arrival(data, shaped_entries)
    payload = _shape_receipt(data, entries_by_receipt=entries_by_receipt)
    payload["teamsNotification"] = {"posted": notified, "detail": notice}
    return jsonify(payload), 201


def _notify_shipment_arrival(receipt, shaped_entries):
    """Tell the client's Teams channel that merchandise landed.

    The goods are already recorded by the time this runs, so nothing here is
    allowed to fail the request. A client with no webhook is not notified.
    """
    fields = receipt.get("fields", {})
    client_ids = _as_list(fields.get(C.F_RECEIPT_CLIENT, []))
    if not client_ids:
        return False, "Shipment has no client."
    try:
        client_record = airtable.get_record(C.CLIENTS_TABLE, client_ids[0], by_field_id=False)
    except requests.HTTPError:
        return False, "Could not read the client."
    client_fields = client_record.get("fields", {})
    webhook = client_fields.get(C.F_CLIENT_TEAMS_WEBHOOK, "")
    if not str(webhook or "").strip():
        return False, "No Teams webhook configured for this client."

    items = []
    for entry in shaped_entries:
        name = entry.get("productName") or entry.get("description") or "Unnamed item"
        identifier = entry.get("skuId") or entry.get("observedIdentifier") or ""
        quantity = entry.get("quantity") or 1
        items.append(f"{quantity} x {name}" + (f" - {identifier}" if identifier else ""))

    card = notifier.build_arrival_card(
        client_name=client_fields.get(C.F_CLIENT_NAME, ""),
        shipment_name=fields.get(C.F_RECEIPT_NAME, ""),
        shipment_id=receipt.get("id", ""),
        carrier=fields.get(C.F_RECEIPT_CARRIER, ""),
        tracking=fields.get(C.F_RECEIPT_TRACKING, ""),
        received=fields.get(C.F_RECEIPT_RECEIVED, ""),
        items=items,
    )
    posted, detail = notifier.post_arrival(webhook, card)
    if not posted:
        current_app.logger.warning("Teams arrival notification not sent: %s", detail)
    return posted, detail


def _receipt_fields_from_body(body):
    fields = {}
    receipt = (body.get("receipt") or body.get("name") or "").strip()
    client_id = (body.get("clientId") or "").strip()
    try:
        carrier = _normalize_receipt_carrier(body.get("carrier"))
    except ValueError as error:
        return err(str(error))
    tracking = (body.get("tracking") or body.get("identifier") or "").strip()
    box_quantity = body.get("boxQuantity", body.get("box_quantity"))
    received = (body.get("received") or body.get("receivedDate") or "").strip()
    receiver_ids = body.get("receiverIds") or body.get("receiverId") or []
    location_ids = body.get("locationIds") or body.get("locationId") or []
    notes = (body.get("notes") or "").strip()
    try:
        box_quantity_number = int(box_quantity)
    except (TypeError, ValueError):
        return err("Box Quantity must be at least 1.")
    if box_quantity_number < 1:
        return err("Box Quantity must be at least 1.")

    if client_id:
        if not _client_permitted(client_id):
            return _forbidden()
        fields[C.F_RECEIPT_CLIENT] = [client_id]
    if receipt:
        fields[C.F_RECEIPT_NAME] = receipt
    else:
        fields[C.F_RECEIPT_NAME] = _receipt_name_for_create(client_id, received)
    if carrier:
        fields[C.F_RECEIPT_CARRIER] = carrier
    if tracking:
        fields[C.F_RECEIPT_TRACKING] = tracking
    fields[C.F_RECEIPT_BOX_QUANTITY] = box_quantity_number
    fields[C.F_RECEIPT_RECEIVED] = received or _now_iso()
    if not receiver_ids:
        current_user_id = _current_user_id()
        if current_user_id:
            receiver_ids = [current_user_id]
    if receiver_ids:
        fields[C.F_RECEIPT_RECEIVER] = receiver_ids if isinstance(receiver_ids, list) else [receiver_ids]
    if location_ids:
        fields[C.F_RECEIPT_LOCATION] = location_ids if isinstance(location_ids, list) else [location_ids]
    if notes:
        fields[C.F_RECEIPT_NOTES] = notes
    return fields


def _receipt_update_fields_from_body(body):
    fields = {}
    if "clientId" in body:
        client_id = (body.get("clientId") or "").strip()
        if client_id:
            if not _client_permitted(client_id):
                return _forbidden()
            fields[C.F_RECEIPT_CLIENT] = [client_id]
        else:
            fields[C.F_RECEIPT_CLIENT] = []
    if "carrier" in body:
        try:
            carrier = _normalize_receipt_carrier(body.get("carrier"))
        except ValueError as error:
            return err(str(error))
        fields[C.F_RECEIPT_CARRIER] = carrier
    if "tracking" in body:
        fields[C.F_RECEIPT_TRACKING] = (body.get("tracking") or "").strip()
    if "boxQuantity" in body or "box_quantity" in body:
        box_quantity = body.get("boxQuantity", body.get("box_quantity"))
        try:
            box_quantity_number = int(box_quantity)
        except (TypeError, ValueError):
            return err("Box Quantity must be at least 1.")
        if box_quantity_number < 1:
            return err("Box Quantity must be at least 1.")
        fields[C.F_RECEIPT_BOX_QUANTITY] = box_quantity_number
    if "received" in body or "receivedDate" in body:
        fields[C.F_RECEIPT_RECEIVED] = (body.get("received") or body.get("receivedDate") or "").strip()
    if "locationIds" in body or "locationId" in body:
        location_ids = body.get("locationIds") if "locationIds" in body else body.get("locationId")
        fields[C.F_RECEIPT_LOCATION] = location_ids if isinstance(location_ids, list) else ([location_ids] if location_ids else [])
    if "notes" in body:
        fields[C.F_RECEIPT_NOTES] = (body.get("notes") or "").strip()
    return fields


SHIPMENT_PHOTO_METADATA_START = "[[MARKS_PHOTO_SHIPMENT_PHOTO_METADATA]]"
SHIPMENT_PHOTO_METADATA_END = "[[/MARKS_PHOTO_SHIPMENT_PHOTO_METADATA]]"


def _split_shipment_notes_metadata(notes):
    raw = str(notes or "")
    start = raw.find(SHIPMENT_PHOTO_METADATA_START)
    end = raw.find(SHIPMENT_PHOTO_METADATA_END, start + len(SHIPMENT_PHOTO_METADATA_START)) if start >= 0 else -1
    if start < 0 or end < 0:
        return raw.strip(), []
    metadata_text = raw[start + len(SHIPMENT_PHOTO_METADATA_START):end].strip()
    visible = (raw[:start] + raw[end + len(SHIPMENT_PHOTO_METADATA_END):]).strip()
    try:
        parsed = json.loads(metadata_text) if metadata_text else []
    except (TypeError, ValueError):
        parsed = []
    return visible, parsed if isinstance(parsed, list) else []


def _shipment_notes_with_photo_metadata(notes, metadata):
    visible, _existing = _split_shipment_notes_metadata(notes)
    manifest = json.dumps(_canonical_photo_manifest(metadata), separators=(",", ":"))
    block = f"{SHIPMENT_PHOTO_METADATA_START}\n{manifest}\n{SHIPMENT_PHOTO_METADATA_END}"
    return f"{visible}\n\n{block}".strip() if visible else block


def _receipt_name_for_create(client_id, received):
    client_name = "Unknown"
    if client_id:
        try:
            client_name = _client_name(_client_record(client_id)) or client_name
        except requests.HTTPError:
            pass
    timestamp = _format_receipt_name_time(received)
    return f"{client_name} - {timestamp}"


def _format_receipt_name_time(value):
    parsed = None
    raw = (value or "").strip()
    if raw:
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
    if parsed is None:
        parsed = datetime.now()
    return parsed.strftime("%Y-%m-%d %H:%M")


def _validate_receipt_entries(entries):
    errors = []
    for index, entry in enumerate(entries, start=1):
        quantity = entry.get("quantity", 1) if isinstance(entry, dict) else None
        try:
            quantity_number = int(quantity)
        except (TypeError, ValueError):
            errors.append(f"Entry {index} quantity must be at least 1.")
            continue
        if quantity_number < 1:
            errors.append(f"Entry {index} quantity must be at least 1.")
    return errors


RECEIPT_CARRIER_OPTIONS = {
    "ups": "UPS",
    "u.p.s.": "UPS",
    "fedex": "FedEx",
    "fed ex": "FedEx",
    "federal express": "FedEx",
    "usps": "USPS",
    "u.s.p.s.": "USPS",
    "dhl": "DHL",
    "courier": "Courier",
    "freight": "Freight",
    "hand delivery": "Hand Delivery",
    "hand-delivery": "Hand Delivery",
    "internal": "Internal",
    "other": "Other",
}


def _normalize_receipt_carrier(value):
    carrier = (value or "").strip()
    if not carrier:
        return ""
    normalized = RECEIPT_CARRIER_OPTIONS.get(re.sub(r"\s+", " ", carrier.lower()))
    if normalized:
        return normalized
    allowed = ", ".join(sorted(set(RECEIPT_CARRIER_OPTIONS.values())))
    raise ValueError(f"Carrier must be one of: {allowed}.")


def _receipt_entry_fields(entry, receipt_id, index, receipt_name):
    entry = entry if isinstance(entry, dict) else {}
    quantity = int(entry.get("quantity") or 1)
    product_name = (
        entry.get("productName")
        or entry.get("product_name")
        or entry.get("name")
        or ""
    ).strip()
    sku_id = (
        entry.get("skuId")
        or entry.get("sku_id")
        or entry.get("observedIdentifier")
        or entry.get("identifier")
        or ""
    ).strip()
    fields = {
        C.F_RECEIPT_ENTRY_NAME: product_name or "Unnamed Product",
        C.F_RECEIPT_ENTRY_RECEIPT: [receipt_id],
        C.F_RECEIPT_ENTRY_QUANTITY: quantity,
        C.F_RECEIPT_ENTRY_MERCH_STATUS: "Received",
        C.F_RECEIPT_ENTRY_PLANNING_STATUS: "New",
    }
    location_ids = entry.get("locationIds") or entry.get("locationId") or []
    condition = (entry.get("condition") or "").strip()
    description = (entry.get("description") or "").strip()
    notes = (entry.get("notes") or "").strip()
    if sku_id:
        fields[C.F_RECEIPT_ENTRY_SKU_ID] = sku_id
    if location_ids:
        fields[C.F_RECEIPT_ENTRY_LOCATION] = location_ids if isinstance(location_ids, list) else [location_ids]
    if condition:
        fields[C.F_RECEIPT_ENTRY_CONDITION] = condition
    if description:
        fields[C.F_RECEIPT_ENTRY_DESCRIPTION] = description
    if notes:
        fields[C.F_RECEIPT_ENTRY_NOTES] = notes
    return fields


def _receipt_entry_match_fields(body, receipt):
    fields = {}
    item_ids = _as_list(body.get("itemIds") or body.get("itemId"))
    item_id = item_ids[0] if item_ids else ""
    match_status = (body.get("matchStatus") or "").strip()
    if item_id:
        try:
            item = airtable.get_record(C.PRODUCTS_TABLE, item_id, by_field_id=False)
        except requests.HTTPError as error:
            return airtable_err(error)
        item_client_ids = _as_list(item.get("fields", {}).get(C.F_ITEM_CLIENT, []))
        if not _client_ids_permitted(item_client_ids):
            return _forbidden()
        receipt_client_ids = _as_list(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, []))
        if receipt_client_ids and item_client_ids and not (set(receipt_client_ids) & set(item_client_ids)):
            return err("Product does not belong to this Shipment client.", 403)
        fields[C.F_RECEIPT_ENTRY_ITEM] = [item_id]
    elif match_status == "Matched":
        return err("Choose a Product before marking this Merchandise matched.")
    else:
        fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = "Received"
    return fields


def _receipt_entry_update_fields_from_body(body):
    fields = {}
    if any(key in body for key in ("productName", "product_name", "name")):
        product_name = (
            body.get("productName")
            or body.get("product_name")
            or body.get("name")
            or ""
        ).strip()
        fields[C.F_RECEIPT_ENTRY_NAME] = product_name or "Unnamed Product"
    if any(key in body for key in ("skuId", "sku_id", "observedIdentifier", "identifier")):
        sku_id = (
            body.get("skuId")
            or body.get("sku_id")
            or body.get("observedIdentifier")
            or body.get("identifier")
            or ""
        ).strip()
        fields[C.F_RECEIPT_ENTRY_SKU_ID] = sku_id
    if "quantity" in body:
        try:
            quantity = int(body.get("quantity") or 1)
        except (TypeError, ValueError):
            return err("Quantity must be at least 1.")
        if quantity < 1:
            return err("Quantity must be at least 1.")
        fields[C.F_RECEIPT_ENTRY_QUANTITY] = quantity
    if "locationIds" in body or "locationId" in body:
        location_ids = body.get("locationIds") if "locationIds" in body else body.get("locationId")
        fields[C.F_RECEIPT_ENTRY_LOCATION] = location_ids if isinstance(location_ids, list) else ([location_ids] if location_ids else [])
    if "condition" in body:
        fields[C.F_RECEIPT_ENTRY_CONDITION] = (body.get("condition") or "").strip()
    if "description" in body:
        fields[C.F_RECEIPT_ENTRY_DESCRIPTION] = (body.get("description") or "").strip()
    if "notes" in body:
        fields[C.F_RECEIPT_ENTRY_NOTES] = (body.get("notes") or "").strip()
    if "photoMetadata" in body:
        manifest = _canonical_photo_manifest(body.get("photoMetadata") or [])
        if manifest:
            fields[C.F_RECEIPT_ENTRY_PHOTO_METADATA] = json.dumps(manifest)
    if "merchStatus" in body:
        merch_status = (body.get("merchStatus") or "").strip()
        if merch_status:
            normalized = _normalized_merch_status(merch_status)
            if normalized in MERCH_STATUS_VALUES:
                fields[C.F_RECEIPT_ENTRY_MERCH_STATUS] = normalized
    if "planningStatusLabel" in body:
        planning_label_value = _validate_planning_status_label(_intake_decision_value(body, "planningStatusLabel"))
        if isinstance(planning_label_value, tuple):
            return planning_label_value
        fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = planning_label_value
    return fields


def _canonical_planning_status_fields(fields):
    """Normalize any Planning Status write to a canonical stored label."""
    fields = dict(fields)
    if C.F_RECEIPT_ENTRY_PLANNING_STATUS in fields:
        fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS] = PLANNING_STATUS_LABELS.get(
            _normalized_planning_status(fields[C.F_RECEIPT_ENTRY_PLANNING_STATUS]),
            "New",
        )
    return fields


def _create_receipt_entry_record(fields):
    return airtable.create_record(
        C.MERCHANDISE_TABLE, _canonical_planning_status_fields(fields), by_field_id=False
    )


def _update_receipt_entry_record(entry_id, fields):
    # Deliberately no typecast: Planning Status writes are normalized above, so an
    # Airtable rejection means something wrote a value outside the canonical set.
    # typecast would silently invent the option instead, which is how stray choices
    # appeared on Workstream Cards.
    return airtable.update_record(
        C.MERCHANDISE_TABLE, entry_id, _canonical_planning_status_fields(fields), by_field_id=False
    )


def _receipt_entries_by_receipt_id(receipt_ids):
    receipt_ids = set(receipt_ids or [])
    grouped = {receipt_id: [] for receipt_id in receipt_ids}
    if not receipt_ids:
        return grouped
    try:
        entries = _list_all_records(C.MERCHANDISE_TABLE)
    except requests.HTTPError:
        return grouped
    linked_product_ids = [
        product_id
        for entry in entries
        for product_id in _as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM, []))
    ]
    products_by_id = _products_by_id_for_ids(linked_product_ids)
    clients_by_id = _clients_by_id()
    for entry in entries:
        linked_receipts = set(_as_list(entry.get("fields", {}).get(C.F_RECEIPT_ENTRY_RECEIPT, [])))
        for receipt_id in linked_receipts & receipt_ids:
            grouped.setdefault(receipt_id, []).append(_shape_receipt_entry(entry, products_by_id=products_by_id, clients_by_id=clients_by_id))
    return grouped


def _photo_with_source(photo, source):
    next_photo = dict(photo or {})
    next_photo["source"] = source
    next_photo["photoType"] = source
    next_photo.setdefault("label", "Shipment Photo" if source == "shipment" else "Item Photo")
    return next_photo


def _item_photo_metadata_from_entry(fields, include_urls=True):
    return [_photo_with_source(photo, "item") for photo in _photo_metadata_from_entry(fields, include_urls=include_urls)]


def _shipment_photo_metadata_from_fields(fields, include_urls=True):
    _visible_notes, notes_metadata = _split_shipment_notes_metadata(fields.get(C.F_RECEIPT_NOTES, ""))
    if notes_metadata:
        metadata = _photo_metadata_from_manifest(notes_metadata, include_urls=include_urls)
    elif fields.get(C.F_RECEIPT_PHOTO_METADATA):
        metadata = _photo_metadata_from_fields(fields, C.F_RECEIPT_PHOTO_METADATA, include_urls=include_urls)
    else:
        metadata = []
    return [_photo_with_source(photo, "shipment") for photo in metadata]


def _combine_item_and_shipment_photos(item_photos=None, shipment_photos=None):
    return list(item_photos or []) + list(shipment_photos or [])


def _with_shipment_photos(shaped_entry, receipt):
    receipt_fields = receipt.get("fields", {}) if receipt else {}
    item_photos = shaped_entry.get("itemPhotos") or shaped_entry.get("photos") or []
    shipment_photos = _shipment_photo_metadata_from_fields(receipt_fields)
    return {
        **shaped_entry,
        "itemPhotos": item_photos,
        "shipmentPhotos": shipment_photos,
        "photos": _combine_item_and_shipment_photos(item_photos, shipment_photos),
    }




def _shape_receipt(r, *, entries_by_receipt=None):
    f = r.get("fields", {})
    entries_by_receipt = entries_by_receipt or {}
    shipment_photos = _shipment_photo_metadata_from_fields(f)
    entries = [_with_shipment_photos(entry, r) for entry in entries_by_receipt.get(r["id"], [])]
    visible_notes, _photo_manifest = _split_shipment_notes_metadata(f.get(C.F_RECEIPT_NOTES, ""))
    return {
        "id": r["id"],
        "name": f.get(C.F_RECEIPT_NAME, ""),
        "receipt": f.get(C.F_RECEIPT_NAME, ""),
        "clientIds": f.get(C.F_RECEIPT_CLIENT, []),
        "itemIds": [],
        "skuIds": [],
        "carrier": f.get(C.F_RECEIPT_CARRIER, ""),
        "tracking": f.get(C.F_RECEIPT_TRACKING, ""),
        "boxQuantity": f.get(C.F_RECEIPT_BOX_QUANTITY, 1),
        "received": f.get(C.F_RECEIPT_RECEIVED, ""),
        "receivedDate": f.get(C.F_RECEIPT_RECEIVED, ""),
        "receiver": "",
        "receiverIds": f.get(C.F_RECEIPT_RECEIVER, []) if isinstance(f.get(C.F_RECEIPT_RECEIVER), list) else [],
        "location": "",
        "locationIds": f.get(C.F_RECEIPT_LOCATION, []) if isinstance(f.get(C.F_RECEIPT_LOCATION), list) else [],
        "photos": shipment_photos,
        "photoMetadata": shipment_photos,
        "shipmentPhotos": shipment_photos,
        "notes": visible_notes,
        "entries": entries,
    }


def _matched_product_summary(product_record, *, clients_by_id=None):
    if not product_record:
        return None
    product = _shape_item(product_record, clients_by_id=clients_by_id or {})
    return {
        "id": product.get("id", ""),
        "name": product.get("name") or product.get("product") or "",
        "product": product.get("product", ""),
        "identifier": product.get("identifier", ""),
        "productId": product.get("productId", ""),
        "gtinUpc": product.get("gtinUpc", ""),
        "primaryMatchKey": product.get("primaryMatchKey", ""),
        "identifierLabel": product.get("identifierLabel", ""),
        "primaryMatchKeyLabel": product.get("primaryMatchKeyLabel", ""),
        "codeType": product.get("codeType", ""),
    }


def _products_by_id_for_ids(product_ids):
    product_ids = {product_id for product_id in _as_list(product_ids) if product_id}
    if not product_ids:
        return {}
    try:
        records = _list_all_records(C.PRODUCTS_TABLE)
    except requests.HTTPError:
        return {}
    return {record["id"]: record for record in records if record.get("id") in product_ids}


def _shape_receipt_entry(r, *, products_by_id=None, clients_by_id=None):
    f = r.get("fields", {})
    manual_product_info = f.get(C.F_RECEIPT_ENTRY_MANUAL_PRODUCT_INFO, "")
    manual_product_info_object = _manual_product_info_object(manual_product_info)
    pending_source_match = manual_product_info_object.get("_sourceMatch") if isinstance(manual_product_info_object.get("_sourceMatch"), dict) else {}
    product_name = f.get(C.F_RECEIPT_ENTRY_NAME, "")
    sku_id = f.get(C.F_RECEIPT_ENTRY_SKU_ID, "")
    item_ids = f.get(C.F_RECEIPT_ENTRY_ITEM, []) if isinstance(f.get(C.F_RECEIPT_ENTRY_ITEM), list) else []
    merch_status = _normalized_merch_status(f.get(C.F_RECEIPT_ENTRY_MERCH_STATUS))
    deliverables = _deliverable_values(f.get(C.F_RECEIPT_ENTRY_DELIVERABLES, ""))
    photo_metadata = _item_photo_metadata_from_entry(f)
    planning_status = _planning_status_for_fields(f)
    planning_label = PLANNING_STATUS_LABELS.get(planning_status, "")
    matched_product = _matched_product_summary(
        (products_by_id or {}).get(item_ids[0]) if item_ids else None,
        clients_by_id=clients_by_id,
    )
    return {
        "id": r["id"],
        "name": product_name,
        "productName": product_name,
        "receiptIds": f.get(C.F_RECEIPT_ENTRY_RECEIPT, []),
        "skuId": sku_id,
        "observedIdentifier": sku_id,
        "quantity": f.get(C.F_RECEIPT_ENTRY_QUANTITY, 0),
        "locationIds": f.get(C.F_RECEIPT_ENTRY_LOCATION, []) if isinstance(f.get(C.F_RECEIPT_ENTRY_LOCATION), list) else [],
        "condition": f.get(C.F_RECEIPT_ENTRY_CONDITION, ""),
        "description": f.get(C.F_RECEIPT_ENTRY_DESCRIPTION, ""),
        "notes": f.get(C.F_RECEIPT_ENTRY_NOTES, ""),
        "photos": photo_metadata,
        "itemPhotos": photo_metadata,
        "shipmentPhotos": [],
        "photoMetadata": photo_metadata,
        "itemIds": item_ids,
        "matchedProduct": matched_product,
        "linkedProduct": matched_product,
        "merchStatus": merch_status,
        # planningStatus is the slug used for queue placement; planningStatusLabel
        # is the stored Airtable label. Both come from the one Planning Status field.
        "planningStatusLabel": planning_label,
        "planningStatus": planning_status,
        "released": bool(f.get(C.F_RECEIPT_ENTRY_RELEASED, False)),
        "released_at": f.get(C.F_RECEIPT_ENTRY_RELEASED_AT, ""),
        "releasedAt": f.get(C.F_RECEIPT_ENTRY_RELEASED_AT, ""),
        "releasedByIds": f.get(C.F_RECEIPT_ENTRY_RELEASED_BY, []) if isinstance(f.get(C.F_RECEIPT_ENTRY_RELEASED_BY), list) else [],
        "merchandiseVerified": bool(f.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED, False)),
        "merchandiseVerifiedAt": f.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED_AT, ""),
        "merchandiseVerifiedBy": str(f.get(C.F_RECEIPT_ENTRY_MERCH_VERIFIED_BY, "") or ""),
        "deliverables": deliverables,
        "manualProductInfo": manual_product_info,
        "pendingSourceMatch": pending_source_match,
    }


def _shape_receipt_entry_with_linked_product(r):
    item_ids = _as_list(r.get("fields", {}).get(C.F_RECEIPT_ENTRY_ITEM, []))
    return _shape_receipt_entry(
        r,
        products_by_id=_products_by_id_for_ids(item_ids),
        clients_by_id=_clients_by_id(),
    )


def _first_permitted_receipt(receipt_ids):
    for receipt_id in _as_list(receipt_ids):
        try:
            receipt = airtable.get_record(C.SHIPMENTS_TABLE, receipt_id, by_field_id=False)
        except requests.HTTPError:
            continue
        if _receipt_client_permitted(receipt.get("fields", {}).get(C.F_RECEIPT_CLIENT, [])):
            return receipt
    return None


def _verification_status_label(status):
    return VERIFICATION_STATUS_LABELS.get(status or "", status or "Awaiting Verification")


def _review_state_for_entry(shaped, linked_item=None, blocking_issues=None):
    merch_status = shaped.get("merchStatus") or ""
    planning_label_value = shaped.get("planningStatusLabel") or ""
    if merch_status == "Issue" or blocking_issues:
        return "Issue"
    if planning_label_value in {"Needs More Information", "Waiting on Information"}:
        return "Waiting for Product Data"
    if planning_label_value == "Awaiting Photo Release":
        return "Validated"
    return "Needs Review"


def _shape_verification_entry(entry, receipt=None, *, item_record=None, issues_by_item_id=None):
    shaped = _shape_receipt_entry(entry)
    entry_fields = entry.get("fields", {})
    receipt_fields = receipt.get("fields", {}) if receipt else {}
    client_ids = receipt_fields.get(C.F_RECEIPT_CLIENT, []) if isinstance(receipt_fields.get(C.F_RECEIPT_CLIENT, []), list) else []
    location_ids = shaped.get("locationIds", [])
    item_ids = shaped.get("itemIds", [])
    linked_item = None
    clients_by_id = {}
    if item_record is None and item_ids:
        try:
            item_record = airtable.get_record(C.PRODUCTS_TABLE, item_ids[0], by_field_id=False)
        except requests.HTTPError:
            item_record = None
    if item_record:
        clients_by_id = _clients_by_id()
        linked_item = _shape_item(item_record, clients_by_id=clients_by_id)
    blocking_issues = _blocking_merchandise_issues((issues_by_item_id or {}).get(item_ids[0], [])) if item_ids else []
    received = receipt_fields.get(C.F_RECEIPT_RECEIVED, "") if receipt else ""
    days_here = _days_here_from_received(received)
    review_state = _review_state_for_entry(shaped, linked_item, blocking_issues)
    required_to_shoot = _evaluate_required_to_shoot({**shaped, "clientIds": client_ids}, linked_item)
    client = clients_by_id.get(client_ids[0]) if client_ids else None
    photo_production = {
        deliverable: _photo_production_status(deliverable, linked_item, client)
        for deliverable in ("Packaging", "Ecomm")
    }
    shaped = _with_shipment_photos(shaped, receipt)
    return {
        **shaped,
        "receipt": {
            "id": receipt.get("id") if receipt else "",
            "name": receipt_fields.get(C.F_RECEIPT_NAME, "") if receipt else "",
            "clientIds": client_ids,
            "carrier": receipt_fields.get(C.F_RECEIPT_CARRIER, "") if receipt else "",
            "tracking": receipt_fields.get(C.F_RECEIPT_TRACKING, "") if receipt else "",
            "boxQuantity": receipt_fields.get(C.F_RECEIPT_BOX_QUANTITY, 1) if receipt else 1,
            "received": receipt_fields.get(C.F_RECEIPT_RECEIVED, "") if receipt else "",
        },
        "clientIds": client_ids,
        "locationId": location_ids[0] if location_ids else "",
        "productName": shaped.get("productName") or shaped.get("description", ""),
        "skuId": shaped.get("skuId", ""),
        "brand": entry_fields.get("Brand", ""),
        "packageSize": entry_fields.get("Package Size", ""),
        "linkedItem": linked_item,
        "blockingIssues": blocking_issues,
        "reviewState": review_state,
        "requiredToShoot": required_to_shoot,
        "photoProduction": photo_production,
        "releaseReady": required_to_shoot.get("ready", False),
        "isUnidentified": not any([shaped.get("productName"), shaped.get("skuId"), shaped.get("description")]),
        "received": received,
        "dateReceived": received,
        "daysHere": days_here,
        "timeHere": _time_here_label(days_here),
        "ageGroup": _age_group_for_days(days_here),
    }


def _photo_metadata_from_entry(fields, include_urls=True):
    return _photo_metadata_from_fields(fields, C.F_RECEIPT_ENTRY_PHOTO_METADATA, include_urls=include_urls)


def _photo_metadata_from_fields(fields, field_name, include_urls=True):
    value = fields.get(field_name, "")
    if isinstance(value, list):
        raw_items = value
    elif isinstance(value, dict):
        raw_items = [value]
    elif not value:
        raw_items = []
    else:
        try:
            parsed = json.loads(value)
        except (TypeError, ValueError):
            raw_items = []
        else:
            raw_items = parsed if isinstance(parsed, list) else []
    return _canonical_photo_manifest(raw_items, include_urls=include_urls)


def _photo_metadata_from_manifest(items, include_urls=True):
    return _canonical_photo_manifest(items if isinstance(items, list) else [], include_urls=include_urls)


def _canonical_photo_manifest(items, include_urls=False):
    manifest = []
    if not isinstance(items, list):
        return manifest
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        object_key = (item.get("object_key") or item.get("objectKey") or "").strip()
        if not object_key:
            continue
        try:
            sort_order = int(item.get("sort_order") or item.get("sortOrder") or index)
        except (TypeError, ValueError):
            sort_order = index
        entry = {
            "object_key": object_key,
            "sort_order": sort_order,
        }
        photo_id = item.get("photo_id") or item.get("photoId")
        if photo_id:
            entry["photo_id"] = str(photo_id)
        shipment_id = item.get("shipment_id") or item.get("shipmentId")
        if shipment_id:
            entry["shipment_id"] = str(shipment_id)
        source = item.get("source") or item.get("photoType")
        if source:
            entry["source"] = str(source)
            entry["photoType"] = str(source)
        label = item.get("label")
        if label:
            entry["label"] = str(label)
        uploaded_by = item.get("uploaded_by") or item.get("uploadedBy")
        if uploaded_by:
            entry["uploaded_by"] = str(uploaded_by)
        if "active" in item:
            entry["active"] = bool(item.get("active"))
        thumbnail_key = (item.get("thumbnail_key") or item.get("thumbnailKey") or "").strip()
        if thumbnail_key:
            entry["thumbnail_key"] = thumbnail_key
        filename = item.get("filename") or item.get("original_filename") or item.get("originalFilename") or item.get("stored_filename")
        if filename:
            entry["filename"] = str(filename)
        original_filename = item.get("original_filename") or item.get("originalFilename")
        if original_filename:
            entry["original_filename"] = str(original_filename)
        stored_filename = item.get("stored_filename") or item.get("storedFilename")
        if stored_filename:
            entry["stored_filename"] = str(stored_filename)
        content_type = item.get("content_type") or item.get("contentType") or item.get("mime_type") or item.get("mimeType")
        if content_type:
            entry["content_type"] = str(content_type)
        size_bytes = item.get("size_bytes") or item.get("sizeBytes")
        if size_bytes:
            try:
                entry["size_bytes"] = int(size_bytes)
            except (TypeError, ValueError):
                pass
        uploaded_at = item.get("uploaded_at") or item.get("uploadedAt")
        if uploaded_at:
            entry["uploaded_at"] = str(uploaded_at)
        if include_urls:
            try:
                display_url = _photo_storage().public_url(object_key)
            except ReceivingPhotoStorageError:
                display_url = ""
            if display_url:
                entry["url"] = display_url
                entry["public_url"] = display_url
        manifest.append(entry)
    manifest.sort(key=lambda photo: photo.get("sort_order") or 0)
    return manifest


# ── Development tools ─────────────────────────────────────────────────────────

def _is_development_mode():
    return C.FLASK_ENV == "development" or C.FLASK_DEBUG


def _demo_date(days_ago):
    value = datetime.now(timezone.utc) - timedelta(days=days_ago)
    return value.isoformat(timespec="seconds").replace("+00:00", "Z")


def _demo_identifier(code_type, index):
    code_type = code_type or "Text"
    if code_type in {"UPC-12", "GTIN-12"}:
        return f"036800{index % 1000000:06d}"[-12:]
    if code_type == "GTIN-14":
        return f"00036800{index % 1000000:06d}"[-14:]
    if code_type == "GTIN-13":
        return f"0036800{index % 1000000:06d}"[-13:]
    if code_type == "GTIN-8":
        return f"{index % 100000000:08d}"
    if code_type == "Numeric":
        return f"{index:06d}"
    return f"DEMO-{index:04d}"


def _item_client_id(record):
    client_ids = record.get("fields", {}).get(C.F_ITEM_CLIENT, []) or []
    return client_ids[0] if client_ids else ""


def _item_receipt_ids(record):
    receipts = record.get("fields", {}).get(C.F_ITEM_RECEIPTS, [])
    return receipts if isinstance(receipts, list) else []


def _issue_item_ids(record):
    item_ids = record.get("fields", {}).get(C.F_ISSUE_ITEM, [])
    return item_ids if isinstance(item_ids, list) else []


def _demo_item_payload(record, client, queue_id, index):
    client = client or {}
    valid_identifier = _demo_identifier(client.get("codeType"), index)
    base = {
        C.F_ITEM_IDENTIFIER: valid_identifier,
        C.F_ITEM_PRODUCT: record.get("fields", {}).get(C.F_ITEM_PRODUCT) or f"Demo Product {index}",
        C.F_ITEM_BRAND: record.get("fields", {}).get(C.F_ITEM_BRAND) or "Demo Brand",
        C.F_ITEM_ARTWORK_RECEIVED: True,
    }
    if queue_id == "waiting_merchandise":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    elif queue_id == "merchandise_issues":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    elif queue_id == "missing_data":
        base.update({
            C.F_ITEM_PRODUCT: "",
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    elif queue_id == "missing_artwork":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: False,
        })
    elif queue_id == "ready_for_photo":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    elif queue_id == "in_creative_force":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    elif queue_id == "completed":
        base.update({
            C.F_ITEM_ARTWORK_RECEIVED: True,
        })
    return base


def _queue_assignment(records, issues_by_item):
    shuffled = list(records)
    random.shuffle(shuffled)
    used = set()
    assignments = {}

    def take(queue_id, predicate=lambda record: True):
        for record in shuffled:
            if record["id"] not in used and predicate(record):
                used.add(record["id"])
                assignments[queue_id] = record
                return record
        return None

    take("waiting_merchandise", lambda record: not _item_receipt_ids(record))
    take("merchandise_issues", lambda record: bool(issues_by_item.get(record["id"])))
    take("missing_data")
    take("missing_artwork")
    take("ready_for_photo")
    take("in_creative_force")
    take("completed")
    return assignments


@api.post("/dev/randomize-demo-data")
def randomize_demo_data():
    if not _is_development_mode():
        return err("Developer tools are only available in development mode.", 404)

    clients_data = _client_records()
    clients_by_id = {record["id"]: _shape_client(record) for record in clients_data}
    items = airtable.list_records(C.PRODUCTS_TABLE, by_field_id=False).get("records", [])
    issues = airtable.list_records(C.ISSUES_TABLE, by_field_id=False).get("records", [])
    issues_by_item = {}
    for issue in issues:
        for item_id in _issue_item_ids(issue):
            issues_by_item.setdefault(item_id, []).append(issue)

    if not items:
        return jsonify({"summary": {"itemsUpdated": 0, "issuesUpdated": 0, "clientsUpdated": 0, "warnings": ["No Products records exist to randomize."]}})

    assignments = _queue_assignment(items, issues_by_item)
    item_queue_by_id = {record["id"]: queue_id for queue_id, record in assignments.items() if record}
    queue_counts = {
        "waiting_merchandise": 0,
        "merchandise_issues": 0,
        "missing_data": 0,
        "missing_artwork": 0,
        "ready_for_photo": 0,
        "in_creative_force": 0,
        "completed": 0,
    }
    queue_cycle = ["ready_for_photo", "waiting_merchandise", "missing_data", "in_creative_force", "completed"]
    warnings = []
    updated_items = 0
    updated_issues = 0
    updated_clients = 0

    if not assignments.get("waiting_merchandise"):
        warnings.append("No unreceived/unlinked Product was available for Waiting for Merchandise without changing relationships.")
    if not assignments.get("merchandise_issues"):
        warnings.append("No existing Issue linked to a Product was available for Merchandise Issues.")

    artwork_client_ids = {
        _item_client_id(assignments["missing_artwork"])
    } if assignments.get("missing_artwork") and _item_client_id(assignments["missing_artwork"]) else set()

    for client_record in clients_data:
        client_id = client_record["id"]
        target_artwork = "Required" if client_id in artwork_client_ids else "Optional"
        fields = {
            C.F_CLIENT_REQUIRED_TO_SHOOT: ["Identifier", "Product Name"],
            C.F_CLIENT_ARTWORK_REQUIREMENT: target_artwork,
            C.F_CLIENT_MERCHANDISE_REQUIRED: True,
        }
        airtable.update_record(C.CLIENTS_TABLE, client_id, fields, by_field_id=False)
        updated_clients += 1

    for index, item in enumerate(items, start=1):
        queue_id = item_queue_by_id.get(item["id"]) or queue_cycle[index % len(queue_cycle)]
        if queue_id == "waiting_merchandise" and _item_receipt_ids(item):
            queue_id = "ready_for_photo"
        if queue_id == "merchandise_issues" and not issues_by_item.get(item["id"]):
            queue_id = "ready_for_photo"
        client = clients_by_id.get(_item_client_id(item), {})
        fields = _demo_item_payload(item, client, queue_id, index)
        airtable.update_record(C.PRODUCTS_TABLE, item["id"], fields, by_field_id=False)
        queue_counts[queue_id] = queue_counts.get(queue_id, 0) + 1
        updated_items += 1

    merchandise_issue_item_id = assignments.get("merchandise_issues", {}).get("id") if assignments.get("merchandise_issues") else ""
    for index, issue in enumerate(issues, start=1):
        item_ids = _issue_item_ids(issue)
        should_block = merchandise_issue_item_id and merchandise_issue_item_id in item_ids
        fields = {
            C.F_ISSUE_TYPE: "Damaged" if should_block else issue.get("fields", {}).get(C.F_ISSUE_TYPE, "Other"),
            C.F_ISSUE_STATUS: "Open" if should_block else "Resolved",
            C.F_ISSUE_PRIORITY: "High" if should_block else issue.get("fields", {}).get(C.F_ISSUE_PRIORITY, "Normal"),
            C.F_ISSUE_OPENED: _demo_date(14 + index),
            C.F_ISSUE_CLOSED: None if should_block else _demo_date(2 + (index % 5)),
        }
        airtable.update_record(C.ISSUES_TABLE, issue["id"], fields, by_field_id=False)
        updated_issues += 1

    return jsonify({
        "summary": {
            "itemsUpdated": updated_items,
            "issuesUpdated": updated_issues,
            "clientsUpdated": updated_clients,
            "queues": queue_counts,
            "warnings": warnings,
        }
    })


@api.post("/dev/clear-core-tables")
def clear_core_tables():
    if not _is_development_mode():
        return err("Developer tools are only available in development mode.", 404)

    record_cache = {}
    photo_keys = set()
    warnings = []
    tables = [
        ("comments", C.COMMENTS_TABLE),
        ("workstreamCards", C.WORKSTREAM_CARDS_TABLE),
        ("thr3dShippingItems", C.THR3D_SHIPPING_ITEMS_TABLE),
        ("activations", C.ACTIVATIONS_TABLE),
        ("issues", C.ISSUES_TABLE),
        ("history", C.HISTORY_TABLE),
        ("imports", C.IMPORTS_TABLE),
        ("merchandise", C.MERCHANDISE_TABLE),
        ("shipments", C.SHIPMENTS_TABLE),
        # Products are test data too: they are re-importable from client source data,
        # so a reset clears them. Clients, Users, and Locations are reference data and
        # are preserved. See docs/DECISIONS.md 2026-08-19.
        ("products", C.PRODUCTS_TABLE),
    ]
    summary = {}
    try:
        for key, table_name in tables:
            record_cache[key] = _list_all_records(table_name)
        for entry in record_cache.get("merchandise", []):
            for photo in _item_photo_metadata_from_entry(entry.get("fields", {}), include_urls=False):
                object_key = photo.get("object_key") or photo.get("objectKey")
                if object_key:
                    photo_keys.add(object_key)
        for shipment in record_cache.get("shipments", []):
            for photo in _shipment_photo_metadata_from_fields(shipment.get("fields", {}), include_urls=False):
                object_key = photo.get("object_key") or photo.get("objectKey")
                if object_key:
                    photo_keys.add(object_key)
        for key, table_name in tables:
            record_ids = [record["id"] for record in record_cache.get(key, [])]
            summary[key] = {
                "table": table_name,
                "deleted": _delete_records_in_batches(table_name, record_ids),
            }
        deleted_photo_keys = []
        failed_photo_keys = []
        if photo_keys:
            try:
                storage = _photo_storage()
                for object_key in sorted(photo_keys):
                    try:
                        storage.delete_photo(object_key)
                        deleted_photo_keys.append(object_key)
                    except (ReceivingPhotoValidationError, ReceivingPhotoConfigError, ReceivingPhotoStorageError):
                        failed_photo_keys.append(object_key)
            except (ReceivingPhotoValidationError, ReceivingPhotoConfigError, ReceivingPhotoStorageError) as error:
                failed_photo_keys = sorted(photo_keys)
                warnings.append(f"Uploaded photos were not deleted: {error}")
        summary["uploadedPhotos"] = {
            "table": "R2 uploaded photos",
            "deleted": len(deleted_photo_keys),
            "failed": len(failed_photo_keys),
        }
    except requests.HTTPError as error:
        return airtable_err(error)

    return jsonify({"summary": summary, "warnings": warnings})


# ── Settings ──────────────────────────────────────────────────────────────────

@api.get("/settings")
def settings():
    return jsonify({
        "settings": {
            "airtableConfigured": C.airtable_ready(),
            "environment": C.FLASK_ENV,
            "development": _is_development_mode(),
            "base": C.AIRTABLE_BASE_ID,
            "tables": {
                "clients": C.CLIENTS_TABLE,
                "products": C.PRODUCTS_TABLE,
                "shipments": C.SHIPMENTS_TABLE,
                "merchandise": C.MERCHANDISE_TABLE,
                "items":   C.ITEMS_TABLE,
                "skus":    C.SKUS_TABLE,
                "receipts": C.RECEIPTS_TABLE,
                "receiptEntries": C.RECEIPT_ENTRIES_TABLE,
                "locations": C.LOCATIONS_TABLE,
                "users": C.USERS_TABLE,
                "issues": C.ISSUES_TABLE,
                "history": C.HISTORY_TABLE,
                "imports": C.IMPORTS_TABLE,
                "comments": C.COMMENTS_TABLE,
                "workstreamCards": C.WORKSTREAM_CARDS_TABLE,
                "thr3dShippingItems": C.THR3D_SHIPPING_ITEMS_TABLE,
                "activations": C.ACTIVATIONS_TABLE,
            },
        }
    })
